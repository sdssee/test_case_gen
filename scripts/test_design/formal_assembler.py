# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import math
import os
import re
import shutil
import unicodedata
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment

from .excel_utils import clear_data_rows, header_map, resize_workbook_structures, write_mapped_row
from .io_utils import atomic_save_workbook, temporary_sibling
from .session_runtime import artifact_paths, load_cases, load_facts, load_plan, semantic_source_digests


SHEETS = [
    "测试设计总览", "需求用户故事拆解", "测试场景矩阵", "功能测试用例",
    "性能测试设计", "风险与待确认问题", "自动化建议", "页面元素覆盖清单",
]
FUNCTION_SHEET = "功能测试用例"
MULTILINE_HEADERS = {
    "测试范围", "不测范围", "主要风险", "准入条件", "准出条件", "待确认问题",
    "用户故事/需求描述", "业务价值", "验收标准", "业务规则", "前置条件", "后置影响", "依赖系统",
    "输入数据/状态条件", "观察点", "测试数据", "操作步骤", "预期结果", "备注", "描述", "影响范围",
    "建议处理方式", "业务链路", "监控指标", "通过标准", "造数策略", "风险说明", "依赖数据",
    "稳定性风险", "页面 URL/菜单路径", "交互方式", "适用DFX场景", "前置状态/权限", "预期行为", "业务依据/规则来源",
    "测试步骤描述", "测试步骤预期结果", "测试用例说明",
}
INTERNAL_EXPORT_PATTERN = re.compile(r"(?:\b(?:UID|UUID|DOM|ARIA|XPATH)\b|\b(?:EL|TX|FN|PAGE|EVT)-[A-Z0-9_-]+\b)", re.IGNORECASE)
EXCEL_ERRORS = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA"}
FORMAL_WIDTHS = {
    "测试设计总览": {
        "测试范围": 30, "不测范围": 26, "主要风险": 28,
        "准入条件": 32, "准出条件": 32, "待确认问题": 28,
    },
    "需求用户故事拆解": {
        "用户故事/需求描述": 34, "业务价值": 28, "验收标准": 34, "业务规则": 32,
        "前置条件": 30, "后置影响": 30, "依赖系统": 28, "待确认问题": 26,
    },
    "测试场景矩阵": {"测试对象/页面元素": 28, "输入数据/状态条件": 30, "观察点": 42, "备注": 28},
    "功能测试用例": {"用例标题": 34, "前置条件": 30, "测试数据": 28, "操作步骤": 44, "预期结果": 44, "备注": 26},
    "页面元素覆盖清单": {
        "页面 URL/菜单路径": 26, "交互方式": 40, "适用DFX场景": 28,
        "前置状态/权限": 30, "预期行为": 40, "业务依据/规则来源": 28,
        "覆盖用例 ID": 26, "待确认问题/备注": 28,
    },
    "性能测试设计": {
        "业务链路": 26, "DFX场景": 28, "监控指标": 28, "通过标准": 30,
        "造数策略": 26, "风险说明": 30, "是否纳入本轮测试": 18,
    },
    "风险与待确认问题": {"描述": 34, "影响范围": 28, "建议处理方式": 32},
}
IMPORT_WIDTHS = {
    "一级模块名称": 16, "二级模块名称": 16, "三级模块名称": 16,
    "四级模块名称": 16, "五级模块名称": 16, "测试用例名称": 34,
    "测试步骤描述": 44, "测试步骤预期结果": 44, "测试用例说明": 24,
    "测试用例序号": 10, "测试类型": 14, "测试用例级别": 14, "执行方式": 12,
    "前置条件": 30, "标签": 24, "备注": 24,
}


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    return str(value)


def _compact_list(values: Iterable[Any], limit: int = 3) -> str:
    unique = list(dict.fromkeys(_text(value).strip() for value in values if _text(value).strip()))
    if len(unique) <= limit:
        return "；".join(unique)
    return "；".join(unique[:limit]) + f"；另有{len(unique) - limit}项"


def _complete_list(values: Iterable[Any]) -> str:
    return "；".join(dict.fromkeys(_text(value).strip() for value in values if _text(value).strip()))


def _humanize_branch(value: Any, descriptions: dict[str, Any] | None = None) -> str:
    text = _text(value).strip()
    if descriptions and _text(descriptions.get(text)).strip():
        return _text(descriptions[text]).strip()
    aliases = {
        "valid": "有效值", "empty": "空值", "invalid": "无效值", "invalid_format": "无效格式",
        "boundary_min": "下边界", "boundary_max": "上边界", "duplicate": "重复值",
        "valid_domain": "有效域名", "valid_hostname": "有效主机名", "valid_ipv4": "有效IPv4地址",
        "valid_ipv6": "有效IPv6地址", "valid_ip": "有效IP地址", "valid_email": "有效邮箱地址",
        "valid_url": "有效URL", "valid_date": "有效日期", "valid_datetime": "有效日期时间",
        "valid_number": "有效数字", "valid_integer": "有效整数", "valid_decimal": "有效小数",
        "valid_port": "有效端口", "valid_phone": "有效电话号码",
    }
    if text in aliases:
        return aliases[text]
    if text.startswith("valid_"):
        return "有效" + text[len("valid_"):].replace("_", "")
    return text.replace("_", " ")


def _element_interaction_summary(element: dict[str, Any], branch_values: list[str]) -> str:
    explicit = _text(element.get("interaction")).strip()
    if explicit:
        return explicit
    element_type = _text(element.get("type") or element.get("element_type")).strip().lower()
    name = _text(element.get("name") or "当前控件").strip()
    descriptions = element.get("valid_input_class_descriptions") if isinstance(element.get("valid_input_class_descriptions"), dict) else {}
    branches = _complete_list(_humanize_branch(value, descriptions) for value in branch_values)
    if element_type == "trigger":
        return f"点击{name}并观察执行结果"
    if element_type == "select":
        return f"逐项选择并观察页面变化：{branches}" if branches else f"逐项选择{name}并观察页面变化"
    if element_type == "input":
        return f"逐类输入并触发功能：{branches}" if branches else f"在{name}中输入并触发功能"
    if element_type == "toggle":
        return f"逐项切换{name}并观察页面变化"
    return f"操作{name}并观察页面变化"


def _numbered(values: Any) -> str:
    if not isinstance(values, list):
        values = [line for line in str(values or "").splitlines() if line.strip()]
    return "\n".join(f"{index}. {str(value).strip()}" for index, value in enumerate(values, 1))


def _paired_columns(steps: Any) -> tuple[str, str]:
    if not isinstance(steps, list):
        return "", ""
    actions = [str(step.get("action", "")).strip() for step in steps if isinstance(step, dict)]
    expected = [str(step.get("expected", "")).strip() for step in steps if isinstance(step, dict)]
    return _numbered(actions), _numbered(expected)


def _display_width(value: Any) -> int:
    return sum(2 if unicodedata.east_asian_width(character) in {"W", "F"} else 1 for character in str(value or ""))


def _visual_line_count(value: Any, characters_per_line: int = 36) -> int:
    text = str(value or "")
    if not text:
        return 1
    return sum(max(1, math.ceil(_display_width(line) / characters_per_line)) for line in text.splitlines())


def _mapped_row_height(worksheet, headers: dict[str, int], row: dict[str, Any], multiline_headers: Iterable[str]) -> float:
    lines = 1
    for header in multiline_headers:
        if header not in headers:
            continue
        column = headers[header]
        dimension = worksheet.column_dimensions.get(get_column_letter(column))
        width = (dimension.width if dimension else worksheet.sheet_format.defaultColWidth) or 10
        lines = max(lines, _visual_line_count(row.get(header, ""), max(4, int(width))))
    return float(max(36, min(409, 8 + lines * 16)))


def _format_data_row(worksheet, headers: dict[str, int], row_index: int, row: dict[str, Any]) -> None:
    wrapped_headers: set[str] = set()
    for header, value in row.items():
        if header not in headers:
            continue
        column = headers[header]
        dimension = worksheet.column_dimensions.get(get_column_letter(column))
        width = (dimension.width if dimension else worksheet.sheet_format.defaultColWidth) or 10
        if header in MULTILINE_HEADERS or "\n" in str(value or "") or _display_width(value) > width:
            worksheet.cell(row_index, column).alignment = Alignment(vertical="top", wrap_text=True)
            wrapped_headers.add(header)
    worksheet.row_dimensions[row_index].height = _mapped_row_height(worksheet, headers, row, wrapped_headers)


def _assert_rows_fit(worksheet, row_numbers: Iterable[int], label: str) -> None:
    headers = header_map(worksheet)
    reverse_headers = {column: header for header, column in headers.items()}
    for row_index in row_numbers:
        values = {reverse_headers[column]: worksheet.cell(row_index, column).value for column in reverse_headers}
        wrapped = {
            header for header, column in headers.items()
            if worksheet.cell(row_index, column).alignment.wrap_text
        }
        required = _mapped_row_height(worksheet, headers, values, wrapped)
        actual = worksheet.row_dimensions[row_index].height or worksheet.sheet_format.defaultRowHeight or 15
        if actual + 0.1 < required:
            raise ValueError(f"{label} row {row_index} may clip wrapped text")


def _assert_visible_content_clean(workbook, label: str) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and INTERNAL_EXPORT_PATTERN.search(value):
                    raise ValueError(f"{label} exposes an internal identifier at {worksheet.title}!{cell.coordinate}")
                if cell.data_type == "e" or value in EXCEL_ERRORS:
                    raise ValueError(f"{label} contains an Excel error at {worksheet.title}!{cell.coordinate}")


def _workbook_structure(workbook) -> dict[str, Any]:
    """Capture template-owned structure while excluding expandable data ranges."""
    result: dict[str, Any] = {"sheetnames": list(workbook.sheetnames), "sheets": {}}
    for worksheet in workbook.worksheets:
        validations = sorted((
            validation.type, validation.operator, str(validation.formula1 or ""), str(validation.formula2 or ""),
            bool(validation.allow_blank), str(validation.errorTitle or ""), str(validation.error or ""),
            str(validation.promptTitle or ""), str(validation.prompt or ""),
        ) for validation in worksheet.data_validations.dataValidation)
        tables = sorted((
            table.name, table.displayName, str(getattr(table.tableStyleInfo, "name", "") or ""),
            tuple(column.name for column in table.tableColumns),
        ) for table in worksheet.tables.values())
        result["sheets"][worksheet.title] = {
            "state": worksheet.sheet_state,
            "headers": tuple(cell.value for cell in worksheet[1]),
            "widths": tuple((key, value.width, value.hidden) for key, value in sorted(worksheet.column_dimensions.items())),
            "header_hidden": bool(worksheet.row_dimensions[1].hidden),
            "freeze": str(worksheet.freeze_panes or ""),
            "merged": tuple(sorted(str(item) for item in worksheet.merged_cells.ranges)),
            "validations": validations,
            "tables": tables,
            "has_filter": bool(worksheet.auto_filter.ref),
            "print_area": str(worksheet.print_area or ""),
            "print_titles": str(worksheet.print_titles or ""),
            "orientation": str(worksheet.page_setup.orientation or ""),
        }
    return result


def _assert_structure_preserved(before: dict[str, Any], workbook, label: str) -> None:
    after = _workbook_structure(workbook)
    if before != after:
        changed = [name for name in before.get("sheets", {}) if before["sheets"].get(name) != after.get("sheets", {}).get(name)]
        raise ValueError(f"{label} template structure changed unexpectedly: {changed or 'sheet order'}")


def _apply_width_profile(workbook, profiles: dict[str, dict[str, float]]) -> None:
    for sheet_name, widths in profiles.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        headers = header_map(worksheet)
        for header, width in widths.items():
            if header in headers:
                worksheet.column_dimensions[get_column_letter(headers[header])].width = width


def _automation_suitable(case: dict[str, Any], plan_functions: dict[str, dict[str, Any]]) -> str:
    explicit = case.get("automation")
    if explicit is True:
        return "是"
    if explicit is False:
        return "否"
    profile = plan_functions.get(str(case.get("function_ref", "")), {}).get("automation_profile", {})
    level = str(profile.get("level", "")).strip().lower() if isinstance(profile, dict) else ""
    if level in {"不适用", "none", "n/a", "na"}:
        return "否"
    risk = str(profile.get("stability_risk", "")).strip().lower()
    substantive_risk = risk and not any(marker in risk for marker in ("无风险", "无已知", "none", "n/a", "不适用"))
    if substantive_risk:
        return "有条件"
    return "是" if str(case.get("automation_value", "")).strip() and str(case.get("automation_priority", "")).strip() else "待评估"


def _sample_formulas(worksheet, row: int = 2) -> dict[int, str]:
    return {
        cell.column: str(cell.value)
        for cell in worksheet[row]
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }


def _apply_sample_formulas(worksheet, formulas: dict[int, str], target_row: int) -> None:
    for column, formula in formulas.items():
        origin = worksheet.cell(2, column).coordinate
        destination = worksheet.cell(target_row, column).coordinate
        worksheet.cell(target_row, column).value = Translator(formula, origin=origin).translate_formula(destination)


def _function_names(facts: dict[str, Any], plan: dict[str, Any]) -> dict[str, str]:
    result = {
        str(row.get("fact_id")): str(row.get("name", ""))
        for row in facts.get("functions", [])
    }
    for row in plan.get("functions", []):
        if row.get("function_ref") and row.get("name"):
            result[str(row["function_ref"])] = str(row["name"])
    return result


def build_sheet_rows(run_dir: Path) -> dict[str, list[dict[str, str]]]:
    facts = load_facts(run_dir)
    plan = load_plan(run_dir)
    case_document = load_cases(run_dir)
    scope = facts.get("scope", {})
    functions = _function_names(facts, plan)
    plan_functions = {str(row.get("function_ref", "")): row for row in plan.get("functions", [])}
    function_facts = {str(row.get("fact_id", "")): row for row in facts.get("functions", [])}
    function_order = [str(row.get("fact_id", "")) for row in facts.get("functions", [])]
    supplied_requirements = list(scope.get("requirements", []))
    plan_requirements = {str(row.get("function_ref", "")): _text(row.get("requirement_id")) for row in plan.get("functions", [])}
    requirement_by_function = {
        function_ref: _text(
            function_facts.get(function_ref, {}).get("requirement_id")
            or plan_requirements.get(function_ref)
            or (supplied_requirements[index - 1].get("requirement_id") if index <= len(supplied_requirements) else "")
            or f"REQ-{index:03d}"
        )
        for index, function_ref in enumerate(function_order, 1)
    }
    pages = {str(row.get("fact_id", "")): row for row in facts.get("pages", [])}
    fact_names = {
        str(row.get("fact_id")): str(row.get("name") or row.get("transaction_type") or "页面实探事实")
        for collection in ("pages", "functions", "elements")
        for row in facts.get(collection, [])
    }
    written_cases = {str(case.get("case_id", "")): case for case in case_document.get("cases", [])}
    transaction_map = {str(row.get("fact_id", "")): row for row in facts.get("transactions", [])}
    primary_elements_by_case: dict[str, list[str]] = {}
    for assignment in plan.get("check_assignments", []):
        if assignment.get("disposition") != "case":
            continue
        transaction = transaction_map.get(str(assignment.get("transaction_ref", "")), {})
        try:
            check = transaction.get("checks", [])[int(assignment.get("check_index", 0)) - 1]
        except (IndexError, TypeError, ValueError):
            continue
        element_ref = str(check.get("element_ref", "")).strip()
        if element_ref:
            primary_elements_by_case.setdefault(str(assignment.get("case_id", "")), []).append(element_ref)
    open_items = facts.get("open_items", [])
    pending = [
        row for row in open_items
        if row.get("category") in {"external_question", "blocked_condition"}
        and row.get("status") not in {"resolved", "accepted", "closed"}
    ]
    raw_risks = [row for row in open_items if row.get("category") == "observed_risk"] + plan.get("risks", [])
    risks: list[dict[str, Any]] = []
    seen_risks: set[tuple[str, str]] = set()
    for row in raw_risks:
        key = (
            re.sub(r"\s+", "", _text(row.get("description") or row.get("reason"))),
            _text(row.get("function_ref") or row.get("impact")),
        )
        if key not in seen_risks:
            risks.append(row)
            seen_risks.add(key)
    risk_level_order = {"无": 0, "低": 1, "中": 2, "高": 3, "严重": 4}
    global_risk_rows = [row for row in risks if not str(row.get("function_ref", "")).strip()]
    risks_by_function: dict[str, list[dict[str, Any]]] = {}
    for row in risks:
        refs = [str(value) for value in row.get("affected_function_refs", []) if str(value).strip()]
        if str(row.get("function_ref", "")).strip():
            refs.append(str(row.get("function_ref", "")).strip())
        for ref in dict.fromkeys(refs):
            risks_by_function.setdefault(ref, []).append(row)

    def risk_rows_for(function_ref: str) -> list[dict[str, Any]]:
        return list({id(row): row for row in global_risk_rows + risks_by_function.get(function_ref, [])}.values())

    def risk_level_for(function_ref: str) -> str:
        linked = risk_rows_for(function_ref)
        return max((_text(row.get("level") or "中") for row in linked), key=lambda value: risk_level_order.get(value, 2)) if linked else "未单独识别"
    function_contexts = [
        row.get("design_context", {}) for row in plan.get("functions", [])
        if isinstance(row.get("design_context"), dict)
    ]
    dependency_summary = list(dict.fromkeys(
        _text(value) for context in function_contexts for value in context.get("dependencies", []) if _text(value)
    ))
    planned_case_count = sum(len(function.get("cases", [])) for function in plan.get("functions", []))
    written_case_count = len(case_document.get("cases", []))
    exit_conditions = [
        "既定页面实探分支均已有用例或明确专项处置",
        f"计划与已编写用例数量一致（{planned_case_count}条）" if planned_case_count == written_case_count else "计划与已编写用例数量需保持一致",
        "无未解决的阻塞问题" if not pending else "阻塞问题已确认处理方式",
        "双Excel内容一致并通过技术校验",
    ]
    rows: dict[str, list[dict[str, str]]] = {name: [] for name in SHEETS}
    rows["测试设计总览"].append({
        "项目/模块": _text(scope.get("module_path")),
        "需求名称": _text(scope.get("requirement_name") or scope.get("module_path")),
        "版本/迭代": _text(scope.get("version") or "未提供"),
        "测试负责人": _text(scope.get("owner") or "未提供"),
        "需求来源": _text(scope.get("source") or "页面实探"),
        "测试范围": _text(scope.get("test_scope") or "；".join(functions.get(ref, ref) for ref in function_order)),
        "不测范围": _text(scope.get("out_of_scope") or "未提供明确排除项"),
        "测试类型": "功能测试；适用的DFX专项设计",
        "测试环境": _text(scope.get("environment") or "未提供"),
        "主要风险": _compact_list((row.get("description") for row in risks), limit=4) or _text(plan.get("risk_not_applicable_reason") or "未发现需单独登记的风险"),
        "准入条件": "；".join(dependency_summary) or "按功能设计上下文准备访问权限与受控数据",
        "准出条件": "；".join(exit_conditions),
        "待确认问题": _compact_list((row.get("description") or row.get("reason") for row in pending), limit=4) or "无待确认项",
    })
    supplied_by_id = {
        _text(row.get("requirement_id")): row for row in supplied_requirements if _text(row.get("requirement_id"))
    }
    elements_by_function: dict[str, list[dict[str, Any]]] = {}
    for element in facts.get("elements", []):
        elements_by_function.setdefault(str(element.get("function_ref", "")), []).append(element)
    derived_requirements: list[dict[str, Any]] = []
    used_supplied: set[int] = set()
    for function_index, function_ref in enumerate(function_order):
        function = function_facts.get(function_ref, {})
        context = plan_functions.get(function_ref, {}).get("design_context", {})
        if not isinstance(context, dict):
            context = {}
        requirement_id = requirement_by_function[function_ref]
        supplied = supplied_by_id.get(requirement_id, {})
        if not supplied and function_index < len(supplied_requirements):
            supplied = supplied_requirements[function_index]
        if supplied:
            used_supplied.add(id(supplied))
        pages_for_function = [
            pages.get(str(element.get("page_ref", "")), {}) for element in elements_by_function.get(function_ref, [])
        ]
        menu_path = next((page.get("menu_path") for page in pages_for_function if page.get("menu_path")), [])
        related_open = [
            _text(item.get("description") or item.get("reason")) for item in open_items
            if function_ref in [str(ref) for ref in item.get("affected_function_refs", [])]
        ]
        derived_requirements.append({
            "requirement_id": requirement_id,
            "description": supplied.get("description") or context.get("user_goal") or function.get("description"),
            "role": supplied.get("role") or context.get("role") or function.get("role"),
            "business_value": supplied.get("business_value") or context.get("business_value"),
            "acceptance_criteria": supplied.get("acceptance_criteria") or context.get("acceptance_criteria"),
            "business_rules": supplied.get("business_rules") or context.get("business_rules"),
            "preconditions": supplied.get("preconditions") or ("可进入" + "-".join(str(value) for value in menu_path) if menu_path else context.get("dependencies")),
            "postconditions": supplied.get("postconditions") or context.get("postcondition"),
            "dependencies": supplied.get("dependencies") or context.get("dependencies"),
            "unresolved": supplied.get("unresolved") or "；".join(value for value in related_open if value),
        })
    derived_requirements.extend(row for row in supplied_requirements if id(row) not in used_supplied)
    for index, requirement in enumerate(derived_requirements, 1):
        requirement_id = _text(requirement.get("requirement_id") or f"REQ-{index:03d}")
        rows["需求用户故事拆解"].append({
            "Story ID/需求 ID": requirement_id,
            "用户故事/需求描述": _text(requirement.get("description")),
            "角色": _text(requirement.get("role")),
            "业务价值": _text(requirement.get("business_value")),
            "验收标准": _text(requirement.get("acceptance_criteria")),
            "业务规则": _text(requirement.get("business_rules")),
            "前置条件": _text(requirement.get("preconditions")),
            "后置影响": _text(requirement.get("postconditions")),
            "依赖系统": _text(requirement.get("dependencies")),
            "待确认问题": _text(requirement.get("unresolved") or "无待确认项"),
        })
    for function in plan.get("functions", []):
        function_ref = str(function.get("function_ref", ""))
        requirement_id = _text(function.get("requirement_id") or requirement_by_function.get(function_ref))
        for case in function.get("cases", []):
            written = written_cases.get(str(case.get("case_id", "")), {})
            expected_points = [
                str(step.get("expected", "")).strip()
                for step in written.get("steps", [])[1:]
                if str(step.get("expected", "")).strip()
            ]
            observation = _text(case.get("verification_focus"))
            if expected_points:
                observation = "；".join(filter(None, [observation, "实际断言：" + "；".join(expected_points)]))
            rows["测试场景矩阵"].append({
                "场景 ID": _text(case.get("case_id")),
                "Story ID/需求 ID": _text(case.get("requirement_id") or requirement_id),
                "功能点": _text(case.get("test_point") or function.get("name")),
                "测试维度": _text(case.get("strategy")),
                "DFX维度": _text(case.get("dfx_dimension")),
                "DFX场景": _text(case.get("dfx_scenario")),
                "测试对象/页面元素": "；".join(dict.fromkeys(
                    fact_names.get(str(ref), "") for ref in primary_elements_by_case.get(str(case.get("case_id", "")), []) if fact_names.get(str(ref), "")
                )),
                "输入数据/状态条件": _text(written.get("test_data") or case.get("test_data")),
                "观察点": observation or _text(case.get("observation")),
                "风险等级": _text(written.get("risk_level") or case.get("risk_level") or risk_level_for(function_ref)),
                "优先级": _text(written.get("priority") or case.get("priority") or "P1"),
                "是否生成用例": "是",
                "备注": _text(case.get("notes")),
            })
    for case in case_document.get("cases", []):
        function_id = str(case.get("function_ref", ""))
        action_text, expected_text = _paired_columns(case.get("steps"))
        case_risks = case.get("risks", [])
        if not isinstance(case_risks, list):
            case_risks = [case_risks] if _text(case_risks) else []
        linked_risks = list(dict.fromkeys(
            [_text(row.get("risk_id") or row.get("fact_id")) for row in risk_rows_for(function_id) if _text(row.get("risk_id") or row.get("fact_id"))]
            + [_text(value) for value in case_risks if _text(value)]
        ))
        rows[FUNCTION_SHEET].append({
            "用例 ID": _text(case.get("case_id")),
            "Story ID/需求 ID": _text(case.get("requirement_id") or requirement_by_function.get(function_id)),
            "模块": _text(scope.get("module_path")),
            "功能点": _text(case.get("test_point") or functions.get(function_id, function_id)),
            "用例标题": _text(case.get("title")),
            "优先级": _text(case.get("priority") or "P1"),
            "测试类型": _text(case.get("test_type") or "功能测试"),
            "DFX维度": _text(case.get("dfx_dimension")),
            "DFX场景": _text(case.get("dfx_scenario")),
            "前置条件": _numbered(case.get("preconditions")),
            "测试数据": _text(case.get("test_data")),
            "操作步骤": action_text,
            "预期结果": expected_text,
            "实际结果": "",
            "执行状态": "未执行",
            "是否适合自动化": _automation_suitable(case, plan_functions),
            "关联风险": "；".join(linked_risks),
            "备注": _text(case.get("notes") or ("主验证：" + _text(case.get("verification_focus")) if case.get("verification_focus") else "")),
        })
    for item in plan.get("performance_scenarios", []):
        mapping = {
            "性能场景 ID": "scenario_id", "Story ID/需求 ID": "requirement_id", "业务链路": "flow",
            "性能测试类型": "test_type", "DFX维度": "dfx_dimension", "DFX场景": "dfx_scenario",
            "目标用户量/并发数": "concurrency", "TPS/QPS 目标": "throughput", "响应时间目标": "response_time",
            "数据量级": "data_scale", "测试时长": "duration", "监控指标": "metrics", "通过标准": "pass_criteria",
            "造数策略": "data_strategy", "风险说明": "risk", "是否纳入本轮测试": "included",
        }
        rows["性能测试设计"].append({header: _text(item.get(key)) for header, key in mapping.items()})
    if not rows["性能测试设计"]:
        reason = _text(plan.get("performance_not_applicable_reason"))
        if not reason:
            raise ValueError("case-plan.json must state why performance design is not applicable")
        rows["性能测试设计"].append({
            "性能场景 ID": "PERF-N/A", "Story ID/需求 ID": "不适用", "业务链路": _text(scope.get("module_path")),
            "性能测试类型": "不适用", "DFX维度": "DFP性能", "DFX场景": "不适用判定",
            "目标用户量/并发数": "不适用", "TPS/QPS 目标": "不适用", "响应时间目标": "不适用",
            "数据量级": "不适用", "测试时长": "不适用", "监控指标": "不适用", "通过标准": "不适用",
            "造数策略": "不适用", "风险说明": reason, "是否纳入本轮测试": "否",
        })
    for index, risk in enumerate(risks + pending, 1):
        affected_names = [functions.get(str(ref), str(ref)) for ref in risk.get("affected_function_refs", []) if str(ref)]
        is_pending = risk in pending
        rows["风险与待确认问题"].append({
            "编号": _text(risk.get("risk_id") or f"RISK-{index:03d}"),
            "类型": _text(risk.get("type") or ("待确认" if is_pending else "风险")),
            "关联DFX维度": _text(risk.get("dfx_dimension") or ("DFT功能" if is_pending else "DFR可靠")),
            "关联DFX场景": _text(risk.get("dfx_scenario") or ("外部业务语义或阻塞待确认" if is_pending else "实探发现风险")),
            "描述": _text(risk.get("description") or risk.get("reason")),
            "影响范围": _text(risk.get("impact") or "；".join(affected_names) or scope.get("module_path")),
            "风险等级": _text(risk.get("level") or "中"),
            "建议处理方式": _text(risk.get("recommendation") or "根据实探事实进行针对性确认或处理"),
            "负责人": _text(risk.get("owner")),
            "状态": _text(risk.get("status") or "待确认"),
        })
    if not rows["风险与待确认问题"]:
        reason = _text(plan.get("risk_not_applicable_reason"))
        if not reason:
            raise ValueError("case-plan.json must state why risk design is not applicable")
        rows["风险与待确认问题"].append({
            "编号": "RISK-N/A", "类型": "不适用", "关联DFX维度": "不适用", "关联DFX场景": "不适用",
            "描述": reason, "影响范围": _text(scope.get("module_path")),
            "风险等级": "无", "建议处理方式": "无需额外处理", "负责人": "不适用", "状态": "已确认",
        })
    for case in case_document.get("cases", []):
        suitable = _automation_suitable(case, plan_functions)
        profile = plan_functions.get(str(case.get("function_ref", "")), {}).get("automation_profile", {})
        if not isinstance(profile, dict):
            profile = {}
        override = case.get("automation_override") if isinstance(case.get("automation_override"), dict) else {}
        rows["自动化建议"].append({
            "用例 ID/场景 ID": _text(case.get("case_id")),
            "自动化层级": _text(override.get("level") or profile.get("level")),
            "自动化价值": _text(case.get("automation_value")),
            "自动化优先级": _text(case.get("automation_priority")),
            "依赖数据": _text(override.get("dependency") or profile.get("dependency")),
            "Mock 需求": _text(override.get("mock_requirement") or profile.get("mock_requirement") or "无"),
            "稳定性风险": _text(override.get("stability_risk") or profile.get("stability_risk")),
            "建议框架/工具": _text(override.get("recommendation") or profile.get("recommendation")),
            "备注": _text(case.get("automation_reason") or case.get("notes") or f"自动化适用性：{suitable}"),
        })
    case_by_fact: dict[str, list[str]] = {}
    auxiliary_case_by_fact: dict[str, list[str]] = {}
    case_by_id = {str(case.get("case_id", "")): case for case in case_document.get("cases", [])}
    for assignment in plan.get("check_assignments", []):
        if assignment.get("disposition") != "case":
            continue
        transaction = transaction_map.get(str(assignment.get("transaction_ref", "")), {})
        try:
            check = transaction.get("checks", [])[int(assignment.get("check_index", 0)) - 1]
        except (IndexError, TypeError, ValueError):
            continue
        case_id = str(assignment.get("case_id", ""))
        primary_ref = str(check.get("element_ref", "")).strip()
        if primary_ref:
            case_by_fact.setdefault(primary_ref, []).append(case_id)
        for ref in check.get("used_element_refs", []):
            ref = str(ref).strip()
            if ref and ref != primary_ref:
                auxiliary_case_by_fact.setdefault(ref, []).append(case_id)
    non_case_by_element: dict[str, list[str]] = {}
    non_case_status_by_element: dict[str, list[str]] = {}
    transactions = {str(row.get("fact_id", "")): row for row in facts.get("transactions", [])}
    for assignment in plan.get("check_assignments", []):
        if assignment.get("disposition") == "case":
            continue
        transaction = transactions.get(str(assignment.get("transaction_ref", "")), {})
        try:
            check_index = int(assignment.get("check_index", 0))
            if check_index < 1:
                continue
            check = transaction.get("checks", [])[check_index - 1]
        except (IndexError, TypeError, ValueError):
            continue
        reason = str(assignment.get("reason") or assignment.get("disposition") or "").strip()
        ref = str(check.get("element_ref", "")).strip()
        if ref and reason:
            non_case_by_element.setdefault(ref, []).append(reason)
            non_case_status_by_element.setdefault(ref, []).append(str(assignment.get("disposition", "")))
    results_by_element: dict[str, list[str]] = {}
    for transaction in transactions.values():
        for check in transaction.get("checks", []):
            refs = {str(ref) for ref in check.get("used_element_refs", []) if str(ref).strip()}
            if check.get("element_ref"):
                refs.add(str(check.get("element_ref")))
            for ref in refs:
                if _text(check.get("result")):
                    anchor = check.get("result_anchor") if isinstance(check.get("result_anchor"), dict) else {}
                    stable = anchor.get("stable_tokens", anchor.get("tokens"))
                    stable_values = stable if isinstance(stable, list) else [stable] if stable not in (None, "", []) else []
                    summary = "稳定显示：" + "、".join(_text(value) for value in stable_values if _text(value)) if stable_values else _text(check.get("result"))
                    results_by_element.setdefault(ref, []).append(summary)
    blocked_functions = {
        str(ref)
        for item in open_items
        if item.get("category") == "blocked_condition" and item.get("status") not in {"resolved", "accepted", "closed"}
        for ref in item.get("affected_function_refs", [])
    }
    for element in facts.get("elements", []):
        element_id = str(element.get("fact_id"))
        page = pages.get(str(element.get("page_ref", "")), {})
        page_name = str(element.get("page_name") or page.get("name") or "").strip()
        menu_path = element.get("menu_path") or page.get("menu_path") or []
        element_label = "-".join(filter(None, [page_name, str(element.get("name", "")).strip()]))
        covered_case_ids = list(dict.fromkeys(case_by_fact.get(element_id, [])))
        if not element.get("exploration_requirements"):
            covered_case_ids = list(dict.fromkeys(covered_case_ids + auxiliary_case_by_fact.get(element_id, [])))
        non_case_reasons = list(dict.fromkeys(non_case_by_element.get(element_id, [])))
        related_cases = [case_by_id[case_id] for case_id in covered_case_ids if case_id in case_by_id]
        function_ref = str(element.get("function_ref", ""))
        dispositions = set(non_case_status_by_element.get(element_id, []))
        if covered_case_ids:
            coverage_status = "已覆盖"
        elif "not_applicable" in dispositions:
            coverage_status = "不适用"
        elif dispositions & {"performance", "risk"}:
            coverage_status = "已实探-专项处理"
        elif function_ref in blocked_functions:
            coverage_status = "受阻"
        else:
            coverage_status = "未覆盖"
        dfx_dimensions = list(dict.fromkeys(_text(case.get("dfx_dimension")) for case in related_cases if _text(case.get("dfx_dimension"))))
        dfx_scenarios = list(dict.fromkeys(_text(case.get("dfx_scenario")) for case in related_cases if _text(case.get("dfx_scenario"))))
        preconditions = list(dict.fromkeys(
            _text(value) for case in related_cases for value in case.get("preconditions", []) if _text(value)
        ))
        results = list(dict.fromkeys(results_by_element.get(element_id, [])))
        branch_values = [
            _text(requirement.get("value"))
            for requirement in element.get("exploration_requirements", [])
            if _text(requirement.get("value"))
        ]
        auxiliary_ids = list(dict.fromkeys(auxiliary_case_by_fact.get(element_id, [])))
        note = _text(element.get("notes")) or "；".join(non_case_reasons)
        if coverage_status == "未覆盖" and not note:
            note = "已登记元素尚未关联独立测试意图，需在最终Review中局部补充"
        elif coverage_status == "受阻" and not note:
            note = "受页面权限、数据或环境条件阻塞"
        if auxiliary_ids:
            auxiliary_note = "辅助使用用例：" + "；".join(auxiliary_ids)
            note = "；".join(filter(None, [note, auxiliary_note]))
        interaction_summary = _element_interaction_summary(element, branch_values)
        expected_summary = _text(element.get("expected_behavior")) or _complete_list(results)
        if not expected_summary:
            expected_summary = "各主验证分支产生已记录的可观察结果"
        rows["页面元素覆盖清单"].append({
            "元素 ID": element_label or _text(element.get("name")), "Story ID/需求 ID": _text(element.get("requirement_id") or requirement_by_function.get(function_ref)),
            "页面/入口": page_name,
            "页面 URL/菜单路径": "-".join(str(value) for value in menu_path) if isinstance(menu_path, list) else _text(menu_path),
            "元素名称/文案": _text(element.get("name")), "元素类型": _text(element.get("type") or "页面控件"),
            "交互方式": interaction_summary,
            "适用DFX维度": _text(element.get("dfx_dimensions") or "；".join(dfx_dimensions) or "DFT功能"),
            "适用DFX场景": _text(element.get("dfx_scenarios") or "；".join(dfx_scenarios) or "正向流程"),
            "前置状态/权限": _text(element.get("precondition") or "；".join(preconditions) or "具备当前页面访问权限"),
            "预期行为": expected_summary,
            "业务依据/规则来源": _text(element.get("rule_source") or "；".join(plan_functions.get(function_ref, {}).get("design_context", {}).get("basis", [])) or "页面实探"),
            "覆盖用例 ID": "；".join(covered_case_ids),
            "覆盖状态": coverage_status,
            "发现方式": "页面实探",
            "素材来源": "", "待确认问题/备注": note,
        })
    return rows


def assemble_formal_workbook(run_dir: Path, template: Path, output: Path) -> dict[str, int]:
    review_path = artifact_paths(run_dir)["review"]
    if not review_path.is_file():
        raise ValueError("review.json does not exist; run the single review first")
    review = json.loads(review_path.read_text(encoding="utf-8"))
    if review.get("status") not in {"ready", "ready_with_notes"}:
        raise ValueError(f"delivery requires a local repair recorded by review: {review.get('status')}")
    current_sources = semantic_source_digests(run_dir)
    if review.get("sources") != current_sources:
        raise ValueError("review.json is stale; run the single review once for the current artifacts")
    if not template.is_file():
        raise ValueError(f"formal workbook template not found: {template}")
    rows_by_sheet = build_sheet_rows(run_dir)
    if not rows_by_sheet[FUNCTION_SHEET]:
        raise ValueError("function-cases.json contains no cases")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = temporary_sibling(output)
    shutil.copy2(template, temporary)
    workbook = load_workbook(temporary)
    if workbook.sheetnames != SHEETS:
        raise ValueError(f"formal template must contain exactly the 8 standard sheets: {SHEETS}")
    _apply_width_profile(workbook, FORMAL_WIDTHS)
    structure = _workbook_structure(workbook)
    counts: dict[str, int] = {}
    for sheet_name in SHEETS:
        worksheet = workbook[sheet_name]
        headers = header_map(worksheet)
        formulas = _sample_formulas(worksheet)
        clear_data_rows(worksheet)
        if not rows_by_sheet[sheet_name] and worksheet.max_row >= 2:
            worksheet.delete_rows(2, worksheet.max_row - 1)
        for row_index, row in enumerate(rows_by_sheet[sheet_name], 2):
            unknown = set(row) - set(headers)
            if unknown:
                raise ValueError(f"{sheet_name} row uses unknown headers: {sorted(unknown)}")
            write_mapped_row(worksheet, headers, row_index, row)
            _apply_sample_formulas(worksheet, formulas, row_index)
            _format_data_row(worksheet, headers, row_index, row)
        counts[sheet_name] = len(rows_by_sheet[sheet_name])
    resize_workbook_structures(workbook)
    _assert_structure_preserved(structure, workbook, "formal workbook")
    atomic_save_workbook(workbook, temporary)
    os.replace(temporary, output)
    return counts


def generate_import_workbook(run_dir: Path, template: Path, output: Path, module_path: str) -> int:
    """Generate the import workbook directly from function-cases.json."""
    if not template.is_file():
        raise ValueError("import template must exist")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = temporary_sibling(output)
    shutil.copy2(template, temporary)
    cases = load_cases(run_dir).get("cases", [])
    target_book = load_workbook(temporary)
    _apply_width_profile(target_book, {target_book.sheetnames[0]: IMPORT_WIDTHS})
    structure = _workbook_structure(target_book)
    target = target_book[target_book.sheetnames[0]]
    target_headers = header_map(target)
    formulas = _sample_formulas(target)
    clear_data_rows(target)
    modules = [part.strip() for part in re_split_module(module_path)][:5]
    modules += [""] * (5 - len(modules))
    count = 0
    for case in cases:
        case_id = str(case.get("case_id", "")).strip()
        title = str(case.get("title", "")).strip()
        if not str(case_id or "").strip() or not str(title or "").strip():
            raise ValueError("function-cases.json contains an empty case ID or title")
        count += 1
        actions, expected = _paired_columns(case.get("steps"))
        row = {
            "一级模块名称": modules[0], "二级模块名称": modules[1], "三级模块名称": modules[2],
            "四级模块名称": modules[3], "五级模块名称": modules[4], "测试用例序号": count,
            "测试用例名称": title,
            "测试步骤描述": actions,
            "测试步骤预期结果": expected,
            "测试类型": case.get("test_type") or "功能测试",
            "测试用例级别": case.get("priority") or "P1",
            "执行方式": "手工",
            "测试用例说明": title.split("-", 1)[0],
            "前置条件": _numbered(case.get("preconditions")),
            "标签": "；".join(filter(None, [
                str(case.get("dfx_dimension") or ""),
                str(case.get("dfx_scenario") or ""),
            ])),
            "备注": _text(case.get("notes")),
        }
        write_mapped_row(target, target_headers, count + 1, row)
        _apply_sample_formulas(target, formulas, count + 1)
        _format_data_row(target, target_headers, count + 1, row)
    if count == 0:
        raise ValueError("no function cases were available for the import workbook")
    resize_workbook_structures(target_book)
    _assert_structure_preserved(structure, target_book, "import workbook")
    atomic_save_workbook(target_book, temporary)
    os.replace(temporary, output)
    return count


def re_split_module(module_path: str) -> list[str]:
    import re
    return [value for value in re.split(r"\s*(?:>|/|\\|→)\s*", module_path) if value]


def _verify_generated_deliverables(
    run_dir: Path,
    formal_path: Path,
    import_path: Path,
    expected_count: int,
    formal_template_path: Path,
    import_template_path: Path,
) -> None:
    formal = load_workbook(formal_path, data_only=False)
    formal_template = load_workbook(formal_template_path, data_only=False)
    _apply_width_profile(formal_template, FORMAL_WIDTHS)
    _assert_structure_preserved(_workbook_structure(formal_template), formal, "saved formal workbook")
    if formal.sheetnames != SHEETS:
        raise ValueError("generated formal workbook does not retain the exact 8-sheet contract")
    required_by_sheet = {
        "测试设计总览": ("项目/模块", "需求名称", "测试范围", "测试类型", "准入条件", "准出条件"),
        "需求用户故事拆解": ("Story ID/需求 ID", "用户故事/需求描述", "角色", "业务价值", "验收标准", "业务规则", "前置条件", "后置影响", "依赖系统"),
        "测试场景矩阵": ("场景 ID", "Story ID/需求 ID", "功能点", "测试维度", "DFX维度", "DFX场景", "输入数据/状态条件", "观察点", "优先级"),
        "功能测试用例": ("用例 ID", "Story ID/需求 ID", "功能点", "用例标题", "优先级", "测试类型", "DFX维度", "DFX场景", "前置条件", "测试数据", "操作步骤", "预期结果"),
        "性能测试设计": ("性能场景 ID", "业务链路", "性能测试类型", "DFX维度", "DFX场景", "是否纳入本轮测试"),
        "风险与待确认问题": ("编号", "类型", "描述", "影响范围", "风险等级", "建议处理方式", "状态"),
        "自动化建议": ("用例 ID/场景 ID", "自动化层级", "自动化价值", "自动化优先级", "依赖数据", "稳定性风险", "建议框架/工具", "备注"),
        "页面元素覆盖清单": ("元素 ID", "Story ID/需求 ID", "页面/入口", "页面 URL/菜单路径", "元素名称/文案", "元素类型", "交互方式", "适用DFX维度", "适用DFX场景", "前置状态/权限", "预期行为", "业务依据/规则来源", "覆盖状态", "发现方式"),
    }
    for sheet_name, required_headers in required_by_sheet.items():
        worksheet = formal[sheet_name]
        headers = header_map(worksheet)
        data_rows = [
            row for row in range(2, worksheet.max_row + 1)
            if any(worksheet.cell(row, column).value not in (None, "") for column in range(1, worksheet.max_column + 1))
        ]
        if not data_rows:
            raise ValueError(f"generated formal workbook leaves {sheet_name} empty")
        if data_rows != list(range(2, 2 + len(data_rows))):
            raise ValueError(f"generated formal workbook has a blank row in {sheet_name}")
        for header in required_headers:
            if header not in headers:
                raise ValueError(f"generated formal workbook lacks required header {sheet_name}.{header}")
            for row in data_rows:
                if worksheet.cell(row, headers[header]).value in (None, ""):
                    raise ValueError(f"generated formal workbook has empty core field {sheet_name}.{header} at row {row}")
        _assert_rows_fit(worksheet, data_rows, f"generated formal workbook {sheet_name}")
    function_ws = formal[FUNCTION_SHEET]
    function_headers = header_map(function_ws)
    formal_rows = [
        row for row in range(2, function_ws.max_row + 1)
        if any(function_ws.cell(row, column).value not in (None, "") for column in range(1, function_ws.max_column + 1))
    ]
    if formal_rows != list(range(2, 2 + expected_count)):
        raise ValueError("generated formal workbook has a missing, blank, or extra function-case row")
    imported = load_workbook(import_path, data_only=False)
    import_template = load_workbook(import_template_path, data_only=False)
    _apply_width_profile(import_template, {import_template.sheetnames[0]: IMPORT_WIDTHS})
    _assert_structure_preserved(_workbook_structure(import_template), imported, "saved import workbook")
    import_ws = imported[imported.sheetnames[0]]
    import_headers = header_map(import_ws)
    import_rows = [
        row for row in range(2, import_ws.max_row + 1)
        if any(import_ws.cell(row, column).value not in (None, "") for column in range(1, import_ws.max_column + 1))
    ]
    if import_rows != list(range(2, 2 + expected_count)):
        raise ValueError("generated import workbook has a missing, blank, or extra case row")
    _assert_rows_fit(import_ws, import_rows, "generated import workbook")
    for name in SHEETS:
        generated_rows = [
            row for row in range(2, formal[name].max_row + 1)
            if any(formal[name].cell(row, column).value not in (None, "") for column in range(1, formal[name].max_column + 1))
        ]
        formulas = _sample_formulas(formal_template[name])
        for row in generated_rows:
            for column in formulas:
                value = formal[name].cell(row, column).value
                if not isinstance(value, str) or not value.startswith("=") or "#REF!" in value:
                    raise ValueError(f"generated formal workbook lost a template formula at {name}!{row},{column}")
    import_formulas = _sample_formulas(import_template[import_template.sheetnames[0]])
    for row in import_rows:
        for column in import_formulas:
            value = import_ws.cell(row, column).value
            if not isinstance(value, str) or not value.startswith("=") or "#REF!" in value:
                raise ValueError(f"generated import workbook lost a template formula at row {row}, column {column}")
    source_cases = load_cases(run_dir).get("cases", [])
    if len(source_cases) != expected_count:
        raise ValueError("generated workbook count differs from function-cases.json")
    mappings = {"用例标题": "测试用例名称", "操作步骤": "测试步骤描述", "预期结果": "测试步骤预期结果"}
    for formal_row, import_row in zip(formal_rows, import_rows):
        for formal_header, import_header in mappings.items():
            formal_value = str(function_ws.cell(formal_row, function_headers[formal_header]).value or "").strip()
            import_value = str(import_ws.cell(import_row, import_headers[import_header]).value or "").strip()
            if formal_value != import_value:
                raise ValueError(f"generated import {import_header} does not match formal case row {formal_row}")
    for index, case in enumerate(source_cases, 2):
        if str(function_ws.cell(index, function_headers["用例标题"]).value or "").strip() != str(case.get("title", "")).strip():
            raise ValueError(f"formal workbook row {index} differs from function-cases.json")
    _assert_visible_content_clean(formal, "generated formal workbook")
    _assert_visible_content_clean(imported, "generated import workbook")


def complete_deliverables(run_dir: Path, project_root: Path) -> dict[str, Any]:
    paths = artifact_paths(run_dir)
    scope = load_facts(run_dir).get("scope", {})
    paths["delivery"].mkdir(parents=True, exist_ok=True)
    formal = paths["delivery"] / "正式测试设计.xlsx"
    import_file = paths["delivery"] / "测试系统导入.xlsx"
    formal_candidate = paths["delivery"] / ".正式测试设计.candidate.xlsx"
    import_candidate = paths["delivery"] / ".测试系统导入.candidate.xlsx"
    counts: dict[str, int] = {}
    import_count = 0
    try:
        formal_template = project_root / "docs" / "test-design" / "codebuddy-test-design-template.xlsx"
        import_template = project_root / "docs" / "test-design" / "测试用例模板.xlsx"
        for attempt in range(2):
            formal_candidate.unlink(missing_ok=True)
            import_candidate.unlink(missing_ok=True)
            try:
                counts = assemble_formal_workbook(run_dir, formal_template, formal_candidate)
                import_count = generate_import_workbook(
                    run_dir, import_template, import_candidate, str(scope.get("module_path", "")),
                )
                _verify_generated_deliverables(
                    run_dir, formal_candidate, import_candidate, import_count, formal_template, import_template,
                )
                break
            except Exception:
                if attempt == 1:
                    raise
        os.replace(formal_candidate, formal)
        os.replace(import_candidate, import_file)
    finally:
        formal_candidate.unlink(missing_ok=True)
        import_candidate.unlink(missing_ok=True)
    return {
        "delivery_dir": str(paths["delivery"].resolve()),
        "formal_workbook": str(formal.resolve()),
        "import_workbook": str(import_file.resolve()),
        "formal_workbook_name": formal.name,
        "import_workbook_name": import_file.name,
        "sheet_rows": counts,
        "import_cases": import_count,
    }
