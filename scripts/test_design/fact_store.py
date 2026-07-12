from __future__ import annotations

import hashlib
import json
import re
from copy import copy
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .io_utils import atomic_save_workbook, atomic_write_json


FACT_SCHEMA_VERSION = "2.0.0"
PRODUCT_MAP_SHEETS = [
    "产品模块地图",
    "业务对象地图",
    "业务链路地图",
    "页面元素地图",
    "用例资产索引",
    "模块能力索引",
    "跨模块依赖关系",
    "可复用测试数据",
    "变更影响分析",
    "变更记录",
]
IDENTITY_FIELDS = {
    "产品模块地图": ["产品/系统", "菜单路径/URL", "页面/入口"],
    "业务对象地图": ["产品/系统", "来源模块", "业务对象"],
    "业务链路地图": ["链路ID"],
    "页面元素地图": ["产品/系统", "模块", "页面/入口", "元素名称/文案", "元素类型"],
    "用例资产索引": ["产品/系统", "模块", "用例ID"],
    "模块能力索引": ["产品/系统", "模块", "功能点"],
    "跨模块依赖关系": ["产品/系统", "当前模块", "依赖模块", "依赖功能点/能力"],
    "可复用测试数据": ["产品/系统", "模块", "测试数据标识"],
    "变更影响分析": ["变更ID"],
    "变更记录": ["版本", "日期", "影响模块", "变更类型"],
}
MODULE_MATCH_FIELDS = {
    "产品模块地图": ["菜单路径/URL", "归档测试设计路径"],
    "业务对象地图": ["来源模块", "消费模块", "归档测试设计路径"],
    "业务链路地图": ["起始模块", "中间模块", "结束模块", "归档测试设计路径"],
    "页面元素地图": ["模块"],
    "用例资产索引": ["模块", "归档测试设计路径"],
    "模块能力索引": ["模块", "归档测试设计路径"],
    "跨模块依赖关系": ["当前模块"],
    "可复用测试数据": ["模块"],
    "变更影响分析": ["变更模块"],
    "变更记录": ["影响模块"],
}
TEMPLATE_EXACT_VALUES = {
    "FLOW-DEMO-001",
    "CHG-DEMO-001",
    "yyyy-mm-dd",
    "docs/test-assets/modules/内部工作流_测试设计.xlsx",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def header_map(ws) -> dict[str, int]:
    return {str(cell.value).strip(): cell.column for cell in ws[1] if cell.value}


def row_dict(ws, headers: dict[str, int], row_index: int) -> dict[str, str]:
    return {
        name: "" if ws.cell(row=row_index, column=column).value is None else str(ws.cell(row=row_index, column=column).value).strip()
        for name, column in headers.items()
    }


def is_template_row(row: dict[str, str]) -> bool:
    values = list(row.values())
    return any(value.startswith("示例") or value in TEMPLATE_EXACT_VALUES for value in values)


def stable_fact_id(sheet_name: str, data: dict[str, str]) -> str:
    fields = IDENTITY_FIELDS[sheet_name]
    identity = "\u001f".join(data.get(field, "").strip() for field in fields)
    if not identity.strip("\u001f"):
        identity = json.dumps(data, ensure_ascii=False, sort_keys=True)
    digest = hashlib.sha256(f"{sheet_name}\u001e{identity}".encode("utf-8")).hexdigest()[:20]
    return f"FACT-{digest}"


def fact_record(sheet_name: str, data: dict[str, str], evidence: dict[str, str]) -> dict[str, Any]:
    normalized = {str(key): "" if value is None else str(value) for key, value in data.items()}
    return {"id": stable_fact_id(sheet_name, normalized), "data": normalized, "evidence": evidence}


def validate_document(document: dict[str, Any], label: str = "fact document") -> None:
    if document.get("schema_version") != FACT_SCHEMA_VERSION:
        raise ValueError(f"{label} schema_version must be {FACT_SCHEMA_VERSION}")
    for field in ["module_key", "product", "module_path", "updated_at", "source", "facts"]:
        if field not in document:
            raise ValueError(f"{label} is missing field: {field}")
    if not isinstance(document["source"], dict) or not isinstance(document["facts"], dict):
        raise ValueError(f"{label} source and facts must be objects")
    unknown_sheets = sorted(set(document["facts"]) - set(PRODUCT_MAP_SHEETS))
    if unknown_sheets:
        raise ValueError(f"{label} contains unknown product-map sheets: {unknown_sheets}")
    seen_ids: set[str] = set()
    for sheet_name in PRODUCT_MAP_SHEETS:
        records = document["facts"].get(sheet_name, [])
        if not isinstance(records, list):
            raise ValueError(f"{label} facts.{sheet_name} must be an array")
        for index, record in enumerate(records, start=1):
            if not isinstance(record, dict) or set(record) != {"id", "data", "evidence"}:
                raise ValueError(f"{label} {sheet_name}[{index}] must contain id, data, evidence")
            if record["id"] in seen_ids:
                raise ValueError(f"{label} contains duplicate fact id: {record['id']}")
            seen_ids.add(record["id"])
            if not isinstance(record["data"], dict) or not isinstance(record["evidence"], dict):
                raise ValueError(f"{label} {sheet_name}[{index}] data/evidence must be objects")
            expected_id = stable_fact_id(sheet_name, record["data"])
            if record["id"] != expected_id:
                raise ValueError(f"{label} {sheet_name}[{index}] fact id is not stable: {record['id']}")


def catalog_dir(product_map: Path) -> Path:
    return product_map.resolve().parent / "catalog"


def module_document_name(module_key: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "_", module_key).strip("_")[:80] or "module"
    digest = hashlib.sha256(module_key.encode("utf-8")).hexdigest()[:12]
    return f"{slug}_{digest}.json"


def workbook_facts(product_map: Path) -> dict[str, list[dict[str, str]]]:
    workbook = load_workbook(product_map, data_only=True, read_only=True)
    result = {sheet_name: [] for sheet_name in PRODUCT_MAP_SHEETS}
    try:
        for sheet_name in PRODUCT_MAP_SHEETS:
            ws = workbook[sheet_name]
            if ws.max_row is None:
                ws.calculate_dimension(force=True)
            headers = header_map(ws)
            for row_index in range(2, (ws.max_row or 1) + 1):
                row = row_dict(ws, headers, row_index)
                if any(row.values()) and not is_template_row(row):
                    result[sheet_name].append(row)
    finally:
        workbook.close()
    return result


def document_from_rows(
    module_key: str,
    product: str,
    module_path: str,
    source: dict[str, str],
    rows_by_sheet: dict[str, list[dict[str, str]]],
) -> dict[str, Any]:
    evidence = {"type": source.get("type", "generated"), "source": source.get("source", "")}
    document = {
        "schema_version": FACT_SCHEMA_VERSION,
        "module_key": module_key,
        "product": product,
        "module_path": module_path,
        "updated_at": utc_now(),
        "source": source,
        "facts": {
            sheet_name: [fact_record(sheet_name, row, evidence) for row in rows_by_sheet.get(sheet_name, [])]
            for sheet_name in PRODUCT_MAP_SHEETS
        },
    }
    validate_document(document)
    return document


def ensure_catalog(product_map: Path) -> Path:
    root = catalog_dir(product_map)
    modules_dir = root / "modules"
    modules_dir.mkdir(parents=True, exist_ok=True)
    marker = root / "migration.json"
    if not marker.exists():
        existing_rows = workbook_facts(product_map)
        if any(existing_rows.values()):
            legacy = document_from_rows(
                "__legacy__",
                "",
                "__legacy__",
                {"type": "excel-migration", "source": product_map.name},
                existing_rows,
            )
            atomic_write_json(modules_dir / "_legacy.json", legacy)
        atomic_write_json(
            marker,
            {
                "schema_version": FACT_SCHEMA_VERSION,
                "migrated_from": product_map.name,
                "migrated_at": utc_now(),
            },
        )
    return root


def load_documents(product_map: Path, require_existing: bool = False) -> list[tuple[Path, dict[str, Any]]]:
    root = catalog_dir(product_map)
    if require_existing:
        if not (root / "migration.json").exists() or not (root / "modules").is_dir():
            raise ValueError(f"Product fact catalog is missing beside product map: {root}")
    else:
        root = ensure_catalog(product_map)
    documents: list[tuple[Path, dict[str, Any]]] = []
    for path in sorted((root / "modules").glob("*.json"), key=lambda item: (item.name != "_legacy.json", item.name)):
        with path.open("r", encoding="utf-8-sig") as fp:
            document = json.load(fp)
        validate_document(document, str(path))
        documents.append((path, document))
    return documents


def prune_legacy_document(product_map: Path, exact_values: set[str]) -> None:
    legacy_path = catalog_dir(product_map) / "modules" / "_legacy.json"
    if not legacy_path.exists():
        return
    with legacy_path.open("r", encoding="utf-8-sig") as fp:
        document = json.load(fp)
    validate_document(document, str(legacy_path))
    for sheet_name, fields in MODULE_MATCH_FIELDS.items():
        document["facts"][sheet_name] = [
            record
            for record in document["facts"].get(sheet_name, [])
            if not ({record["data"].get(field, "") for field in fields} & exact_values)
        ]
    document["updated_at"] = utc_now()
    atomic_write_json(legacy_path, document)


def save_module_document(
    product_map: Path,
    module_key: str,
    product: str,
    module_path: str,
    archive_path: str,
    rows_by_sheet: dict[str, list[dict[str, str]]],
    source: dict[str, str],
) -> Path:
    root = ensure_catalog(product_map)
    prune_legacy_document(product_map, {module_path, archive_path})
    document = document_from_rows(module_key, product, module_path, source, rows_by_sheet)
    path = root / "modules" / module_document_name(module_key)
    atomic_write_json(path, document)
    rebuild_index(product_map)
    return path


def rebuild_index(product_map: Path) -> None:
    root = ensure_catalog(product_map)
    modules = []
    for path, document in load_documents(product_map):
        modules.append(
            {
                "module_key": document["module_key"],
                "product": document["product"],
                "module_path": document["module_path"],
                "file": path.relative_to(root).as_posix(),
                "updated_at": document["updated_at"],
            }
        )
    atomic_write_json(
        root / "index.json",
        {"schema_version": FACT_SCHEMA_VERSION, "generated_at": utc_now(), "modules": modules},
    )


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for source_cell in ws[source_row]:
        target_cell = ws.cell(row=target_row, column=source_cell.column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def project_catalog_to_workbook(product_map: Path) -> None:
    documents = load_documents(product_map)
    workbook = load_workbook(product_map)
    for sheet_name in PRODUCT_MAP_SHEETS:
        ws = workbook[sheet_name]
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)
        if ws.max_row < 2:
            ws.insert_rows(2)
        for cell in ws[2]:
            cell.value = None

        records_by_id: dict[str, dict[str, Any]] = {}
        for _, document in documents:
            for record in document["facts"].get(sheet_name, []):
                records_by_id[record["id"]] = record

        headers = header_map(ws)
        for offset, record in enumerate(records_by_id.values(), start=2):
            if offset > 2:
                copy_row_style(ws, 2, offset)
            for field, value in record["data"].items():
                column = headers.get(field)
                if column:
                    ws.cell(row=offset, column=column, value=value)

        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    atomic_save_workbook(workbook, product_map)


def validate_catalog(product_map: Path, require_existing: bool = False) -> dict[str, int]:
    documents = load_documents(product_map, require_existing=require_existing)
    counts = {sheet_name: 0 for sheet_name in PRODUCT_MAP_SHEETS}
    seen_module_keys: dict[str, Path] = {}
    products_by_module_path: dict[str, set[str]] = {}
    seen_fact_ids: dict[str, Path] = {}
    for path, document in documents:
        module_key = str(document.get("module_key", "")).strip()
        product = str(document.get("product", "")).strip()
        module_path = str(document.get("module_path", "")).strip()
        previous_module = seen_module_keys.get(module_key)
        if previous_module is not None:
            raise ValueError(
                f"Product fact catalog contains duplicate module_key {module_key!r}: {previous_module} and {path}"
            )
        seen_module_keys[module_key] = path
        if path.name != "_legacy.json":
            expected_name = module_document_name(module_key)
            if path.name != expected_name:
                raise ValueError(
                    f"Product fact module document filename is not canonical for {module_key!r}: "
                    f"{path.name} != {expected_name}"
                )
            expected_key = (
                module_path
                if module_path == product or module_path.startswith(f"{product}>")
                else f"{product}>{module_path}"
            )
            if module_key != expected_key:
                raise ValueError(
                    f"Product fact module_key must bind product and module_path: {module_key!r} != {expected_key!r}"
                )
            fact_total = sum(len(document["facts"].get(sheet_name, [])) for sheet_name in PRODUCT_MAP_SHEETS)
            if fact_total == 0:
                raise ValueError(f"Product fact module document must not be empty: {path}")
            products_by_module_path.setdefault(module_path, set()).add(product)
        for sheet_name in PRODUCT_MAP_SHEETS:
            records = document["facts"].get(sheet_name, [])
            counts[sheet_name] += len(records)
            for record in records:
                fact_id = str(record.get("id", ""))
                previous_fact = seen_fact_ids.get(fact_id)
                if previous_fact is not None:
                    raise ValueError(
                        f"Product fact catalog contains duplicate fact id {fact_id}: {previous_fact} and {path}"
                    )
                seen_fact_ids[fact_id] = path
                fact_product = str(record.get("data", {}).get("产品/系统", "")).strip()
                if fact_product and product and fact_product != product:
                    raise ValueError(
                        f"Product fact {fact_id} 产品/系统={fact_product!r} does not match document product={product!r}"
                    )
    conflicting_paths = {
        module_path: sorted(products)
        for module_path, products in products_by_module_path.items()
        if len(products) > 1 and module_path.split(">", 1)[0] in products
    }
    if conflicting_paths:
        raise ValueError(
            "Product fact catalog maps the same module_path to conflicting products and one product is actually "
            f"the path's first-level module: {conflicting_paths}"
        )
    index_path = catalog_dir(product_map) / "index.json"
    if require_existing and not index_path.exists():
        raise ValueError(f"Product fact catalog index is missing: {index_path}")
    if not index_path.exists():
        rebuild_index(product_map)
    with index_path.open("r", encoding="utf-8-sig") as fp:
        index = json.load(fp)
    if index.get("schema_version") != FACT_SCHEMA_VERSION or not isinstance(index.get("modules"), list):
        raise ValueError(f"Product fact catalog index has an invalid schema: {index_path}")
    root = catalog_dir(product_map)
    indexed_files = {str(item.get("file", "")) for item in index["modules"]}
    document_files = {path.relative_to(root).as_posix() for path, _ in documents}
    if indexed_files != document_files:
        raise ValueError(
            f"Product fact catalog index does not match module documents: indexed={sorted(indexed_files)} "
            f"actual={sorted(document_files)}"
        )
    return counts
