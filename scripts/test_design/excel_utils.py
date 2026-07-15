from __future__ import annotations

from copy import copy

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.cell_range import MultiCellRange


def header_map(worksheet, header_row: int = 1) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in worksheet[header_row]
        if cell.value is not None and str(cell.value).strip()
    }


def clear_data_rows(worksheet, start_row: int = 2) -> None:
    if worksheet.max_row > start_row:
        worksheet.delete_rows(start_row + 1, worksheet.max_row - start_row)
    if worksheet.max_row >= start_row:
        for cell in worksheet[start_row]:
            cell.value = None


def _copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    for source_cell in worksheet[source_row]:
        target_cell = worksheet.cell(row=target_row, column=source_cell.column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height


def write_mapped_row(worksheet, headers: dict[str, int], row_index: int, values: dict[str, object]) -> None:
    _copy_row_style(worksheet, 2 if worksheet.max_row >= 2 else 1, row_index)
    for field, value in values.items():
        if field in headers:
            worksheet.cell(row=row_index, column=headers[field], value=value)


def resize_workbook_structures(workbook) -> None:
    """Resize template-owned tables, filters and validations without deleting them."""
    for worksheet in workbook.worksheets:
        last_row = max(worksheet.max_row, 2)
        for table in worksheet.tables.values():
            min_col, min_row, max_col, _ = range_boundaries(table.ref)
            table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"
        if worksheet.auto_filter.ref:
            min_col, min_row, max_col, _ = range_boundaries(worksheet.auto_filter.ref)
            worksheet.auto_filter.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"
        for validation in worksheet.data_validations.dataValidation:
            ranges = []
            for cell_range in validation.ranges.ranges:
                min_col, min_row, max_col, _ = range_boundaries(str(cell_range))
                if min_row >= 2:
                    ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}")
                else:
                    ranges.append(str(cell_range))
            validation.sqref = MultiCellRange(ranges)
