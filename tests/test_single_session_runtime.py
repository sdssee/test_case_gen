# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import time
import unittest
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from test_design.formal_assembler import SHEETS, complete_deliverables
from test_design.session_runtime import (
    append_events,
    artifact_paths,
    build_plan_skeleton,
    checkpoint_facts,
    compile_facts,
    ensure_run,
    load_facts,
    load_cases,
    load_plan,
    pipeline_status,
    review_run,
    save_cases,
    save_plan,
    validate_cases,
    validate_discovery,
    validate_plan,
)
from test_design_cli import _payload, _project_scoped_run_dir, execute_request


def _mark_final_scan(run_dir: Path, page_ref: str = "PAGE") -> None:
    append_events(run_dir, [{
        "kind": "page", "fact_id": page_ref,
        "data": {"final_scan_status": "stable", "unhandled_element_refs": []},
    }])


def _plan_metadata(goal: str = "验证页面功能") -> dict[str, object]:
    return {
        "design_context": {
            "user_goal": goal,
            "role": "具备页面访问权限的用户",
            "business_value": "保证用户操作得到明确且正确的页面反馈",
            "acceptance_criteria": ["各独立场景产生与实探一致的稳定结果"],
            "business_rules": ["有限选项分别验证，不进行默认组合"],
            "dependencies": ["具备页面访问权限和受控测试数据"],
            "postcondition": "页面保持在可继续验证的稳定状态",
            "basis": ["页面实探"],
        },
        "automation_profile": {
            "level": "UI", "dependency": "受控测试数据",
            "stability_risk": "页面控件定位变化", "recommendation": "沿用项目UI自动化框架",
        },
    }


def _semantic_review(run_dir: Path) -> dict[str, object]:
    return {
        "reviewed_case_ids": [str(row.get("case_id", "")) for row in load_cases(run_dir).get("cases", [])],
        "reviewed_sections": ["cases", "performance", "risks", "automation", "elements", "cross_sheet"],
        "summary": "逐条复核当前用例，未发现需要局部修正的语义问题。",
        "issues": [], "local_fixes": [],
    }


class SingleSessionRuntimeTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.run_dir = Path(self.temporary.name) / "run-001"
        ensure_run(self.run_dir, "告警管理>告警列表", "大数据平台", "需求说明")
        append_events(self.run_dir, [
            {"kind": "page", "fact_id": "PAGE-LIST", "data": {
                "name": "告警列表", "menu_path": ["告警管理", "告警列表"], "result_anchor": "告警列表",
                "final_scan_status": "stable", "unhandled_element_refs": []
            }},
            {"kind": "function", "fact_id": "FN-VIEW", "data": {"name": "告警视图模式"}},
            {"kind": "element", "fact_id": "EL-VIEW-MODE", "data": {
                "function_ref": "FN-VIEW", "page_ref": "PAGE-LIST", "name": "视图模式",
                "type": "下拉框", "interactive": True, "option_set": "finite",
                "options": ["精简", "标准", "详细", "仅未确认", "仅已确认", "仅严重", "全部"]
            }},
            {"kind": "transaction", "fact_id": "TX-VIEW", "data": {
                "function_ref": "FN-VIEW", "element_refs": ["EL-VIEW-MODE"],
                "transaction_type": "selection", "recovery_result": "恢复标准视图",
                "checks": [
                    {"element_ref": "EL-VIEW-MODE", "action": "选择精简视图", "option_value": "精简", "result": "列表仅显示核心告警字段",
                     "result_anchor": {"assertion": "field_visible", "target": "列表", "field": "核心告警字段", "tokens": ["核心告警字段"]}},
                    {"element_ref": "EL-VIEW-MODE", "action": "选择标准视图", "option_value": "标准", "result": "列表显示默认告警字段", "result_anchor": {"assertion": "contains", "value": "默认告警字段"}},
                    {"element_ref": "EL-VIEW-MODE", "action": "选择详细视图", "option_value": "详细", "result": "列表显示全部可见告警字段", "result_anchor": {"assertion": "contains", "value": "全部可见告警字段"}},
                    {"element_ref": "EL-VIEW-MODE", "action": "选择仅未确认", "option_value": "仅未确认", "result": "列表只显示未确认告警", "result_anchor": {"assertion": "contains", "value": "未确认告警"}},
                    {"element_ref": "EL-VIEW-MODE", "action": "选择仅已确认", "option_value": "仅已确认", "result": "列表只显示已确认告警", "result_anchor": {"assertion": "contains", "value": "已确认告警"}},
                    {"element_ref": "EL-VIEW-MODE", "action": "选择仅严重", "option_value": "仅严重", "result": "列表只显示严重告警", "result_anchor": {"assertion": "contains", "value": "严重告警"}},
                    {"element_ref": "EL-VIEW-MODE", "action": "选择全部", "option_value": "全部", "result": "列表恢复显示全部告警", "result_anchor": {"assertion": "contains", "value": "全部告警"}},
                ],
            }},
        ])
        _mark_final_scan(self.run_dir, "PAGE-LIST")
        checkpoint_facts(self.run_dir)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def _write_plan_and_cases(self) -> None:
        refs = ["FN-VIEW", "EL-VIEW-MODE", "TX-VIEW"]
        branches = [
            ("精简", "列表仅显示核心告警字段", "精简视图显示"),
            ("标准", "列表显示默认告警字段", "标准视图显示"),
            ("详细", "列表显示全部可见告警字段", "详细视图显示"),
            ("仅未确认", "列表只显示未确认告警", "仅未确认告警筛选"),
            ("仅已确认", "列表只显示已确认告警", "仅已确认告警筛选"),
            ("仅严重", "列表只显示严重告警", "仅严重告警筛选"),
            ("全部", "列表恢复显示全部告警", "全部告警显示"),
        ]
        plan = {
            "schema_version": "2.0", "risks": [],
            "risk_not_applicable_reason": "实探未发现需要单独登记的风险",
            "performance_scenarios": [], "performance_not_applicable_reason": "本功能不包含可独立定义指标的性能链路",
            "functions": [{"function_ref": "FN-VIEW", "name": "告警视图模式", **_plan_metadata("选择不同视图并查看对应告警字段"), "cases": [
                {"case_id": f"TC-VIEW-{index:03d}", "title": title, "strategy": "baseline", "page_ref": "PAGE-LIST"}
                for index, (_, _, title) in enumerate(branches, 1)
            ]}],
            "check_assignments": [
                {"transaction_ref": "TX-VIEW", "check_index": index, "disposition": "case", "case_id": f"TC-VIEW-{index:03d}"}
                for index in range(1, len(branches) + 1)
            ],
        }
        save_plan(self.run_dir, plan)
        navigation = {"action": "进入告警管理>告警列表", "expected": "显示告警列表和查询区"}
        cases = {"schema_version": "2.0", "source_plan": "case-plan.json", "cases": [
            {"case_id": f"TC-VIEW-{index:03d}", "function_ref": "FN-VIEW", "title": f"告警视图模式-{title}",
             "priority": "P1" if index < 6 else "P2", "test_type": "功能测试",
             "preconditions": ["告警列表存在可查看且能够区分当前选项效果的数据"], "test_data": f"视图模式：{option}",
             "steps": [navigation, {"action": f"选择{option}视图模式", "expected": expected}],
             "fact_refs": refs, "automation_value": "高频稳定回归", "automation_priority": "P1" if index < 6 else "P2"}
            for index, (option, expected, title) in enumerate(branches, 1)
        ]}
        save_cases(self.run_dir, cases)

    def test_one_transaction_assigns_each_finite_option_to_its_own_case(self) -> None:
        self.assertEqual([], validate_discovery(self.run_dir))
        facts = load_facts(self.run_dir)
        self.assertEqual(1, len(facts["transactions"]))
        self.assertEqual(7, len(facts["transactions"][0]["checks"]))
        skeleton = build_plan_skeleton(self.run_dir)
        self.assertEqual(7, len(skeleton["functions"][0]["checks"]))
        self._write_plan_and_cases()
        self.assertEqual(7, sum(len(item["cases"]) for item in json.loads(artifact_paths(self.run_dir)["plan"].read_text(encoding="utf-8"))["functions"]))
        written = load_cases(self.run_dir)
        self.assertEqual("进入告警管理-告警列表", written["cases"][0]["steps"][0]["action"])
        self.assertEqual({"transaction_ref": "TX-VIEW", "check_index": 1}, written["cases"][0]["steps"][1]["source_check"])
        self.assertIn("EL-VIEW-MODE", written["cases"][0]["fact_refs"])

    def test_named_test_data_reference_requires_the_same_controlled_source(self) -> None:
        self._write_plan_and_cases()
        cases = load_cases(self.run_dir)
        cases["cases"][0]["test_data"] = "有效目标地址"
        with self.assertRaisesRegex(ValueError, "natural-language data placeholder"):
            save_cases(self.run_dir, cases)
        cases["cases"][0]["test_data"] = "TEST_VALID_TARGET"
        with self.assertRaisesRegex(ValueError, "controlled source"):
            save_cases(self.run_dir, cases)

    def test_review_is_ready_and_has_no_evidence_dependency(self) -> None:
        self._write_plan_and_cases()
        self.assertEqual([], validate_plan(self.run_dir))
        self.assertEqual([], validate_cases(self.run_dir))
        facts_before = artifact_paths(self.run_dir)["facts"].read_bytes()
        partial_review = _semantic_review(self.run_dir)
        partial_review["reviewed_sections"] = ["cases"]
        self.assertEqual("needs_local_fix", review_run(self.run_dir, partial_review)["status"])
        self.assertEqual("ready", review_run(self.run_dir, _semantic_review(self.run_dir))["status"])
        self.assertEqual(facts_before, artifact_paths(self.run_dir)["facts"].read_bytes())
        self.assertFalse(artifact_paths(self.run_dir)["diagnostics"].exists())
        self.assertEqual("ready", pipeline_status(self.run_dir)["state"])

    def test_review_without_model_semantics_cannot_be_delivered(self) -> None:
        self._write_plan_and_cases()
        review = review_run(self.run_dir)
        self.assertEqual("needs_local_fix", review["status"])
        self.assertEqual("missing", review["semantic"]["status"])
        with self.assertRaisesRegex(ValueError, "delivery requires"):
            complete_deliverables(self.run_dir, ROOT)

    def test_scope_binding_resumes_same_target_and_rejects_a_different_target(self) -> None:
        resumed = ensure_run(self.run_dir, "告警管理>告警列表")
        self.assertEqual("告警管理>告警列表", resumed["module_path"])
        self.assertNotIn("menu_path", resumed)
        with self.assertRaisesRegex(ValueError, "choose a new run directory"):
            ensure_run(self.run_dir, "告警管理>告警详情")

    def test_scope_updates_merge_and_keep_run_identity(self) -> None:
        before = load_facts(self.run_dir)["scope"]
        append_events(self.run_dir, [{"kind": "scope", "fact_id": "SCOPE", "data": {
            "owner": "测试负责人", "version": "迭代一"
        }}])
        after = compile_facts(self.run_dir)["scope"]
        self.assertEqual(before["run_id"], after["run_id"])
        self.assertEqual(before["created_at"], after["created_at"])
        self.assertEqual(before["module_path"], after["module_path"])
        self.assertEqual("测试负责人", after["owner"])

    def test_fact_compilation_keeps_first_discovery_order(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "order"
            ensure_run(run_dir, "工具>顺序")
            append_events(run_dir, [
                {"kind": "function", "fact_id": "FN-Z", "data": {"name": "先发现"}},
                {"kind": "function", "fact_id": "FN-A", "data": {"name": "后发现"}},
                {"kind": "function", "fact_id": "FN-Z", "data": {"name": "先发现-已更新"}},
            ])
            facts = compile_facts(run_dir)
            self.assertEqual(["FN-Z", "FN-A"], [row["fact_id"] for row in facts["functions"]])
            self.assertEqual("先发现-已更新", facts["functions"][0]["name"])

    def test_cli_rejects_a_run_directory_outside_the_project(self) -> None:
        with tempfile.TemporaryDirectory() as outside:
            result = subprocess.run(
                [sys.executable, str(ROOT / "scripts" / "test_design_cli.py"), "status", "--run-dir", outside],
                cwd=ROOT, capture_output=True, text=True, check=False,
            )
        self.assertNotEqual(0, result.returncode)
        self.assertIn("must stay inside the current project root", result.stderr)

    def test_short_run_id_resolves_to_the_canonical_current_directory(self) -> None:
        resolved = _project_scoped_run_dir(Path("run-canonical-test"))
        self.assertEqual((ROOT / "docs" / "test-design" / "current" / "run-canonical-test").resolve(), resolved)

    def test_temporary_cli_payload_is_removed_after_read(self) -> None:
        path = Path(tempfile.gettempdir()) / "test-design-payload-cleanup.json"
        path.write_text('{"ok": true}', encoding="utf-8")
        self.assertEqual({"ok": True}, _payload(path))
        self.assertFalse(path.exists())

    def test_exact_duplicate_event_is_absorbed_without_another_line(self) -> None:
        event = {"kind": "open_item", "fact_id": "OPEN-IDEMPOTENT", "data": {
            "category": "observed_risk", "description": "一次性记录", "material": False,
            "affected_function_refs": ["FN-VIEW"],
        }}
        first = append_events(self.run_dir, [event])[0]
        before = len(artifact_paths(self.run_dir)["events"].read_text(encoding="utf-8").splitlines())
        second = append_events(self.run_dir, [event])[0]
        after = len(artifact_paths(self.run_dir)["events"].read_text(encoding="utf-8").splitlines())
        self.assertEqual(first["fact_id"], second["fact_id"])
        self.assertEqual(before, after)

    def test_plan_and_cases_upsert_one_function_without_rewriting_other_functions(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "incremental"
            ensure_run(run_dir, "工具>双功能")
            append_events(run_dir, [
                {"kind": "page", "fact_id": "PAGE", "data": {"name": "双功能", "menu_path": ["工具", "双功能"], "final_scan_status": "stable", "unhandled_element_refs": []}},
                {"kind": "function", "fact_id": "FN-A", "data": {"name": "功能A"}},
                {"kind": "function", "fact_id": "FN-B", "data": {"name": "功能B"}},
                {"kind": "element", "fact_id": "EL-A", "data": {"page_ref": "PAGE", "function_ref": "FN-A", "name": "执行A", "type": "按钮", "interactive": True}},
                {"kind": "element", "fact_id": "EL-B", "data": {"page_ref": "PAGE", "function_ref": "FN-B", "name": "执行B", "type": "按钮", "interactive": True}},
                {"kind": "transaction", "fact_id": "TX-A", "data": {"function_ref": "FN-A", "element_refs": ["EL-A"], "checks": [
                    {"element_ref": "EL-A", "used_element_refs": ["EL-A"], "trigger_element_ref": "EL-A", "action": "点击执行A", "result": "显示A结果", "result_anchor": {"assertion": "contains", "value": "A结果"}}
                ]}},
                {"kind": "transaction", "fact_id": "TX-B", "data": {"function_ref": "FN-B", "element_refs": ["EL-B"], "checks": [
                    {"element_ref": "EL-B", "used_element_refs": ["EL-B"], "trigger_element_ref": "EL-B", "action": "点击执行B", "result": "显示B结果", "result_anchor": {"assertion": "contains", "value": "B结果"}}
                ]}},
            ])
            _mark_final_scan(run_dir)
            self.assertTrue(checkpoint_facts(run_dir)["ready"])
            for suffix in ("A", "B"):
                save_plan(run_dir, {
                    "schema_version": "2.0",
                    "performance_scenarios": [], "performance_not_applicable_reason": "无独立性能指标",
                    "risks": [], "risk_not_applicable_reason": "未发现风险",
                    "functions": [{"function_ref": f"FN-{suffix}", "name": f"功能{suffix}", **_plan_metadata(f"执行功能{suffix}"), "cases": [
                        {"case_id": f"TC-{suffix}", "page_ref": "PAGE", "title": f"执行{suffix}", "strategy": "baseline"}
                    ]}],
                    "check_assignments": [{"transaction_ref": f"TX-{suffix}", "check_index": 1, "disposition": "case", "case_id": f"TC-{suffix}"}],
                })
                if suffix == "A":
                    self.assertEqual("continue", pipeline_status(run_dir)["state"])
            self.assertEqual(["FN-A", "FN-B"], [row["function_ref"] for row in load_plan(run_dir)["functions"]])
            with self.assertRaisesRegex(ValueError, "no planned disposition"):
                save_plan(run_dir, {
                    "schema_version": "2.0",
                    "performance_scenarios": [], "performance_not_applicable_reason": "无独立性能指标",
                    "risks": [], "risk_not_applicable_reason": "未发现风险",
                    "functions": [{"function_ref": "FN-A", "name": "功能A", **_plan_metadata("执行功能A"), "cases": [
                        {"case_id": "TC-A", "page_ref": "PAGE", "title": "执行A", "strategy": "baseline"}
                    ]}],
                    "check_assignments": [],
                })
            navigation = {"action": "进入工具-双功能", "expected": "显示双功能页面"}
            for suffix in ("A", "B"):
                save_cases(run_dir, {"schema_version": "2.0", "cases": [{
                    "case_id": f"TC-{suffix}", "function_ref": f"FN-{suffix}", "title": f"功能{suffix}-执行{suffix}",
                    "priority": "P1", "test_type": "功能测试", "preconditions": ["具备页面访问权限"], "test_data": f"功能{suffix}受控数据",
                    "automation_value": "稳定回归", "automation_priority": "P1",
                    "steps": [navigation, {"action": f"点击执行{suffix}", "expected": f"显示{suffix}结果"}],
                }]})
                if suffix == "A":
                    self.assertEqual("continue", pipeline_status(run_dir)["state"])
            before = artifact_paths(run_dir)["cases"].read_bytes()
            current_b = next(row for row in load_cases(run_dir)["cases"] if row["function_ref"] == "FN-B")
            save_cases(run_dir, {"schema_version": "2.0", "cases": [current_b]})
            self.assertEqual(before, artifact_paths(run_dir)["cases"].read_bytes())
            self.assertEqual("ready", review_run(run_dir, _semantic_review(run_dir))["status"])

    def test_resume_automatically_rebuilds_a_stale_facts_view(self) -> None:
        append_events(self.run_dir, [{"kind": "open_item", "data": {
            "category": "observed_risk", "description": "非阻塞提示", "material": False,
            "affected_function_refs": ["FN-VIEW"],
        }}])
        resumed = ensure_run(self.run_dir, "告警管理>告警列表")
        self.assertEqual("告警管理>告警列表", resumed["module_path"])
        self.assertEqual(1, len(load_facts(self.run_dir)["open_items"]))

    def test_recording_defers_fact_rebuild_until_checkpoint(self) -> None:
        facts_path = artifact_paths(self.run_dir)["facts"]
        before = json.loads(facts_path.read_text(encoding="utf-8"))["event_count"]
        append_events(self.run_dir, [{"kind": "open_item", "data": {
            "category": "observed_risk", "description": "页面响应偶发波动", "material": False,
            "affected_function_refs": ["FN-VIEW"],
        }}])
        self.assertEqual(before, json.loads(facts_path.read_text(encoding="utf-8"))["event_count"])
        checkpoint_facts(self.run_dir)
        self.assertEqual(before + 1, json.loads(facts_path.read_text(encoding="utf-8"))["event_count"])

    def test_recompile_without_business_changes_keeps_review_valid(self) -> None:
        self._write_plan_and_cases()
        self.assertEqual(["FN-VIEW"], load_plan(self.run_dir)["performance_basis_refs"])
        self.assertEqual(["页面控件定位变化"], [row["description"] for row in load_plan(self.run_dir)["risks"]])
        self.assertEqual("ready", review_run(self.run_dir, _semantic_review(self.run_dir))["status"])
        time.sleep(1.1)
        compile_facts(self.run_dir)
        paths = artifact_paths(self.run_dir)
        plan = json.loads(paths["plan"].read_text(encoding="utf-8"))
        plan["generated_at"] = "2099-01-01T00:00:00Z"
        paths["plan"].write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
        cases = json.loads(paths["cases"].read_text(encoding="utf-8"))
        cases["updated_at"] = "2099-01-02T00:00:00Z"
        paths["cases"].write_text(json.dumps(cases, ensure_ascii=False), encoding="utf-8")
        self.assertEqual("ready", pipeline_status(self.run_dir)["state"])
        receipt = complete_deliverables(self.run_dir, ROOT)
        self.assertEqual(7, receipt["import_cases"])

    def test_crud_lifecycle_dfx_hint_is_transaction_scoped_once(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "lifecycle"
            ensure_run(run_dir, "告警管理>告警规则")
            append_events(run_dir, [
                {"kind": "page", "fact_id": "PAGE", "data": {"name": "告警规则", "menu_path": ["告警管理", "告警规则"], "final_scan_status": "stable", "unhandled_element_refs": []}},
                {"kind": "function", "fact_id": "FN", "data": {"name": "新建规则"}},
                {"kind": "test_object", "fact_id": "OBJ", "data": {"name": "AI_TEST_RULE_001", "owner": "current_run", "state": "cleaned"}},
                {"kind": "element", "fact_id": "EL-NAME", "data": {"page_ref": "PAGE", "function_ref": "FN", "name": "规则名称", "type": "文本框", "interactive": True}},
                {"kind": "element", "fact_id": "EL-SAVE", "data": {"page_ref": "PAGE", "function_ref": "FN", "name": "保存", "type": "按钮", "interactive": True}},
                {"kind": "transaction", "fact_id": "TX-CREATE", "data": {
                    "function_ref": "FN", "element_refs": ["EL-NAME", "EL-SAVE"], "transaction_type": "create",
                    "test_object_ref": "OBJ", "outcome": "success", "commit_result": "保存成功",
                    "persistence_result": "重新打开后规则名称一致", "effect_result": "规则出现在列表中",
                    "checks": [
                        {"element_ref": "EL-NAME", "input_class": "valid", "action": "输入AI_TEST_RULE_001", "result": "名称字段显示AI_TEST_RULE_001", "result_anchor": {"assertion": "field_equals", "field": "规则名称", "value": "AI_TEST_RULE_001"}},
                        {"element_ref": "EL-SAVE", "action": "单击保存", "result": "页面提示保存成功", "result_anchor": {"assertion": "contains", "value": "保存成功"}},
                    ],
                }},
            ])
            _mark_final_scan(run_dir)
            checkpoint = checkpoint_facts(run_dir)
            self.assertTrue(checkpoint["ready"])
            hints = build_plan_skeleton(run_dir)["functions"][0]["dfx_hints"]
            lifecycle = [item for item in hints if item["code"] == "lifecycle"]
            self.assertEqual(1, len(lifecycle))
            self.assertEqual("transaction", lifecycle[0]["scope"])
            self.assertEqual([
                {"transaction_ref": "TX-CREATE", "check_index": 1},
                {"transaction_ref": "TX-CREATE", "check_index": 2},
            ], lifecycle[0]["related_checks"])

    def test_runtime_allocates_ids_and_resolves_batch_local_references(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "runtime-ids"
            ensure_run(run_dir, "告警管理>告警列表")
            recorded = append_events(run_dir, [
                {"kind": "page", "local_ref": "page", "data": {
                    "name": "告警列表", "menu_path": ["告警管理", "告警列表"],
                    "final_scan_status": "stable", "unhandled_element_refs": [],
                }},
                {"kind": "function", "local_ref": "function", "data": {"name": "刷新"}},
                {"kind": "element", "local_ref": "element", "data": {
                    "page_ref": "@page", "function_ref": "@function", "name": "刷新", "type": "按钮", "interactive": True,
                }},
                {"kind": "transaction", "data": {
                    "function_ref": "@function", "element_refs": ["@element"], "checks": [
                        {"element_ref": "@element", "action": "单击刷新", "result": "列表数据重新加载", "result_anchor": {"assertion": "contains", "value": "列表数据"}}
                    ],
                }},
            ])
            self.assertTrue(all(item["fact_id"] for item in recorded))
            facts = compile_facts(run_dir)
            self.assertEqual(facts["pages"][0]["fact_id"], facts["elements"][0]["page_ref"])
            self.assertEqual(facts["functions"][0]["fact_id"], facts["transactions"][0]["function_ref"])

    def test_persistent_client_ref_merges_exact_updates_across_batches(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "persistent-ref"
            ensure_run(run_dir, "工具>通用页面")
            first = append_events(run_dir, [{
                "kind": "page", "client_ref": "page:main",
                "data": {"name": "通用页面", "menu_path": ["工具", "通用页面"]},
            }])[0]
            second = append_events(run_dir, [{
                "kind": "page", "client_ref": "page:main",
                "data": {"final_scan_status": "stable", "unhandled_element_refs": []},
            }])[0]
            self.assertEqual(first["fact_id"], second["fact_id"])
            append_events(run_dir, [{
                "kind": "function", "client_ref": "function:primary",
                "data": {"name": "主功能", "page_ref": "@page:main"},
            }])
            facts = compile_facts(run_dir)
            self.assertEqual(1, len(facts["pages"]))
            self.assertEqual("stable", facts["pages"][0]["final_scan_status"])
            self.assertEqual(facts["pages"][0]["fact_id"], facts["functions"][0]["page_ref"])

    def test_misplaced_event_envelope_fields_are_normalized_before_append(self) -> None:
        append_events(self.run_dir, [{
            "kind": "page",
            "data": {"fact_id": "PAGE-LIST", "status": "active", "name": "告警记录列表"},
        }])
        facts = compile_facts(self.run_dir)
        self.assertEqual(1, len(facts["pages"]))
        self.assertEqual("PAGE-LIST", facts["pages"][0]["fact_id"])
        self.assertEqual("告警记录列表", facts["pages"][0]["name"])
        self.assertNotIn("fact_id", facts["pages"][0].get("data", {}))

    def test_model_shaped_controls_are_normalized_without_product_rules(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "normalization"
            ensure_run(run_dir, "工具>通用页面")
            recorded = append_events(run_dir, [
                {"kind": "element", "data": {
                    "name": "目标", "type": "textbox", "required": True,
                    "valid_input_classes": [
                        {"class": "textual", "description": "有效文本"},
                        {"class": "numeric", "description": "有效数字"},
                    ],
                }},
                {"kind": "element", "data": {
                    "name": "模式", "type": "combobox",
                    "options": [{"value": "A", "label": "模式A"}, {"value": "B", "label": "模式B"}],
                }},
            ])
            input_element, select_element = [item["data"] for item in recorded]
            self.assertEqual("input", input_element["type"])
            self.assertEqual(["valid_textual", "valid_numeric", "empty"], [row["value"] for row in input_element["exploration_requirements"]])
            self.assertEqual("select", select_element["type"])
            self.assertEqual("finite", select_element["option_set"])
            self.assertEqual(["A", "B"], [row["value"] for row in select_element["exploration_requirements"]])

    def test_cli_and_fallback_share_the_same_execution_adapter(self) -> None:
        canonical_root = ROOT / "docs" / "test-design" / "current"
        canonical_root.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(dir=canonical_root) as value:
            run_dir = Path(value) / "adapter"
            result = execute_request("record", run_dir, {
                "kind": "page", "data": {
                    "name": "通用页面", "menu_path": ["工具", "通用页面"],
                    "final_scan_status": "stable", "unhandled_element_refs": [],
                },
            }, module_path="工具>通用页面")
            self.assertEqual(1, result["recorded"])
            self.assertEqual(execute_request("status", run_dir)["state"], pipeline_status(run_dir)["state"])

    def test_resume_drops_only_a_truncated_final_event_line(self) -> None:
        path = artifact_paths(self.run_dir)["events"]
        with path.open("a", encoding="utf-8", newline="") as handle:
            handle.write('{"kind":"element","data":')
        events_before = len(path.read_text(encoding="utf-8").splitlines())
        facts = compile_facts(self.run_dir)
        self.assertGreater(facts["event_count"], 0)
        self.assertEqual(events_before - 1, len(path.read_text(encoding="utf-8").splitlines()))

    def test_open_items_use_review_notes_or_real_fact_blocking_without_retry_loop(self) -> None:
        self._write_plan_and_cases()
        append_events(self.run_dir, [{"kind": "open_item", "fact_id": "OPEN-RISK", "data": {
            "category": "observed_risk", "description": "末页返回响应较慢", "material": False,
            "affected_function_refs": ["FN-VIEW"]
        }}])
        compile_facts(self.run_dir)
        review = review_run(self.run_dir, _semantic_review(self.run_dir))
        self.assertEqual("ready_with_notes", review["status"])
        append_events(self.run_dir, [{"kind": "open_item", "fact_id": "OPEN-QUESTION", "data": {
            "category": "external_question", "description": "告警确认后的外部处置语义待确认", "material": True,
            "affected_function_refs": ["FN-VIEW"]
        }}])
        compile_facts(self.run_dir)
        self.assertEqual("blocked_by_fact", review_run(self.run_dir, _semantic_review(self.run_dir))["status"])
        append_events(self.run_dir, [{"kind": "open_item", "fact_id": "OPEN-QUESTION", "status": "resolved", "data": {
            "category": "external_question", "description": "用户已确认外部处置语义", "material": True,
            "affected_function_refs": ["FN-VIEW"]
        }}])
        compile_facts(self.run_dir)
        self.assertEqual("ready_with_notes", review_run(self.run_dir, _semantic_review(self.run_dir))["status"])

    def test_delivery_detects_a_stale_review_without_returning_to_discovery(self) -> None:
        self._write_plan_and_cases()
        self.assertEqual("ready", review_run(self.run_dir, _semantic_review(self.run_dir))["status"])
        append_events(self.run_dir, [{"kind": "open_item", "fact_id": "OPEN-NOTE", "data": {
            "category": "observed_risk", "description": "非阻塞提示", "material": False,
            "affected_function_refs": ["FN-VIEW"]
        }}])
        compile_facts(self.run_dir)
        with self.assertRaisesRegex(ValueError, "stale"):
            complete_deliverables(self.run_dir, ROOT)

    def test_review_recomputes_discovery_instead_of_trusting_semantic_claims(self) -> None:
        self._write_plan_and_cases()
        append_events(self.run_dir, [{
            "kind": "element", "fact_id": "EL-VIEW-MODE",
            "data": {"type": "unclassified-widget"},
        }])
        compile_facts(self.run_dir)
        review = review_run(self.run_dir, _semantic_review(self.run_dir))
        self.assertEqual("blocked_by_fact", review["status"])
        self.assertIn("unclassified_interactive_element", [row["code"] for row in review["deterministic"]["issues"]])

    def test_configuration_covers_default_and_each_single_factor_value(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            run_dir = Path(value) / "config-run"
            ensure_run(run_dir, "告警管理>新增告警规则")
            checks = []
            for option in ["不配置", "邮件", "短信"]:
                checks.append({
                    "element_ref": "EL-CONFIG", "action": f"设置通知方式为{option}", "option_value": option,
                    "result": f"页面接受{option}", "result_anchor": {"assertion": "contains", "value": option}, "commit_result": "保存成功", "persistence_result": "重开后回显一致",
                    "effect_result": f"规则按{option}生效", "recovery_result": "恢复基线",
                })
            checks[0]["action"] = (
                f"先设置通知方式为{checks[1]['option_value']}，再切回{checks[0]['option_value']}"
            )
            append_events(run_dir, [
                {"kind": "page", "fact_id": "PAGE", "data": {"name": "新增告警规则", "menu_path": ["告警管理", "告警规则", "新增告警规则"], "final_scan_status": "stable", "unhandled_element_refs": []}},
                {"kind": "function", "fact_id": "FN-CONFIG", "data": {"name": "通知方式配置"}},
                {"kind": "test_object", "fact_id": "OBJ", "data": {"name": "AI_TEST_RULE_001", "owner": "current_run", "state": "cleaned"}},
                {"kind": "element", "fact_id": "EL-CONFIG", "data": {
                    "function_ref": "FN-CONFIG", "page_ref": "PAGE", "name": "通知方式", "interactive": True,
                    "configuration": True, "option_set": "finite", "options": ["不配置", "邮件", "短信"], "default_value": "不配置"
                }},
                {"kind": "transaction", "fact_id": "TX-CONFIG", "data": {
                    "function_ref": "FN-CONFIG", "element_refs": ["EL-CONFIG"], "transaction_type": "configuration",
                    "test_object_ref": "OBJ", "outcome": "success", "combination": False, "checks": checks
                }},
            ])
            _mark_final_scan(run_dir)
            compile_facts(run_dir)
            self.assertEqual([], validate_discovery(run_dir))

    def test_delivery_has_two_independent_workbooks_and_no_blank_case_rows(self) -> None:
        self._write_plan_and_cases()
        self.assertEqual("ready", review_run(self.run_dir, _semantic_review(self.run_dir))["status"])
        receipt = complete_deliverables(self.run_dir, ROOT)
        delivery = artifact_paths(self.run_dir)["delivery"]
        formal = Path(receipt["formal_workbook"])
        import_file = Path(receipt["import_workbook"])
        self.assertEqual(str(delivery.resolve()), receipt["delivery_dir"])
        self.assertEqual("正式测试设计.xlsx", formal.name)
        self.assertEqual("测试系统导入.xlsx", import_file.name)
        self.assertEqual({"正式测试设计.xlsx", "测试系统导入.xlsx"}, {path.name for path in delivery.iterdir()})
        self.assertTrue(all(count > 0 for count in receipt["sheet_rows"].values()))
        workbook = load_workbook(formal, data_only=False)
        formal_template = load_workbook(ROOT / "docs" / "test-design" / "codebuddy-test-design-template.xlsx", data_only=False)
        self.assertEqual(SHEETS, workbook.sheetnames)
        function_headers = {cell.value: cell.column for cell in workbook["功能测试用例"][1]}
        self.assertGreaterEqual(workbook["功能测试用例"].column_dimensions[get_column_letter(function_headers["操作步骤"])].width, 44)
        self.assertEqual(
            {"有条件"},
            {workbook["功能测试用例"].cell(row, function_headers["是否适合自动化"]).value for row in range(2, 9)},
        )
        imported_book = load_workbook(import_file, data_only=False)
        import_headers = {cell.value: cell.column for cell in imported_book[imported_book.sheetnames[0]][1]}
        self.assertGreaterEqual(imported_book[imported_book.sheetnames[0]].column_dimensions[get_column_letter(import_headers["测试步骤预期结果"])].width, 44)
        for name in SHEETS:
            self.assertEqual(
                len(formal_template[name].data_validations.dataValidation),
                len(workbook[name].data_validations.dataValidation),
            )
            self.assertEqual(set(formal_template[name].tables), set(workbook[name].tables))
        ws = workbook["功能测试用例"]
        self.assertEqual("TC-VIEW-001", ws.cell(2, 1).value)
        self.assertEqual("TC-VIEW-003", ws.cell(4, 1).value)
        self.assertEqual("TC-VIEW-007", ws.cell(8, 1).value)
        self.assertIsNone(ws.cell(9, 1).value)
        self.assertGreaterEqual(ws.row_dimensions[2].height, 40)
        written_first = load_cases(self.run_dir)["cases"][0]
        self.assertEqual(written_first["test_point"], ws.cell(2, function_headers["功能点"]).value)
        self.assertNotEqual("告警视图模式", ws.cell(2, function_headers["功能点"]).value)
        matrix = workbook["测试场景矩阵"]
        matrix_text = "\n".join(str(matrix.cell(row, column).value or "") for row in range(2, matrix.max_row + 1) for column in range(1, matrix.max_column + 1))
        self.assertNotIn("TX-VIEW", matrix_text)
        self.assertNotIn("EL-VIEW", matrix_text)
        matrix_headers = {cell.value: cell.column for cell in matrix[1]}
        self.assertEqual(
            ws.cell(2, function_headers["功能点"]).value,
            matrix.cell(2, matrix_headers["功能点"]).value,
        )
        self.assertEqual("P1", matrix.cell(4, matrix_headers["优先级"]).value)
        self.assertEqual("中", matrix.cell(4, matrix_headers["风险等级"]).value)
        self.assertEqual("视图模式：详细", matrix.cell(4, matrix_headers["输入数据/状态条件"]).value)
        self.assertIn("全部可见告警字段", matrix.cell(4, matrix_headers["观察点"]).value)
        coverage = workbook["页面元素覆盖清单"]
        coverage_text = "\n".join(str(coverage.cell(row, column).value or "") for row in range(2, coverage.max_row + 1) for column in range(1, coverage.max_column + 1))
        self.assertNotIn("DOM", coverage_text)
        self.assertNotIn("EL-VIEW", coverage_text)
        coverage_headers = {cell.value: cell.column for cell in coverage[1]}
        self.assertEqual("已覆盖", coverage.cell(2, coverage_headers["覆盖状态"]).value)
        self.assertNotIn("另有", coverage.cell(2, coverage_headers["交互方式"]).value)
        self.assertIn("精简；标准；详细；仅未确认；仅已确认；仅严重；全部", coverage.cell(2, coverage_headers["交互方式"]).value)
        self.assertNotIn("选择精简视图", coverage.cell(2, coverage_headers["交互方式"]).value)
        requirements = workbook["需求用户故事拆解"]
        requirement_headers = {cell.value: cell.column for cell in requirements[1]}
        for header in ("Story ID/需求 ID", "用户故事/需求描述", "角色", "业务价值", "验收标准", "业务规则"):
            self.assertTrue(str(requirements.cell(2, requirement_headers[header]).value or "").strip())
        self.assertEqual("选择不同视图并查看对应告警字段", requirements.cell(2, requirement_headers["用户故事/需求描述"]).value)
        overview_text = "\n".join(str(cell.value or "") for cell in workbook["测试设计总览"][2])
        self.assertNotIn("跨产物", overview_text)
        self.assertNotIn("结构化用例", overview_text)
        performance = workbook["性能测试设计"]
        self.assertEqual("PERF-N/A", performance.cell(2, 1).value)
        self.assertIn("不包含可独立定义指标", "\n".join(str(cell.value or "") for cell in performance[2]))
        risk = workbook["风险与待确认问题"]
        self.assertEqual("RISK-001", risk.cell(2, 1).value)
        self.assertIn("页面控件定位变化", "\n".join(str(cell.value or "") for cell in risk[2]))
        automation = workbook["自动化建议"]
        self.assertEqual(8, automation.max_row)
        automation_headers = {cell.value: cell.column for cell in automation[1]}
        self.assertEqual("高频稳定回归", automation.cell(2, automation_headers["自动化价值"]).value)
        import_book = load_workbook(import_file, data_only=False)
        import_template = load_workbook(ROOT / "docs" / "test-design" / "测试用例模板.xlsx", data_only=False)
        import_ws = import_book[import_book.sheetnames[0]]
        import_template_ws = import_template[import_template.sheetnames[0]]
        self.assertEqual(len(import_template_ws.data_validations.dataValidation), len(import_ws.data_validations.dataValidation))
        self.assertEqual(set(import_template_ws.tables), set(import_ws.tables))
        self.assertEqual(8, import_ws.max_row)
        self.assertGreaterEqual(import_ws.row_dimensions[2].height, 40)
        validated = subprocess.run(
            [sys.executable, str(ROOT / "scripts" / "validate-test-design-deliverable.py"),
             "--workbook", str(formal), "--import-workbook", str(import_file)],
            cwd=ROOT, capture_output=True, text=True, check=False,
        )
        self.assertEqual(0, validated.returncode, validated.stderr)
        status = pipeline_status(self.run_dir)
        self.assertEqual("completed", status["state"])
        self.assertEqual(str(formal.resolve()), status["deliverables"]["formal_workbook"])


if __name__ == "__main__":
    unittest.main()
