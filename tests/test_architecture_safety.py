# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import importlib.util
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
FRAMEWORK_VERSION = next(
    line.split("=", 1)[1].strip()
    for line in (REPO_ROOT / "VERSION").read_text(encoding="utf-8-sig").splitlines()
    if line.startswith("framework_version=")
)
MODULE_PATH = REPO_ROOT / "scripts" / "test_design_excel_tools.py"
sys.path.insert(0, str(REPO_ROOT / "scripts"))
from test_design.contracts.sheet_data import SHEET_DATA_HEADERS

SPEC = importlib.util.spec_from_file_location("test_design_excel_tools", MODULE_PATH)
assert SPEC and SPEC.loader
TOOLS = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(TOOLS)
SYNC_MODULE_PATH = REPO_ROOT / "scripts" / "sync-rule-entrypoints.py"
SYNC_SPEC = importlib.util.spec_from_file_location("sync_rule_entrypoints", SYNC_MODULE_PATH)
assert SYNC_SPEC and SYNC_SPEC.loader
SYNC_TOOLS = importlib.util.module_from_spec(SYNC_SPEC)
SYNC_SPEC.loader.exec_module(SYNC_TOOLS)
RUN_VALIDATION_PATH = REPO_ROOT / "scripts" / "run-validation.py"
RUN_VALIDATION_SPEC = importlib.util.spec_from_file_location("run_validation", RUN_VALIDATION_PATH)
assert RUN_VALIDATION_SPEC and RUN_VALIDATION_SPEC.loader
RUN_VALIDATION = importlib.util.module_from_spec(RUN_VALIDATION_SPEC)
RUN_VALIDATION_SPEC.loader.exec_module(RUN_VALIDATION)


class ArchitectureSafetyTests(unittest.TestCase):
    def test_full_validation_accepts_platform_only_upgrade_skips_on_linux(self) -> None:
        upgrade_test = mock.Mock()
        upgrade_test.id.return_value = "suite.test_upgrade_failure_restores_framework_and_protected_assets"
        result = mock.Mock(skipped=[(upgrade_test, "PowerShell integration runs on Windows")])
        self.assertEqual([], RUN_VALIDATION.skipped_required_upgrade_tests(result, "posix"))
        self.assertEqual([upgrade_test.id()], RUN_VALIDATION.skipped_required_upgrade_tests(result, "nt"))

    def test_entry_contract_rejects_incomplete_runtime_graph(self) -> None:
        with self.assertRaisesRegex(ValueError, "exactly codex and codebuddy"):
            SYNC_TOOLS.validate_runtime_graphs({}, ".codebuddy/rules/test-design-rule.md", ".codebuddy/.rules/test-design-rule.mdc")

    def test_generated_block_rewrite_preserves_local_override(self) -> None:
        begin = "<!-- TEST-DESIGN-GENERATED:BEGIN -->"
        end = "<!-- TEST-DESIGN-GENERATED:END -->"
        source = f"head\n{begin}\nold\n{end}\n<!-- LOCAL-OVERRIDES:BEGIN -->\nLOCAL_KEEP\n<!-- LOCAL-OVERRIDES:END -->\n"
        rewritten = SYNC_TOOLS.replace_generated_block(source, begin, end, "new")
        self.assertIn("LOCAL_KEEP", rewritten)
        self.assertIn(f"{begin}\nnew\n{end}", rewritten)

    def test_formal_assembler_populates_all_sheets_and_removes_template_examples(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            root = Path(value)
            run_dir = root / "run"
            data_dir = run_dir / "artifacts" / "data"
            data_dir.mkdir(parents=True)
            template = REPO_ROOT / "docs" / "test-design" / "codebuddy-test-design-template.xlsx"
            workbook = load_workbook(template, data_only=False)
            sources = {
                "测试设计总览": "overview.json",
                "需求用户故事拆解": "requirements.json",
                "测试场景矩阵": "scenarios.json",
                "性能测试设计": "performance.json",
                "风险与待确认问题": "risks.json",
                "自动化建议": "automation.json",
                "页面元素覆盖清单": "page_elements.json",
            }
            for sheet_name, filename in sources.items():
                headers = [cell.value for cell in workbook[sheet_name][1] if cell.value]
                row = {str(header): f"ASSEMBLED-{index + 1}" for index, header in enumerate(headers)}
                (data_dir / filename).write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")

            function_headers = [cell.value for cell in workbook["功能测试用例"][1] if cell.value]
            function_row = {str(header): f"ASSEMBLED-{index + 1}" for index, header in enumerate(function_headers)}
            function_row["用例 ID"] = "TC-ASSEMBLED-001"
            function_row["Story ID/需求 ID"] = "REQ-ASSEMBLED-001"
            function_row["用例标题"] = "组装验证-正式用例"
            part_name = "function_cases_part_001.json"
            (data_dir / part_name).write_text(json.dumps([function_row], ensure_ascii=False), encoding="utf-8")
            (data_dir / "function_cases_manifest.json").write_text(
                json.dumps({"part_size": 10, "total_cases": 1, "parts": [part_name]}, ensure_ascii=False),
                encoding="utf-8",
            )

            output = root / "assembled.xlsx"
            counts = TOOLS.assemble_formal_workbook(run_dir, template, output)
            self.assertEqual(1, counts["功能测试用例"])
            assembled = load_workbook(output, data_only=True)
            self.assertEqual("TC-ASSEMBLED-001", assembled["功能测试用例"]["A2"].value)
            for sheet in assembled.worksheets:
                combined = "".join(str(cell.value or "") for row in sheet.iter_rows(min_row=2) for cell in row)
                self.assertNotIn("TC-LOGIN-001", combined)
                self.assertNotIn("示例项目", combined)

    def test_deliverable_validator_rejects_unmodified_formal_template(self) -> None:
        template = REPO_ROOT / "docs" / "test-design" / "codebuddy-test-design-template.xlsx"
        result = subprocess.run(
            [
                sys.executable,
                str(REPO_ROOT / "scripts" / "validate-test-design-deliverable.py"),
                "--workbook",
                str(template),
            ],
            cwd=REPO_ROOT,
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        self.assertNotEqual(0, result.returncode)
        self.assertIn("formal-template example marker", result.stderr + result.stdout)

    def create_project_root(self, root: Path) -> None:
        source = REPO_ROOT / "docs" / "test-assets" / "batch-runs" / "templates"
        target = root / "docs" / "test-assets" / "batch-runs" / "templates"
        shutil.copytree(source, target)

    def write_csv_rows(self, path: Path, rows: list[dict[str, str]]) -> None:
        with path.open("r", encoding="utf-8-sig", newline="") as stream:
            headers = next(csv.reader(stream))
        with path.open("w", encoding="utf-8-sig", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=headers)
            writer.writeheader()
            for values in rows:
                row = {header: "" for header in headers}
                row.update(values)
                writer.writerow(row)

    def write_valid_sheet_files(self, data_dir: Path) -> None:
        for filename, headers in SHEET_DATA_HEADERS.items():
            row = {header: "正式数据" for header in headers}
            (data_dir / filename).write_text(json.dumps([row], ensure_ascii=False), encoding="utf-8")

    def make_valid_plan_run(self, project_root: Path, run_id: str = "risk-probe") -> Path:
        run_dir = TOOLS.init_batch_run(project_root, run_id, "产品>模块>页面", "BATCH-001", "产品")
        self.write_csv_rows(
            run_dir / "page-discovery.csv",
            [
                {
                    "批次ID": "BATCH-001",
                    "最小标题路径": "模块>页面",
                    "页面/入口": "风险页面",
                    "元素名称/文案": "危险操作按钮",
                    "元素类型": "按钮",
                    "交互方式": "点击",
                    "完整点击路径": "系统>模块>页面>危险操作按钮",
                    "预期/观察行为": "打开确认弹窗",
                    "适用DFX维度": "DFT功能",
                    "适用DFX场景": "正向流程",
                }
            ],
        )
        self.write_csv_rows(
            run_dir / "element-case-plan.csv",
            [
                {
                    "批次ID": "BATCH-001",
                    "最小标题路径": "模块>页面",
                    "页面/入口": "风险页面",
                    "功能点": "危险操作",
                    "元素名称/文案": "危险操作按钮",
                    "元素类型": "按钮",
                    "交互方式": "点击",
                    "适用DFX维度": "DFT功能",
                    "适用DFX场景": "正向流程",
                    "测试设计方向": "打开并关闭确认弹窗",
                    "应生成用例数": "3",
                    "计划用例ID": "TC-RISK-001,TC-RISK-002,TC-RISK-003",
                    "操作类别": "查看",
                    "验证要求": "结果分支",
                    "数据策略": "无数据变更",
                    "执行状态": "已完成",
                    "是否必须真实执行": "是",
                    "是否涉及配置生效": "否",
                    "是否涉及CRUD闭环": "否",
                }
            ],
        )
        status_path = run_dir / "batch-status.csv"
        with status_path.open("r", encoding="utf-8-sig", newline="") as stream:
            status = next(csv.DictReader(stream))
        status.update({"状态": "执行中", "页面数": "1", "元素总数": "1", "已覆盖元素数": "1"})
        self.write_csv_rows(status_path, [status])
        return run_dir

    def function_case(self, case_id: str) -> dict[str, str]:
        suffix = int(case_id.rsplit("-", 1)[-1])
        interaction, outcome = [
            ("点击弹窗右上角关闭按钮", "弹窗关闭且原页面数据不变"),
            ("点击弹窗取消按钮", "弹窗取消并返回原页面且数据不变"),
            ("按 Esc 键关闭弹窗", "弹窗关闭且焦点返回危险操作按钮"),
            ("点击弹窗遮罩区域", "弹窗按遮罩策略关闭且原页面数据不变"),
            ("用 Tab 聚焦取消按钮后按 Enter", "弹窗通过键盘取消并返回原页面"),
            ("用 Tab 聚焦关闭按钮后按 Space", "弹窗通过键盘关闭且焦点顺序保持正确"),
            ("缩小浏览器窗口后点击关闭按钮", "小窗口下弹窗仍可关闭且页面布局未错乱"),
            ("切换浏览器页签再返回并点击取消", "页签切换后弹窗状态保留且可正常取消"),
            ("等待三秒后点击关闭按钮", "等待期间弹窗保持稳定且随后正常关闭"),
            ("连续双击关闭按钮", "弹窗只关闭一次且页面无重复操作或报错"),
        ][(suffix - 1) % 10]
        return {
            "用例 ID": case_id,
            "Story ID/需求 ID": "REQ-001",
            "模块": "模块>页面",
            "功能点": "危险操作",
            "用例标题": f"危险操作-{case_id}确认弹窗",
            "优先级": "P1",
            "测试类型": "功能测试",
            "DFX维度": "DFT功能",
            "DFX场景": "正向流程",
            "前置条件": "1. 已获得测试账号\n2. 系统服务可用",
            "测试数据": "无数据变更",
            "操作步骤": (
                "1. 打开系统登录入口\n2. 登录后进入模块菜单\n3. 打开风险页面\n"
                f"4. 点击危险操作按钮后{interaction}"
            ),
            "预期结果": (
                "1. 登录后显示模块导航\n2. 风险页面加载完成\n"
                f"3. {outcome}"
            ),
            "实际结果": "",
            "执行状态": "未执行",
            "是否适合自动化": "否",
            "关联风险": "",
            "备注": "",
        }

    def test_discovery_delete_requires_tagged_data_and_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = TOOLS.init_batch_run(project_root, "delete-evidence", "产品>模块>页面", "BATCH-001", "产品")
            row = {
                "批次ID": "BATCH-001",
                "最小标题路径": "模块>页面",
                "页面/入口": "列表页面",
                "元素名称/文案": "删除按钮",
                "元素类型": "按钮",
                "交互方式": "点击删除并确认",
                "完整点击路径": "系统>模块>列表页面>删除按钮>确认",
                "预期/观察行为": "删除成功且记录不再显示",
                "结果分支/后续状态": "列表刷新并持久化删除结果",
                "适用DFX维度": "DFT功能",
                "适用DFX场景": "正向流程",
            }
            self.write_csv_rows(run_dir / "page-discovery.csv", [row])
            with self.assertRaisesRegex(ValueError, "tagged test data"):
                TOOLS.validate_batch_artifacts(run_dir, "discovery")

            row["测试数据来源"] = "本次创建 CODEX_TEST_DELETE_001"
            self.write_csv_rows(run_dir / "page-discovery.csv", [row])
            with self.assertRaisesRegex(ValueError, "existing evidence"):
                TOOLS.validate_batch_artifacts(run_dir, "discovery")

            evidence = run_dir / "artifacts" / "screenshots" / "delete-success.txt"
            evidence.write_text("CODEX_TEST_DELETE_001 删除成功并从列表消失", encoding="utf-8")
            row["证据路径"] = "artifacts/screenshots/delete-success.txt"
            self.write_csv_rows(run_dir / "page-discovery.csv", [row])
            TOOLS.validate_batch_artifacts(run_dir, "discovery")

    def test_icon_interaction_and_structured_mutation_cannot_escape_plan_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "icon-mutation")
            discovery_path = run_dir / "page-discovery.csv"
            with discovery_path.open("r", encoding="utf-8-sig", newline="") as stream:
                discovery = next(csv.DictReader(stream))
            discovery.update(
                {
                    "元素名称/文案": "缓存动作图标",
                    "元素类型": "图标",
                    "交互方式": "单击",
                    "完整点击路径": "系统>模块>风险页面>缓存动作图标",
                    "预期/观察行为": "操作成功并持久化",
                    "结果分支/后续状态": "关联页面状态更新",
                }
            )
            self.write_csv_rows(discovery_path, [discovery])

            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                plan = next(csv.DictReader(stream))
            ids = [f"TC-ICON-{index:03d}" for index in range(1, 6)]
            plan.update(
                {
                    "功能点": "缓存动作",
                    "元素名称/文案": "缓存动作图标",
                    "元素类型": "图标",
                    "交互方式": "单击",
                    "测试设计方向": "验证状态更新、回显和实际生效",
                    "应生成用例数": "5",
                    "计划用例ID": ",".join(ids),
                    "操作类别": "状态变更",
                    "验证要求": "回显,持久化,实际生效",
                    "数据策略": "本次创建测试数据",
                    "执行状态": "已完成",
                    "是否涉及CRUD闭环": "是",
                }
            )
            self.write_csv_rows(plan_path, [plan])
            lifecycle = {
                "批次ID": "BATCH-001",
                "最小标题路径": "模块>页面",
                "关联页面/入口": "风险页面",
                "修改项/元素": "缓存动作图标",
                "测试数据ID/名称": "CODEX_TEST_ICON_001",
                "数据类型": "状态变更测试数据",
                "创建入口": "风险页面",
                "创建结果": "创建成功",
                "查看结果": "详情回显成功",
                "编辑前值": "状态A",
                "编辑后值": "状态B",
                "编辑结果": "状态更新成功",
                "保存后回显": "重新进入仍显示状态B",
                "实际生效结果": "关联页面按状态B生效",
                "清理状态": "待清理",
            }
            self.write_csv_rows(run_dir / "test-data-lifecycle.csv", [lifecycle])
            TOOLS.validate_batch_artifacts(run_dir, "discovery")
            with self.assertRaisesRegex(ValueError, "mutation discovery must use tagged test data"):
                TOOLS.validate_batch_artifacts(run_dir, "plan")

            evidence = run_dir / "artifacts/screenshots/icon-state.txt"
            evidence.write_text("CODEX_TEST_ICON_001 状态更新成功并生效", encoding="utf-8")
            discovery.update(
                {
                    "测试数据来源": "CODEX_TEST_ICON_001",
                    "证据路径": "artifacts/screenshots/icon-state.txt",
                }
            )
            self.write_csv_rows(discovery_path, [discovery])
            TOOLS.validate_batch_artifacts(run_dir, "plan")

    def test_batch_init_requires_explicit_resume_and_preserves_ledgers(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = TOOLS.init_batch_run(project_root, "probe", "产品>模块>页面", "BATCH-001")
            status_path = run_dir / "batch-status.csv"
            with status_path.open("r", encoding="utf-8-sig", newline="") as fp:
                reader = csv.DictReader(fp)
                headers = reader.fieldnames or []
                rows = list(reader)
            rows[0]["状态"] = "执行中"
            rows[0]["页面数"] = "3"
            with status_path.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=headers)
                writer.writeheader()
                writer.writerows(rows)

            with self.assertRaisesRegex(ValueError, "already exists"):
                TOOLS.init_batch_run(project_root, "probe", "产品>模块>页面", "BATCH-001")
            (run_dir / "risk-confirmation.csv").unlink()
            TOOLS.init_batch_run(project_root, "probe", "产品>模块>页面", "BATCH-001", resume=True)

            with status_path.open("r", encoding="utf-8-sig", newline="") as fp:
                resumed = next(csv.DictReader(fp))
            self.assertEqual("执行中", resumed["状态"])
            self.assertEqual("3", resumed["页面数"])
            self.assertTrue((run_dir / "risk-confirmation.csv").exists())

            TOOLS.init_batch_run(
                project_root,
                "probe",
                "产品>模块>页面",
                "BATCH-001",
                force_reinitialize=True,
            )
            with status_path.open("r", encoding="utf-8-sig", newline="") as fp:
                reset = next(csv.DictReader(fp))
            self.assertEqual("待开始", reset["状态"])
            self.assertEqual("0", reset["页面数"])
            self.assertEqual(1, len(list(run_dir.parent.glob("probe_backup_*"))))

            custom = TOOLS.init_batch_run(project_root, "custom-id", "产品>模块>页面", "BATCH-007")
            for filename in ["batch-status.csv", "page-discovery.csv", "element-case-plan.csv", "test-data-lifecycle.csv", "risk-confirmation.csv"]:
                with (custom / filename).open("r", encoding="utf-8-sig", newline="") as stream:
                    self.assertEqual("BATCH-007", next(csv.DictReader(stream))["批次ID"])

    def test_batch_run_rejects_multiple_batch_rows_and_cross_batch_ledgers(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "single-batch-scope")
            status_path = run_dir / "batch-status.csv"
            with status_path.open("r", encoding="utf-8-sig", newline="") as stream:
                status = next(csv.DictReader(stream))
            second = dict(status)
            second.update({"批次ID": "BATCH-002", "最小标题路径": "模块>另一个页面"})
            self.write_csv_rows(status_path, [status, second])
            with self.assertRaisesRegex(ValueError, "exactly one batch-status.csv row"):
                TOOLS.validate_batch_artifacts(run_dir, "discovery")

            self.write_csv_rows(status_path, [status])
            discovery_path = run_dir / "page-discovery.csv"
            with discovery_path.open("r", encoding="utf-8-sig", newline="") as stream:
                discovery = next(csv.DictReader(stream))
            discovery["批次ID"] = "BATCH-002"
            self.write_csv_rows(discovery_path, [discovery])
            with self.assertRaisesRegex(ValueError, "do not mix multiple leaf batches"):
                TOOLS.validate_batch_artifacts(run_dir, "discovery")

            discovery.update({"批次ID": "BATCH-001", "最小标题路径": "模块>另一个页面"})
            self.write_csv_rows(discovery_path, [discovery])
            with self.assertRaisesRegex(ValueError, "does not match the run leaf"):
                TOOLS.validate_batch_artifacts(run_dir, "discovery")

    def test_resume_migrates_legacy_risk_driven_deep_dive_ledger(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = TOOLS.init_batch_run(project_root, "legacy-risk", "产品>模块>页面", "BATCH-001")
            risk_path = run_dir / "risk-confirmation.csv"
            old_headers = [
                "批次ID", "风险ID", "风险/待确认问题", "用户确认结论", "处置策略", "是否需要补充深探",
                "补充深探目标", "关联页面/入口", "关联元素名称/文案", "补充证据路径", "补充深探状态",
                "关联用例ID", "备注",
            ]
            with risk_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=old_headers)
                writer.writeheader()
                writer.writerow({"批次ID": "BATCH-001", "风险ID": "RISK-001", "风险/待确认问题": "旧问题"})

            TOOLS.init_batch_run(project_root, "legacy-risk", "产品>模块>页面", "BATCH-001", resume=True)

            with risk_path.open("r", encoding="utf-8-sig", newline="") as stream:
                row = next(csv.DictReader(stream))
            self.assertEqual("旧问题", row["模型不理解内容/待确认问题"])
            self.assertEqual("待确认", row["确认状态"])
            self.assertEqual(1, len(list(run_dir.glob("risk-confirmation.pre-default-deep-dive-*.csv"))))

    def test_resume_migrates_structured_plan_and_lifecycle_ledgers_with_backups(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = TOOLS.init_batch_run(project_root, "legacy-structured", "产品>模块>页面", "BATCH-001")
            plan_path = run_dir / "element-case-plan.csv"
            lifecycle_path = run_dir / "test-data-lifecycle.csv"
            removed_plan = {"操作类别", "验证要求", "数据策略", "执行状态"}
            removed_lifecycle = {"关联页面/入口", "修改项/元素", "保存后回显", "实际生效结果"}
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                old_plan_headers = [header for header in (reader.fieldnames or []) if header not in removed_plan]
                old_plan = next(reader)
            with plan_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=old_plan_headers)
                writer.writeheader()
                writer.writerow({header: old_plan.get(header, "") for header in old_plan_headers})
            with lifecycle_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                old_lifecycle_headers = [header for header in (reader.fieldnames or []) if header not in removed_lifecycle]
                old_lifecycle = next(reader)
            with lifecycle_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=old_lifecycle_headers)
                writer.writeheader()
                writer.writerow({header: old_lifecycle.get(header, "") for header in old_lifecycle_headers})

            TOOLS.init_batch_run(project_root, "legacy-structured", "产品>模块>页面", "BATCH-001", resume=True)

            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                migrated_plan = next(csv.DictReader(stream))
            self.assertIn(migrated_plan["操作类别"], {"查看", "其他"})
            self.assertEqual("待执行", migrated_plan["执行状态"])
            self.assertEqual(1, len(list(run_dir.glob("element-case-plan.pre-structured-ledger-*.csv"))))
            self.assertEqual(1, len(list(run_dir.glob("test-data-lifecycle.pre-structured-ledger-*.csv"))))

    def test_resume_rejects_unknown_damaged_ledger_header(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = TOOLS.init_batch_run(project_root, "damaged-header", "产品>模块>页面", "BATCH-001")
            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                headers = [header for header in (reader.fieldnames or []) if header != "功能点"]
                row = next(reader)
            with plan_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=headers)
                writer.writeheader()
                writer.writerow({header: row.get(header, "") for header in headers})
            with self.assertRaisesRegex(ValueError, "unsupported element-case-plan.csv header"):
                TOOLS.init_batch_run(project_root, "damaged-header", "产品>模块>页面", "BATCH-001", resume=True)

    def test_plan_gate_requires_confirmed_model_uncertainty(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root)

            TOOLS.validate_batch_artifacts(run_dir, "plan")
            with self.assertRaisesRegex(ValueError, "risk-confirmation.csv is not ready"):
                TOOLS.validate_batch_artifacts(run_dir, "risk")

    def test_risk_none_is_system_recorded_and_prepare_cleans_all_stale_case_data(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "risk-none")
            data_dir = run_dir / "artifacts" / "data"
            stale_names = [
                "function_cases_part_001.json", "function_cases_manifest.json", "overview.json", "requirements.json",
                "scenarios.json", "performance.json", "risks.json", "automation.json", "page_elements.json",
            ]
            for name in stale_names:
                (data_dir / name).write_text("{}", encoding="utf-8")

            with self.assertRaisesRegex(ValueError, "risk-confirmation.csv is not ready"):
                TOOLS.prepare_function_case_generation(run_dir)
            self.assertTrue(all((data_dir / name).exists() for name in stale_names))

            TOOLS.record_no_model_uncertainty(run_dir)
            with (run_dir / "risk-confirmation.csv").open("r", encoding="utf-8-sig", newline="") as stream:
                row = next(csv.DictReader(stream))
            self.assertEqual("RISK-NONE", row["风险ID"])
            self.assertEqual("无需用户确认", row["确认状态"])
            TOOLS.validate_batch_artifacts(run_dir, "risk")
            TOOLS.prepare_function_case_generation(run_dir)
            self.assertTrue(all(not (data_dir / name).exists() for name in stale_names))

    def test_pipeline_status_derives_risk_assessment_without_manual_status(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "pipeline")
            status = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("RISK_ASSESSMENT_REQUIRED", status["state"])
            self.assertIn("record-risk-none", status["command"])
            TOOLS.record_no_model_uncertainty(run_dir)
            status = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("CASE_PREPARATION_REQUIRED", status["state"])
            TOOLS.prepare_function_case_generation(run_dir)
            status = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("CASE_GENERATION_REQUIRED", status["state"])

    def test_plan_uses_page_scoped_element_identity(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "page-scoped")
            discovery_path = run_dir / "page-discovery.csv"
            with discovery_path.open("r", encoding="utf-8-sig", newline="") as stream:
                first = next(csv.DictReader(stream))
            second = dict(first)
            second["页面/入口"] = "另一个页面"
            self.write_csv_rows(discovery_path, [first, second])
            with self.assertRaisesRegex(ValueError, "missing interactive page elements"):
                TOOLS.validate_batch_artifacts(run_dir, "plan")

    def test_plan_cache_invalidates_when_structured_plan_changes(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "cache")
            TOOLS.validate_batch_artifacts(run_dir, "plan", use_cache=True)
            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                row = next(csv.DictReader(stream))
            row["执行状态"] = "待执行"
            self.write_csv_rows(plan_path, [row])
            with self.assertRaisesRegex(ValueError, "执行状态=已完成"):
                TOOLS.validate_batch_artifacts(run_dir, "plan", use_cache=True)

    def test_discovery_cache_tracks_project_relative_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = TOOLS.init_batch_run(project_root, "evidence", "产品>模块>页面", "BATCH-001", "产品")
            evidence = run_dir / "artifacts" / "screenshots" / "proof.txt"
            evidence.write_text("proof", encoding="utf-8")
            relative_evidence = evidence.relative_to(project_root).as_posix()
            self.write_csv_rows(
                run_dir / "page-discovery.csv",
                [
                    {
                        "批次ID": "BATCH-001",
                        "最小标题路径": "模块>页面",
                        "页面/入口": "编辑页面",
                        "元素名称/文案": "保存按钮",
                        "元素类型": "按钮",
                        "交互方式": "点击",
                        "适用DFX维度": "DFT功能",
                        "适用DFX场景": "正向流程",
                        "结果分支/后续状态": "保存成功并返回详情",
                        "完整点击路径": "系统>模块>编辑页面>保存按钮",
                        "预期/观察行为": "保存成功并持久化回显",
                        "测试数据来源": "AI_TEST_EVIDENCE_001",
                        "证据路径": relative_evidence,
                    }
                ],
            )
            TOOLS.validate_batch_artifacts(run_dir, "discovery", use_cache=True)
            TOOLS.validate_batch_artifacts(run_dir, "discovery", use_cache=True)
            evidence.unlink()
            with self.assertRaisesRegex(ValueError, "reference existing evidence"):
                TOOLS.validate_batch_artifacts(run_dir, "discovery", use_cache=True)

    def test_risk_none_cannot_be_mixed_with_real_uncertainty(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "risk-exclusive")
            TOOLS.record_no_model_uncertainty(run_dir)
            risk_path = run_dir / "risk-confirmation.csv"
            with risk_path.open("r", encoding="utf-8-sig", newline="") as stream:
                none_row = next(csv.DictReader(stream))
            real_row = dict(none_row)
            real_row.update(
                {
                    "风险ID": "RISK-001",
                    "模型不理解内容/待确认问题": "审批是否异步",
                    "用户确认结论": "是",
                    "确认状态": "已确认",
                }
            )
            self.write_csv_rows(risk_path, [none_row, real_row])
            with self.assertRaisesRegex(ValueError, "RISK-NONE must be the only"):
                TOOLS.validate_batch_artifacts(run_dir, "risk")

    def test_risk_none_requires_explicit_non_user_confirmation_semantics(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "risk-none-semantics")
            TOOLS.record_no_model_uncertainty(run_dir)
            risk_path = run_dir / "risk-confirmation.csv"
            with risk_path.open("r", encoding="utf-8-sig", newline="") as stream:
                row = next(csv.DictReader(stream))
            row["用户确认结论"] = "待用户确认"
            self.write_csv_rows(risk_path, [row])
            with self.assertRaisesRegex(ValueError, "用户确认结论=无需用户确认"):
                TOOLS.validate_batch_artifacts(run_dir, "risk")

            row["用户确认结论"] = "无需用户确认"
            row["确认状态"] = "已确认"
            self.write_csv_rows(risk_path, [row])
            with self.assertRaisesRegex(ValueError, "确认状态=无需用户确认"):
                TOOLS.validate_batch_artifacts(run_dir, "risk")

    def test_record_risk_none_refuses_to_overwrite_real_uncertainty(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "risk-overwrite")
            risk_path = run_dir / "risk-confirmation.csv"
            real = {
                "批次ID": "BATCH-001",
                "风险ID": "RISK-001",
                "模型不理解内容/待确认问题": "审批是否异步",
                "已完成深探依据": "页面未展示审批规则",
                "用户确认结论": "待用户确认",
                "处置策略": "待确认",
                "是否阻塞用例设计": "是",
                "确认状态": "待确认",
            }
            self.write_csv_rows(risk_path, [real])
            with self.assertRaisesRegex(ValueError, "refuses to overwrite"):
                TOOLS.record_no_model_uncertainty(run_dir)
            with risk_path.open("r", encoding="utf-8-sig", newline="") as stream:
                self.assertEqual("RISK-001", next(csv.DictReader(stream))["风险ID"])

    def test_pipeline_rejects_confirmed_but_incomplete_risk_ledger(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "risk-incomplete")
            self.write_csv_rows(
                run_dir / "risk-confirmation.csv",
                [
                    {
                        "批次ID": "BATCH-001",
                        "风险ID": "RISK-001",
                        "用户确认结论": "异步审批",
                        "是否阻塞用例设计": "否",
                        "确认状态": "已确认",
                    }
                ],
            )
            status = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("DISCOVERY_REQUIRED", status["state"])

    def test_generation_session_becomes_stale_when_plan_semantics_change(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "stale-session")
            skill = project_root / ".codebuddy/skills/test-design/SKILL.md"
            skill.parent.mkdir(parents=True)
            skill.write_text("skill-version-1", encoding="utf-8")
            catalog_index = project_root / "docs/test-assets/catalog/index.json"
            catalog_index.parent.mkdir(parents=True)
            catalog_index.write_text('{"version":1}', encoding="utf-8")
            TOOLS.record_no_model_uncertainty(run_dir)
            TOOLS.prepare_function_case_generation(run_dir)
            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                plan = next(csv.DictReader(stream))
            plan["测试设计方向"] += "并验证审计记录"
            self.write_csv_rows(plan_path, [plan])
            status = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("CASE_PREPARATION_REQUIRED", status["state"])

            TOOLS.prepare_function_case_generation(run_dir)
            skill.write_text("skill-version-2", encoding="utf-8")
            self.assertFalse(TOOLS.generation_session_is_current(run_dir))

            TOOLS.prepare_function_case_generation(run_dir)
            catalog_index.write_text('{"version":2}', encoding="utf-8")
            self.assertFalse(TOOLS.generation_session_is_current(run_dir))

    def test_generation_result_backfills_keep_session_current_through_complete_status(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "result-backfill")
            product_map = project_root / "docs/test-assets/product-map.xlsx"
            product_map.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(REPO_ROOT / "docs/test-assets/product-map.xlsx", product_map)
            ids = ["TC-RISK-001", "TC-RISK-002", "TC-RISK-003"]
            TOOLS.record_no_model_uncertainty(run_dir)
            TOOLS.prepare_function_case_generation(run_dir)

            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                plan = next(csv.DictReader(stream))
            plan["实际用例ID"] = ",".join(ids)
            self.write_csv_rows(plan_path, [plan])

            discovery_path = run_dir / "page-discovery.csv"
            with discovery_path.open("r", encoding="utf-8-sig", newline="") as stream:
                discovery = next(csv.DictReader(stream))
            discovery.update({"是否已生成用例": "是", "关联用例ID": ",".join(ids), "覆盖状态": "已覆盖"})
            self.write_csv_rows(discovery_path, [discovery])

            lifecycle_path = run_dir / "test-data-lifecycle.csv"
            with lifecycle_path.open("r", encoding="utf-8-sig", newline="") as stream:
                lifecycle = next(csv.DictReader(stream))
            lifecycle["创建步骤关联用例"] = ids[0]
            self.write_csv_rows(lifecycle_path, [lifecycle])

            status_path = run_dir / "batch-status.csv"
            with status_path.open("r", encoding="utf-8-sig", newline="") as stream:
                status = next(csv.DictReader(stream))
            status["功能用例数"] = "3"
            self.write_csv_rows(status_path, [status])

            data_dir = run_dir / "artifacts" / "data"
            session = json.loads((data_dir / "generation-session.json").read_text(encoding="utf-8"))
            (data_dir / "function_cases_part_001.json").write_text(
                json.dumps([self.function_case(case_id) for case_id in ids], ensure_ascii=False), encoding="utf-8"
            )
            (data_dir / "function_cases_manifest.json").write_text(
                json.dumps(
                    {
                        "part_size": 10,
                        "total_cases": 3,
                        "parts": ["function_cases_part_001.json"],
                        "generation_session_id": session["generation_session_id"],
                        "source_fingerprint": session["source_fingerprint"],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            self.write_valid_sheet_files(data_dir)
            TOOLS.validate_batch_artifacts(run_dir, "cases")

            formal_rel = Path("docs/test-assets/modules/模块_页面_测试设计.xlsx")
            import_rel = Path("docs/test-assets/imports/模块_页面_导入用例.xlsx")
            published = [
                project_root / formal_rel,
                project_root / import_rel,
                project_root / "docs/test-design/current" / formal_rel.name,
                project_root / "docs/test-design/deliverables" / formal_rel.name,
                project_root / "docs/test-design/deliverables" / import_rel.name,
            ]
            formal_template = REPO_ROOT / "docs/test-design/codebuddy-test-design-template.xlsx"
            import_template = REPO_ROOT / "docs/test-design/测试用例模板.xlsx"
            for index, path in enumerate(published):
                path.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(import_template if index in {1, 4} else formal_template, path)
            synced_map = load_workbook(product_map)
            synced_map["产品模块地图"]["A2"] = "交付阶段同步后的产品事实"
            synced_map.save(product_map)
            TOOLS.validate_catalog(product_map)
            TOOLS.write_delivery_receipt(
                project_root,
                status_path,
                published,
                product_map,
                discovery_path,
                "模块>页面",
                None,
            )
            status.update(
                {
                    "状态": "已完成",
                    "归档路径": formal_rel.as_posix(),
                    "导入文件路径": import_rel.as_posix(),
                    "导入文件已生成": "是",
                    "产品版图已更新": "是",
                    "覆盖质量自检": "通过",
                }
            )
            self.write_csv_rows(status_path, [status])
            self.assertEqual("COMPLETE", TOOLS.derive_pipeline_status(run_dir)["state"])

            delivered_map_bytes = product_map.read_bytes()
            externally_changed_map = load_workbook(product_map)
            externally_changed_map["产品模块地图"]["A2"] = "交付后出现的新产品事实"
            externally_changed_map.save(product_map)
            self.assertFalse(TOOLS.generation_session_is_current(run_dir))
            with self.assertRaisesRegex(ValueError, "catalog source fingerprint is stale"):
                TOOLS.validate_batch_artifacts(run_dir, "cases", use_cache=True)
            self.assertEqual("CASE_PREPARATION_REQUIRED", TOOLS.derive_pipeline_status(run_dir)["state"])
            product_map.write_bytes(delivered_map_bytes)
            self.assertEqual("COMPLETE", TOOLS.derive_pipeline_status(run_dir)["state"])

            published[0].write_bytes(b"tampered")
            tampered = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("DELIVERY_REQUIRED", tampered["state"])
            self.assertIn("changed since validation", tampered["reasons"][0])

    def test_pipeline_does_not_trust_manual_complete_flags_without_files(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "false-complete")
            TOOLS.record_no_model_uncertainty(run_dir)
            (run_dir / "artifacts" / "data" / "function_cases_manifest.json").write_text("{}", encoding="utf-8")
            status_path = run_dir / "batch-status.csv"
            with status_path.open("r", encoding="utf-8-sig", newline="") as stream:
                status_row = next(csv.DictReader(stream))
            status_row.update({"状态": "已完成", "导入文件已生成": "是", "产品版图已更新": "是", "覆盖质量自检": "通过"})
            self.write_csv_rows(status_path, [status_row])
            with mock.patch("test_design.pipeline.generation_session_is_current", return_value=True), mock.patch(
                "test_design.pipeline.validate_batch_artifacts", return_value=None
            ):
                status = TOOLS.derive_pipeline_status(run_dir)
            self.assertEqual("DELIVERY_REQUIRED", status["state"])

    def test_lifecycle_requires_evidence_for_each_editable_item(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "lifecycle")
            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                plan = next(csv.DictReader(stream))
            plan.update(
                {
                    "功能点": "编辑-危险操作",
                    "测试设计方向": "编辑危险操作状态并验证保存回显和实际生效",
                    "操作类别": "编辑",
                    "验证要求": "回显,持久化,实际生效",
                    "数据策略": "本次创建测试数据",
                    "是否涉及CRUD闭环": "是",
                    "应生成用例数": "5",
                    "计划用例ID": ",".join(f"TC-EDIT-{index:03d}" for index in range(1, 6)),
                }
            )
            self.write_csv_rows(plan_path, [plan])
            discovery_path = run_dir / "page-discovery.csv"
            with discovery_path.open("r", encoding="utf-8-sig", newline="") as stream:
                discovery = next(csv.DictReader(stream))
            evidence = run_dir / "artifacts/screenshots/edit-success.txt"
            evidence.write_text("AI_TEST_EDIT_001 修改后回显并实际生效", encoding="utf-8")
            discovery.update(
                {
                    "测试数据来源": "AI_TEST_EDIT_001",
                    "预期/观察行为": "编辑成功并持久化回显",
                    "结果分支/后续状态": "依赖功能按编辑后值实际生效",
                    "证据路径": "artifacts/screenshots/edit-success.txt",
                }
            )
            self.write_csv_rows(discovery_path, [discovery])
            lifecycle_path = run_dir / "test-data-lifecycle.csv"
            lifecycle = {
                "批次ID": "BATCH-001",
                "最小标题路径": "模块>页面",
                "关联页面/入口": "错误页面",
                "修改项/元素": "危险操作按钮",
                "测试数据ID/名称": "AI_TEST_EDIT_001",
                "数据类型": "编辑测试数据",
                "创建入口": "风险页面",
                "创建结果": "创建成功",
                "查看结果": "详情回显成功",
                "编辑前值": "关闭",
                "编辑后值": "开启",
                "编辑结果": "保存成功",
                "保存后回显": "重新进入页面显示开启",
                "实际生效结果": "关联功能已按开启状态执行",
                "配置生效验证点": "重新进入页面检查回显并调用关联功能验证生效",
                "清理状态": "待清理",
            }
            self.write_csv_rows(lifecycle_path, [lifecycle])
            with self.assertRaisesRegex(ValueError, "record every mutating item separately"):
                TOOLS.validate_batch_artifacts(run_dir, "plan")
            lifecycle["关联页面/入口"] = "风险页面"
            self.write_csv_rows(lifecycle_path, [lifecycle])
            TOOLS.validate_batch_artifacts(run_dir, "plan")

    def test_cases_gate_rejects_duplicate_ids_across_shards(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "duplicate-cases")
            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                plan = next(csv.DictReader(stream))
            ids = [f"TC-RISK-{index:03d}" for index in range(1, 11)]
            plan.update({"应生成用例数": "10", "计划用例ID": ",".join(ids), "实际用例ID": ",".join(ids)})
            self.write_csv_rows(plan_path, [plan])
            TOOLS.record_no_model_uncertainty(run_dir)
            TOOLS.prepare_function_case_generation(run_dir)
            data_dir = run_dir / "artifacts" / "data"
            session = json.loads((data_dir / "generation-session.json").read_text(encoding="utf-8"))
            (data_dir / "function_cases_part_001.json").write_text(
                json.dumps([self.function_case(case_id) for case_id in ids], ensure_ascii=False), encoding="utf-8"
            )
            (data_dir / "function_cases_part_002.json").write_text(
                json.dumps([self.function_case(ids[0])], ensure_ascii=False), encoding="utf-8"
            )
            (data_dir / "function_cases_manifest.json").write_text(
                json.dumps(
                    {
                        "part_size": 10,
                        "total_cases": 11,
                        "parts": ["function_cases_part_001.json", "function_cases_part_002.json"],
                        "generation_session_id": session["generation_session_id"],
                        "source_fingerprint": session["source_fingerprint"],
                    },
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            self.write_valid_sheet_files(data_dir)
            with self.assertRaisesRegex(ValueError, "unique across all manifest shards"):
                TOOLS.validate_batch_artifacts(run_dir, "cases")

    def test_cases_gate_rejects_empty_sheet_json_before_delivery(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "empty-sheet")
            ids = ["TC-RISK-001", "TC-RISK-002", "TC-RISK-003"]
            plan_path = run_dir / "element-case-plan.csv"
            with plan_path.open("r", encoding="utf-8-sig", newline="") as stream:
                plan = next(csv.DictReader(stream))
            plan["实际用例ID"] = ",".join(ids)
            self.write_csv_rows(plan_path, [plan])
            status_path = run_dir / "batch-status.csv"
            with status_path.open("r", encoding="utf-8-sig", newline="") as stream:
                status = next(csv.DictReader(stream))
            status["功能用例数"] = "3"
            self.write_csv_rows(status_path, [status])
            TOOLS.record_no_model_uncertainty(run_dir)
            TOOLS.prepare_function_case_generation(run_dir)
            data_dir = run_dir / "artifacts" / "data"
            session = json.loads((data_dir / "generation-session.json").read_text(encoding="utf-8"))
            (data_dir / "function_cases_part_001.json").write_text(
                json.dumps([self.function_case(case_id) for case_id in ids], ensure_ascii=False), encoding="utf-8"
            )
            manifest_payload = {
                "part_size": 10,
                "total_cases": 3,
                "parts": ["function_cases_part_001.json"],
                "generation_session_id": session["generation_session_id"],
                "source_fingerprint": session["source_fingerprint"],
            }
            manifest_path = data_dir / "function_cases_manifest.json"
            manifest_path.write_text(
                json.dumps(["function_cases_part_001.json"], ensure_ascii=False), encoding="utf-8"
            )
            with self.assertRaisesRegex(ValueError, "must be a JSON object"):
                TOOLS.validate_batch_artifacts(run_dir, "cases")
            manifest_path.write_text(json.dumps(manifest_payload, ensure_ascii=False), encoding="utf-8")
            self.write_valid_sheet_files(data_dir)
            (data_dir / "overview.json").write_text('[{"错误字段":"正式数据"}]', encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "exact target Sheet headers"):
                TOOLS.validate_batch_artifacts(run_dir, "cases")
            empty_overview = {header: "" for header in SHEET_DATA_HEADERS["overview.json"]}
            (data_dir / "overview.json").write_text(json.dumps([empty_overview], ensure_ascii=False), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "at least one non-empty"):
                TOOLS.validate_batch_artifacts(run_dir, "cases")
            self.write_valid_sheet_files(data_dir)
            (data_dir / "risks.json").write_text("[]", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "non-empty row list"):
                TOOLS.validate_batch_artifacts(run_dir, "cases")

            risk_path = run_dir / "risk-confirmation.csv"
            with risk_path.open("r", encoding="utf-8-sig", newline="") as stream:
                reader = csv.DictReader(stream)
                risk_headers = reader.fieldnames or []
            risk = {header: "" for header in risk_headers}
            risk.update(
                {
                    "批次ID": "BATCH-001",
                    "风险ID": "RISK-001",
                    "模型不理解内容/待确认问题": "确认后是否触发异步审批",
                    "已完成深探依据": "已完成确认、取消和关闭路径，页面未展示审批规则",
                    "用户确认结论": "确认后触发异步审批",
                    "处置策略": "按异步审批设计用例",
                    "是否阻塞用例设计": "否",
                    "关联页面/入口": "风险页面",
                    "关联元素名称/文案": "危险操作按钮",
                    "确认状态": "待确认",
                    "关联用例ID": "TC-RISK-001",
                }
            )
            with risk_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=risk_headers)
                writer.writeheader()
                writer.writerow(risk)
            with self.assertRaisesRegex(ValueError, "risk-confirmation.csv is not ready"):
                TOOLS.validate_batch_artifacts(run_dir, "risk")

    def test_delivery_rollback_restores_existing_and_removes_new_files(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            root = Path(value)
            existing = root / "existing.txt"
            created = root / "created.txt"
            existing.write_text("before", encoding="utf-8")

            with self.assertRaisesRegex(RuntimeError, "probe failure"):
                with TOOLS.rollback_files_on_error([existing, created]):
                    existing.write_text("after", encoding="utf-8")
                    created.write_text("new", encoding="utf-8")
                    raise RuntimeError("probe failure")

            self.assertEqual("before", existing.read_text(encoding="utf-8"))
            self.assertFalse(created.exists())

    def test_delivery_lock_rejects_concurrent_writer(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            lock_path = Path(value) / "delivery.lock"
            with TOOLS.exclusive_process_lock(lock_path):
                with self.assertRaisesRegex(RuntimeError, "Another delivery process"):
                    with TOOLS.exclusive_process_lock(lock_path):
                        self.fail("Concurrent delivery should never acquire the same lock")
            with TOOLS.exclusive_process_lock(lock_path):
                self.assertTrue(lock_path.exists())

    def test_complete_delivery_rolls_back_when_prevalidation_fails(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            formal = project_root / "working" / "formal.xlsx"
            import_template = project_root / "working" / "import-template.xlsx"
            formal.parent.mkdir(parents=True)
            shutil.copy2(REPO_ROOT / "docs" / "test-design" / "codebuddy-test-design-template.xlsx", formal)
            shutil.copy2(REPO_ROOT / "docs" / "test-design" / "测试用例模板.xlsx", import_template)
            before = formal.read_bytes()

            with mock.patch.object(TOOLS, "run_python_script", side_effect=RuntimeError("prevalidation failed")):
                with self.assertRaisesRegex(RuntimeError, "prevalidation failed"):
                    TOOLS.complete_deliverables(
                        project_root,
                        formal,
                        import_template,
                        "模块>页面",
                    )

            self.assertEqual(before, formal.read_bytes())
            self.assertFalse((project_root / "docs" / "test-assets" / "imports" / "模块_页面_导入用例.xlsx").exists())
            self.assertFalse((project_root / "docs" / "test-assets" / "modules" / "模块_页面_测试设计.xlsx").exists())
            lock_path = project_root / ".test-design-locks" / "delivery.lock"
            with TOOLS.exclusive_process_lock(lock_path):
                self.assertTrue(lock_path.exists())

    def test_complete_delivery_rejects_module_path_outside_run_leaf(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            self.create_project_root(project_root)
            run_dir = self.make_valid_plan_run(project_root, "delivery-scope")
            working = project_root / "working"
            working.mkdir()
            formal = working / "formal.xlsx"
            import_template = working / "import-template.xlsx"
            shutil.copy2(REPO_ROOT / "docs/test-design/codebuddy-test-design-template.xlsx", formal)
            shutil.copy2(REPO_ROOT / "docs/test-design/测试用例模板.xlsx", import_template)
            with self.assertRaisesRegex(ValueError, "do not deliver one leaf batch under another module path"):
                TOOLS.complete_deliverables(
                    project_root,
                    formal,
                    import_template,
                    "另一个模块>页面",
                    batch_status=run_dir / "batch-status.csv",
                    batch_id="BATCH-001",
                )
            self.assertFalse((project_root / "docs/test-design/deliverables/另一个模块_页面_测试设计.xlsx").exists())

    def test_complete_delivery_rolls_back_product_catalog_after_post_sync_failure(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            project_root = Path(value)
            working = project_root / "working"
            working.mkdir(parents=True)
            formal = working / "formal.xlsx"
            import_template = working / "import-template.xlsx"
            product_map = project_root / "docs/test-assets/product-map.xlsx"
            product_map.parent.mkdir(parents=True)
            discovery = working / "page-discovery.csv"
            shutil.copy2(REPO_ROOT / "docs/test-design/codebuddy-test-design-template.xlsx", formal)
            shutil.copy2(REPO_ROOT / "docs/test-design/测试用例模板.xlsx", import_template)
            shutil.copy2(REPO_ROOT / "docs/test-assets/product-map.xlsx", product_map)
            before_product_map = product_map.read_bytes()
            template = REPO_ROOT / "docs/test-assets/batch-runs/templates/page-discovery-template.csv"
            shutil.copy2(template, discovery)
            self.write_csv_rows(
                discovery,
                [
                    {
                        "批次ID": "BATCH-001",
                        "最小标题路径": "模块>页面",
                        "页面/入口": "页面",
                        "元素名称/文案": "查询按钮",
                        "元素类型": "按钮",
                        "交互方式": "点击",
                        "覆盖状态": "已覆盖",
                    }
                ],
            )
            catalog_paths = TOOLS.product_map_mutable_paths(product_map, "产品>模块>页面")
            with mock.patch.object(TOOLS, "run_python_script", side_effect=RuntimeError("post-sync validation failed")):
                with self.assertRaisesRegex(RuntimeError, "post-sync validation failed"):
                    TOOLS.complete_deliverables(
                        project_root,
                        formal,
                        import_template,
                        "产品>模块>页面",
                        product_map=product_map,
                        page_discovery=discovery,
                    )
            self.assertEqual(before_product_map, product_map.read_bytes())
            self.assertTrue(all(not path.exists() for path in catalog_paths if path != product_map))
            self.assertFalse((project_root / "docs/test-assets/modules/产品_模块_页面_测试设计.xlsx").exists())

    def test_product_map_sync_uses_current_schema_and_real_dependencies(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            root = Path(value)
            formal = root / "formal.xlsx"
            product_map = root / "product-map.xlsx"
            discovery = root / "page-discovery.csv"
            shutil.copy2(REPO_ROOT / "docs" / "test-design" / "codebuddy-test-design-template.xlsx", formal)
            shutil.copy2(REPO_ROOT / "docs" / "test-assets" / "product-map.xlsx", product_map)

            workbook = load_workbook(formal)
            function_sheet = workbook["功能测试用例"]
            function_headers = TOOLS.header_map(function_sheet)
            TOOLS.write_mapped_row(
                function_sheet,
                function_headers,
                2,
                {
                    "用例 ID": "TC-SYNC-001",
                    "模块": "模块>页面",
                    "功能点": "查询",
                    "用例标题": "查询-正常查询",
                    "测试类型": "功能测试",
                    "DFX维度": "DFX功能",
                    "DFX场景": "正常处理",
                    "测试数据": "CODEX_TEST_001",
                },
            )
            requirement_sheet = workbook["需求用户故事拆解"]
            TOOLS.write_mapped_row(
                requirement_sheet,
                TOOLS.header_map(requirement_sheet),
                2,
                {"Story ID/需求 ID": "REQ-001", "依赖系统": "依赖系统A"},
            )
            performance_sheet = workbook["性能测试设计"]
            TOOLS.write_mapped_row(
                performance_sheet,
                TOOLS.header_map(performance_sheet),
                2,
                {"性能场景 ID": "PERF-001", "业务链路": "查询链路"},
            )
            workbook.save(formal)

            template = REPO_ROOT / "docs" / "test-assets" / "batch-runs" / "templates" / "page-discovery-template.csv"
            with template.open("r", encoding="utf-8-sig", newline="") as fp:
                headers = next(csv.reader(fp))
            row = {header: "" for header in headers}
            row.update(
                {
                    "批次ID": "BATCH-001",
                    "最小标题路径": "模块>页面",
                    "页面/入口": "查询页面",
                    "元素名称/文案": "查询",
                    "元素类型": "按钮",
                    "覆盖状态": "已覆盖",
                    "关联用例ID": "TC-SYNC-001",
                }
            )
            with discovery.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=headers)
                writer.writeheader()
                writer.writerow(row)

            TOOLS.sync_product_map(
                product_map,
                formal,
                discovery,
                "产品>模块>页面",
                "docs/test-assets/modules/模块_页面_测试设计.xlsx",
            )
            synced = load_workbook(product_map, data_only=True)
            dependencies = TOOLS.non_empty_rows(
                synced["跨模块依赖关系"], TOOLS.header_map(synced["跨模块依赖关系"])
            )
            reusable_data = TOOLS.non_empty_rows(
                synced["可复用测试数据"], TOOLS.header_map(synced["可复用测试数据"])
            )
            impacts = TOOLS.non_empty_rows(
                synced["变更影响分析"], TOOLS.header_map(synced["变更影响分析"])
            )
            changes = TOOLS.non_empty_rows(synced["变更记录"], TOOLS.header_map(synced["变更记录"]))
            self.assertEqual("依赖系统A", dependencies[-1]["依赖模块"])
            self.assertEqual("CODEX_TEST_001", reusable_data[-1]["数据用途"])
            self.assertTrue(impacts[-1]["变更ID"].startswith("CHANGE-"))
            self.assertTrue(changes[-1]["日期"])
            catalog = root / "catalog"
            self.assertTrue((catalog / "index.json").exists())
            module_documents = list((catalog / "modules").glob("*.json"))
            documents = [json.loads(path.read_text(encoding="utf-8")) for path in module_documents]
            matching_documents = [
                item for item in documents if item["module_key"] == "产品>模块>页面"
            ]
            self.assertEqual(1, len(matching_documents))
            document = matching_documents[0]
            self.assertEqual("2.0.0", document["schema_version"])
            self.assertEqual("产品>模块>页面", document["module_key"])
            self.assertEqual("依赖系统A", document["facts"]["跨模块依赖关系"][0]["data"]["依赖模块"])
            for sheet in synced.worksheets:
                for row_index in range(2, sheet.max_row + 1):
                    combined = "".join("" if cell.value is None else str(cell.value) for cell in sheet[row_index])
                    self.assertNotIn("示例", combined, f"{sheet.title} still contains a sample row")
                    self.assertNotIn("FLOW-DEMO-001", combined, f"{sheet.title} still contains a demo row")

            TOOLS.sync_product_map(
                product_map,
                formal,
                discovery,
                "产品>模块>页面",
                "docs/test-assets/modules/模块_页面_测试设计.xlsx",
            )
            repeated_documents = [
                json.loads(path.read_text(encoding="utf-8"))
                for path in (catalog / "modules").glob("*.json")
            ]
            self.assertEqual(
                1,
                len([item for item in repeated_documents if item["module_key"] == "产品>模块>页面"]),
            )
            rebuilt = load_workbook(product_map, data_only=True)
            rebuilt_dependencies = TOOLS.non_empty_rows(
                rebuilt["跨模块依赖关系"], TOOLS.header_map(rebuilt["跨模块依赖关系"])
            )
            self.assertEqual(
                1,
                len([item for item in rebuilt_dependencies if item["当前模块"] == "产品>模块>页面"]),
            )

    def test_product_fact_migration_preserves_existing_real_excel_rows(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            root = Path(value)
            product_map = root / "product-map.xlsx"
            shutil.copy2(REPO_ROOT / "docs" / "test-assets" / "product-map.xlsx", product_map)
            workbook = load_workbook(product_map)
            sheet = workbook["产品模块地图"]
            headers = TOOLS.header_map(sheet)
            for cell in sheet[2]:
                cell.value = None
            TOOLS.write_mapped_row(
                sheet,
                headers,
                2,
                {
                    "产品/系统": "真实产品",
                    "一级模块": "资产模块",
                    "页面/入口": "资产页面",
                    "菜单路径/URL": "资产模块>资产页面",
                    "模块功能摘要": "迁移前已存在的真实资产",
                    "归档测试设计路径": "docs/test-assets/modules/资产模块_测试设计.xlsx",
                    "覆盖状态": "已覆盖",
                    "最后更新时间": "2026-07-10",
                },
            )
            workbook.save(product_map)

            TOOLS.ensure_catalog(product_map)
            TOOLS.rebuild_index(product_map)
            TOOLS.project_catalog_to_workbook(product_map)

            legacy_path = root / "catalog" / "modules" / "_legacy.json"
            self.assertTrue(legacy_path.exists())
            legacy = json.loads(legacy_path.read_text(encoding="utf-8"))
            migrated_rows = legacy["facts"]["产品模块地图"]
            self.assertEqual("真实产品", migrated_rows[0]["data"]["产品/系统"])
            projected = load_workbook(product_map, data_only=True)
            projected_rows = TOOLS.non_empty_rows(
                projected["产品模块地图"], TOOLS.header_map(projected["产品模块地图"])
            )
            self.assertEqual("真实产品", projected_rows[0]["产品/系统"])

    @unittest.skipIf(os.name != "nt", "PowerShell upgrade rollback integration runs on Windows")
    @unittest.skipIf(os.environ.get("TEST_DESIGN_SKIP_UPGRADE_INTEGRATION") == "1", "Avoid recursive upgrade test")
    def test_upgrade_failure_restores_framework_and_protected_assets(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            root = Path(value)
            sandbox = root / "repo"
            ignore = shutil.ignore_patterns(".git", "__pycache__", "*.pyc", "dist", "dist-test", ".upgrade-backups", ".test-design-locks")
            shutil.copytree(REPO_ROOT, sandbox, ignore=ignore)
            environment = os.environ.copy()
            environment["TEST_DESIGN_SKIP_UPGRADE_INTEGRATION"] = "1"

            package_result = subprocess.run(
                [
                    "powershell",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(sandbox / "scripts" / "new-framework-upgrade-package.ps1"),
                    "-OutputDir",
                    "dist-test",
                ],
                cwd=sandbox,
                env=environment,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, package_result.returncode, package_result.stderr)
            valid_package = sandbox / "dist-test" / f"framework-upgrade-{FRAMEWORK_VERSION}.zip"
            extracted = root / "broken-package"
            with zipfile.ZipFile(valid_package) as archive:
                archive.extractall(extracted)
            canonical_rule = Path(".codebuddy") / "rules" / "test-design-rule.md"
            (extracted / canonical_rule).write_text("BROKEN UPGRADE PACKAGE\n", encoding="utf-8")
            probe_file = extracted / "tests" / "new-file-probe.txt"
            probe_file.parent.mkdir(parents=True, exist_ok=True)
            probe_file.write_text("must be removed by rollback", encoding="utf-8")
            broken_package = root / "broken-upgrade.zip"
            with zipfile.ZipFile(broken_package, "w", zipfile.ZIP_DEFLATED) as archive:
                for path in extracted.rglob("*"):
                    if path.is_file():
                        archive.write(path, path.relative_to(extracted))

            rule_before = (sandbox / canonical_rule).read_bytes()
            product_map_before = (sandbox / "docs" / "test-assets" / "product-map.xlsx").read_bytes()
            upgrade_result = subprocess.run(
                [
                    "powershell",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(sandbox / "scripts" / "upgrade-framework.ps1"),
                    "-PackagePath",
                    str(broken_package),
                ],
                cwd=sandbox,
                env=environment,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(0, upgrade_result.returncode)
            self.assertIn("restoring framework and protected assets", upgrade_result.stderr + upgrade_result.stdout)
            self.assertEqual(rule_before, (sandbox / canonical_rule).read_bytes())
            self.assertEqual(product_map_before, (sandbox / "docs" / "test-assets" / "product-map.xlsx").read_bytes())
            self.assertFalse((sandbox / "tests" / "new-file-probe.txt").exists())

    @unittest.skipIf(os.name != "nt", "PowerShell asset migration integration runs on Windows")
    @unittest.skipIf(os.environ.get("TEST_DESIGN_SKIP_UPGRADE_INTEGRATION") == "1", "Avoid recursive upgrade test")
    def test_upgrade_migrates_asset_schema_1_to_2_without_losing_excel_facts(self) -> None:
        with tempfile.TemporaryDirectory() as value:
            root = Path(value)
            sandbox = root / "repo"
            ignore = shutil.ignore_patterns(".git", "__pycache__", "*.pyc", "dist", "dist-test", ".upgrade-backups", ".test-design-locks")
            shutil.copytree(REPO_ROOT, sandbox, ignore=ignore)
            environment = os.environ.copy()
            environment["TEST_DESIGN_SKIP_UPGRADE_INTEGRATION"] = "1"

            package_result = subprocess.run(
                [
                    "powershell",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(sandbox / "scripts" / "new-framework-upgrade-package.ps1"),
                    "-OutputDir",
                    "dist-test",
                ],
                cwd=sandbox,
                env=environment,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, package_result.returncode, package_result.stderr)
            package = sandbox / "dist-test" / f"framework-upgrade-{FRAMEWORK_VERSION}.zip"

            agents_path = sandbox / "AGENTS.md"
            agents_text = agents_path.read_text(encoding="utf-8")
            legacy_entry = "# Legacy local instructions\n\n- MUST_NOT_BE_OVERWRITTEN\n"
            agents_path.write_text(legacy_entry, encoding="utf-8")
            marker_result = subprocess.run(
                [
                    "powershell",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(sandbox / "scripts" / "upgrade-framework.ps1"),
                    "-PackagePath",
                    str(package),
                ],
                cwd=sandbox,
                env=environment,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertNotEqual(0, marker_result.returncode)
            self.assertIn("no LOCAL-OVERRIDES block", marker_result.stderr + marker_result.stdout)
            self.assertEqual(legacy_entry, agents_path.read_text(encoding="utf-8"))
            agents_path.write_text(agents_text, encoding="utf-8")
            local_override = "\n- LOCAL_OVERRIDE_MUST_SURVIVE\n"
            agents_text = re.sub(
                r"(?s)(?<=<!-- LOCAL-OVERRIDES:BEGIN -->).*?(?=<!-- LOCAL-OVERRIDES:END -->)",
                local_override,
                agents_text,
            )
            agents_path.write_text(agents_text, encoding="utf-8")

            (sandbox / "VERSION").write_text(
                "framework_version=1.2.0\nasset_schema_version=1.0.0\n",
                encoding="utf-8",
            )
            product_map = sandbox / "docs" / "test-assets" / "product-map.xlsx"
            workbook = load_workbook(product_map)
            sheet = workbook["产品模块地图"]
            for cell in sheet[2]:
                cell.value = None
            TOOLS.write_mapped_row(
                sheet,
                TOOLS.header_map(sheet),
                2,
                {
                    "产品/系统": "迁移产品",
                    "一级模块": "迁移模块",
                    "页面/入口": "迁移页面",
                    "菜单路径/URL": "迁移模块>迁移页面",
                    "模块功能摘要": "升级前真实事实",
                    "归档测试设计路径": "docs/test-assets/modules/迁移模块_测试设计.xlsx",
                    "覆盖状态": "已覆盖",
                    "最后更新时间": "2026-07-10",
                },
            )
            workbook.save(product_map)

            upgrade_result = subprocess.run(
                [
                    "powershell",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(sandbox / "scripts" / "upgrade-framework.ps1"),
                    "-PackagePath",
                    str(package),
                    "-RunMigrations",
                ],
                cwd=sandbox,
                env=environment,
                check=False,
                capture_output=True,
                text=True,
            )
            self.assertEqual(0, upgrade_result.returncode, upgrade_result.stderr + upgrade_result.stdout)
            self.assertIn("asset_schema_version=2.0.0", (sandbox / "VERSION").read_text(encoding="utf-8-sig"))
            legacy_path = sandbox / "docs" / "test-assets" / "catalog" / "modules" / "_legacy.json"
            self.assertTrue(legacy_path.exists())
            legacy = json.loads(legacy_path.read_text(encoding="utf-8"))
            migrated = legacy["facts"]["产品模块地图"]
            self.assertEqual("迁移产品", migrated[0]["data"]["产品/系统"])
            self.assertIn("LOCAL_OVERRIDE_MUST_SURVIVE", agents_path.read_text(encoding="utf-8"))


if __name__ == "__main__":
    unittest.main()
