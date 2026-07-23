# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import uuid
from collections import OrderedDict
from copy import copy, deepcopy
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter, range_boundaries
except ImportError as exc:  # pragma: no cover - depends on local runtime packaging.
    raise SystemExit(
        "ERROR: openpyxl is required. Run this script in the CodeBuddy/Codex spreadsheet runtime "
        "or install openpyxl in the active Python environment."
    ) from exc


FORMAL_FUNCTION_SHEET = "功能测试用例"
FORMAL_SHEETS = [
    "测试设计总览",
    "需求用户故事拆解",
    "测试场景矩阵",
    "功能测试用例",
    "性能测试设计",
    "风险与待确认问题",
    "自动化建议",
    "页面元素覆盖清单",
]
SHEET_ROW_KEYS = {
    "测试设计总览": None,
    "需求用户故事拆解": "Story ID/需求 ID",
    "测试场景矩阵": "场景 ID",
    "功能测试用例": "用例 ID",
    "性能测试设计": "性能场景 ID",
    "风险与待确认问题": "编号",
    "自动化建议": "用例 ID/场景 ID",
    "页面元素覆盖清单": "元素 ID",
}
SHEET_REQUIRED_FIELDS = {
    "测试设计总览": [
        "项目/模块",
        "需求名称",
        "版本/迭代",
        "测试负责人",
        "需求来源",
        "测试范围",
        "不测范围",
        "测试类型",
        "测试环境",
        "主要风险",
        "准入条件",
        "准出条件",
        "待确认问题",
    ],
    "需求用户故事拆解": ["Story ID/需求 ID", "用户故事/需求描述", "验收标准"],
    "测试场景矩阵": [
        "场景 ID",
        "功能点",
        "测试维度",
        "DFX维度",
        "DFX场景",
        "测试对象/页面元素",
        "输入数据/状态条件",
        "观察点",
        "优先级",
        "是否生成用例",
    ],
    "功能测试用例": [
        "用例 ID",
        "Story ID/需求 ID",
        "模块",
        "功能点",
        "用例标题",
        "优先级",
        "测试类型",
        "DFX维度",
        "DFX场景",
        "前置条件",
        "操作步骤",
        "预期结果",
        "是否适合自动化",
    ],
    "性能测试设计": [
        "性能场景 ID",
        "业务链路",
        "性能测试类型",
        "DFX维度",
        "DFX场景",
        "响应时间目标",
        "监控指标",
        "通过标准",
        "是否纳入本轮测试",
    ],
    "风险与待确认问题": ["编号", "类型", "描述", "影响范围", "风险等级", "建议处理方式", "状态"],
    "自动化建议": [
        "用例 ID/场景 ID",
        "自动化层级",
        "自动化价值",
        "自动化优先级",
        "依赖数据",
        "稳定性风险",
        "建议框架/工具",
    ],
    "页面元素覆盖清单": [
        "元素 ID",
        "页面/入口",
        "页面 URL/菜单路径",
        "元素名称/文案",
        "元素类型",
        "交互方式",
        "适用DFX维度",
        "适用DFX场景",
        "预期行为",
        "覆盖状态",
        "发现方式",
    ],
}
FACT_STATUSES = {"已实测", "页面观察", "DFX设计", "待确认"}
AMBIGUOUS_EXPECTED_PATTERNS = (
    r"结果或错误",
    r"成功或失败",
    r"接受.{0,30}或.{0,30}截断",
    r"提示.{0,30}或.{0,30}报错",
    r"页面正常响应",
    r"系统处理正确",
    r"结果符合预期",
)
DESTRUCTIVE_PAYLOAD_PATTERNS = (
    r"(?i)\brm\s+-rf\b",
    r"(?i)\bformat\s+[a-z]:",
    r"(?i)\bdel\s+/[fsq]",
    r"(?i)\bdrop\s+(database|schema)\b",
    r"(?i)\btruncate\s+table\b",
)
PERFORMANCE_SOURCE_MARKERS = ("需求阈值", "实测基线", "建议目标", "待确认", "不适用")
IMPORT_MULTILINE_FIELDS = ["测试步骤描述", "测试步骤预期结果", "前置条件", "测试用例说明", "备注"]
FORMAL_MULTILINE_FIELDS = {
    "测试设计总览": ["测试范围", "不测范围", "主要风险", "准入条件", "准出条件", "待确认问题"],
    "需求用户故事拆解": ["用户故事/需求描述", "业务价值", "验收标准", "业务规则", "前置条件", "后置影响", "待确认问题"],
    "测试场景矩阵": ["测试对象/页面元素", "输入数据/状态条件", "观察点", "备注"],
    "功能测试用例": ["前置条件", "测试数据", "操作步骤", "预期结果", "备注"],
    "性能测试设计": ["业务链路", "监控指标", "通过标准", "造数策略", "风险说明"],
    "风险与待确认问题": ["描述", "影响范围", "建议处理方式"],
    "自动化建议": ["依赖数据", "Mock 需求", "稳定性风险", "建议框架/工具", "备注"],
    "页面元素覆盖清单": ["预期行为", "业务依据/规则来源", "待确认问题/备注"],
}

IMPORT_AUTO_FIELDS = {"测试用例系统编号", "作者"}
IMPORT_ALLOWED_VALUES = {
    "测试类型": {"功能测试", "性能规格测试", "可靠性测试", "兼容性测试", "可维护性测试", "安全性测试", "易用性测试"},
    "测试用例级别": {"L1", "L2", "L3", "L4"},
    "执行方式": {"自动化", "手动"},
}


def header_map(ws, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def row_dict(ws, headers: dict[str, int], row_index: int) -> dict[str, str]:
    result: dict[str, str] = {}
    for name, column in headers.items():
        value = ws.cell(row=row_index, column=column).value
        result[name] = "" if value is None else str(value).strip()
    return result


def non_empty_rows(ws, headers: dict[str, int]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row_index in range(2, ws.max_row + 1):
        row = row_dict(ws, headers, row_index)
        if any(value for value in row.values()):
            rows.append(row)
    return rows


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for source_cell in ws[source_row]:
        target_cell = ws.cell(row=target_row, column=source_cell.column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def copy_cell_format(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)
    if source_cell.protection:
        target_cell.protection = copy(source_cell.protection)


def copy_template_row_format(template_ws, target_ws, template_row: int, target_row: int) -> None:
    max_col = min(target_ws.max_column, template_ws.max_column)
    for column in range(1, max_col + 1):
        copy_cell_format(template_ws.cell(row=template_row, column=column), target_ws.cell(row=target_row, column=column))
    target_ws.row_dimensions[target_row].height = template_ws.row_dimensions[template_row].height


def copy_column_dimensions(template_ws, target_ws) -> None:
    for key, dimension in template_ws.column_dimensions.items():
        target_dimension = target_ws.column_dimensions[key]
        target_dimension.width = dimension.width
        target_dimension.hidden = dimension.hidden
        target_dimension.bestFit = dimension.bestFit


def extend_validation_ranges(ws, max_row: int) -> None:
    if max_row < 2:
        return
    for validation in ws.data_validations.dataValidation:
        ranges: list[str] = []
        for cell_range in validation.sqref.ranges:
            min_col, min_row, max_col, old_max_row = range_boundaries(str(cell_range))
            if min_row <= 2 <= old_max_row:
                old_max_row = max(old_max_row, max_row)
            ranges.append(
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{old_max_row}"
            )
        validation.sqref = " ".join(ranges)


def apply_template_sheet_format(template_ws, target_ws) -> None:
    copy_column_dimensions(template_ws, target_ws)
    if target_ws.max_row >= 2 and template_ws.max_row >= 2:
        for row_index in range(2, target_ws.max_row + 1):
            copy_template_row_format(template_ws, target_ws, 2, row_index)
    target_ws.data_validations = deepcopy(template_ws.data_validations)
    extend_validation_ranges(target_ws, max(target_ws.max_row, 200))


def apply_template_workbook_format(target_wb, template_wb) -> None:
    for sheet_name in target_wb.sheetnames:
        if sheet_name not in template_wb.sheetnames:
            continue
        apply_template_sheet_format(template_wb[sheet_name], target_wb[sheet_name])


def set_wrap(ws, headers: dict[str, int], row_index: int, field_names: list[str]) -> None:
    for field in field_names:
        column = headers.get(field)
        if not column:
            continue
        cell = ws.cell(row=row_index, column=column)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal,
            vertical="top",
            text_rotation=cell.alignment.text_rotation,
            wrap_text=True,
            shrink_to_fit=cell.alignment.shrink_to_fit,
            indent=cell.alignment.indent,
        )


def normalize_case_level(priority: str) -> str:
    value = (priority or "").upper()
    if value in {"L1", "L2", "L3", "L4"}:
        return value
    if value in {"P0", "P1", "高", "高优先级"}:
        return "L1"
    if value in {"P2", "中", "中优先级"}:
        return "L2"
    if value in {"P3", "低", "低优先级"}:
        return "L3"
    return "L2"


def normalize_test_type(value: str) -> str:
    if value in IMPORT_ALLOWED_VALUES["测试类型"]:
        return value
    if "性能" in value:
        return "性能规格测试"
    if "兼容" in value:
        return "兼容性测试"
    if "安全" in value or "权限" in value:
        return "安全性测试"
    if "可靠" in value or "稳定" in value:
        return "可靠性测试"
    if "易用" in value:
        return "易用性测试"
    if "维护" in value:
        return "可维护性测试"
    return "功能测试"


def execution_mode(row: dict[str, str]) -> str:
    note = "".join([row.get("备注", ""), row.get("是否适合自动化", "")])
    automation_markers = ["自动化资产", "脚本", "流水线", "API自动化", "UI自动化", "已实现"]
    if "自动化" in note and any(marker in note for marker in automation_markers):
        return "自动化"
    return "手动"


def module_names(module_path: str) -> list[str]:
    parts = [part.strip() for part in module_path.replace("/", ">").split(">") if part.strip()]
    return (parts + [""] * 5)[:5]


def import_module_names(module_path: str, product_name: str | None = None) -> list[str]:
    """Map a real 1-N level path to import fields without changing file naming.

    The current import template requires the first three module name columns. If
    the real menu path is shallower, only the import fields reuse the deepest
    known name; the canonical path and deliverable filename remain unchanged.
    """
    parts = canonical_module_parts(module_path, product_name)
    if not parts:
        parts = [module_path.strip() or "测试设计"]
    while len(parts) < 3:
        parts.append(parts[-1])
    return (parts + [""] * 5)[:5]


def canonical_module_parts(module_path: str, product_name: str | None = None) -> list[str]:
    parts = [part.strip() for part in module_path.replace("\\", ">").replace("/", ">").split(">") if part.strip()]
    if product_name and parts and parts[0] == product_name.strip():
        parts = parts[1:]
    return parts


def deliverable_names(module_path: str, product_name: str | None = None) -> tuple[str, str, str]:
    parts = canonical_module_parts(module_path, product_name)
    stem = safe_filename(">".join(parts) if parts else module_path)
    return stem, f"{stem}_测试设计.xlsx", f"{stem}_导入用例.xlsx"


def module_leaf_name(module_path: str) -> str:
    parts = [part.strip() for part in module_path.replace("/", ">").split(">") if part.strip()]
    return parts[-1] if parts else module_path


def clear_data_rows(ws, start_row: int = 2) -> None:
    if ws.max_row > start_row:
        ws.delete_rows(start_row + 1, ws.max_row - start_row)
    if ws.max_row >= start_row:
        for cell in ws[start_row]:
            cell.value = None


def worksheet_used_range(ws) -> str:
    return f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"


def remove_worksheet_tables_and_refresh_filter(ws) -> None:
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]
    sheet_ref = worksheet_used_range(ws)
    if ws.max_row > 1 and ws.max_column > 1:
        ws.auto_filter.ref = sheet_ref


def remove_workbook_tables_and_refresh_filters(wb) -> None:
    for ws in wb.worksheets:
        remove_worksheet_tables_and_refresh_filter(ws)


def relative_project_path(project_root: Path, path: Path) -> str:
    return path.resolve().relative_to(project_root.resolve()).as_posix()


def safe_filename(value: str) -> str:
    cleaned = value.replace("\\", ">").replace("/", ">")
    for char in '<>:"|?*':
        cleaned = cleaned.replace(char, "_")
    cleaned = "_".join(part.strip() for part in cleaned.split("_") if part.strip())
    cleaned = cleaned.replace(" ", "")
    return cleaned or "测试设计"


def copy_workbook(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() == target.resolve():
        return
    shutil.copy2(source, target)


def atomic_publish_copies(copies: list[tuple[Path, Path]]) -> None:
    """Publish a set of files as one rollback-capable operation."""
    token = uuid.uuid4().hex
    staged: list[tuple[Path, Path]] = []
    backups: list[tuple[Path, Path]] = []
    published: list[Path] = []
    try:
        for source, target in copies:
            target.parent.mkdir(parents=True, exist_ok=True)
            stage = target.with_name(f".{target.name}.{token}.tmp")
            shutil.copy2(source, stage)
            staged.append((stage, target))
        for _, target in staged:
            if target.exists():
                backup = target.with_name(f".{target.name}.{token}.bak")
                os.replace(target, backup)
                backups.append((backup, target))
        for stage, target in staged:
            os.replace(stage, target)
            published.append(target)
    except Exception:
        for target in published:
            if target.exists():
                target.unlink()
        for backup, target in backups:
            if backup.exists():
                os.replace(backup, target)
        raise
    finally:
        for stage, _ in staged:
            if stage.exists():
                stage.unlink()
        for backup, _ in backups:
            if backup.exists():
                backup.unlink()


def cleanup_excel_lock_files(directories: set[Path]) -> None:
    for directory in directories:
        if not directory.exists():
            continue
        for lock_file in directory.glob("~$*.xlsx"):
            try:
                lock_file.unlink()
            except OSError:
                # An open Excel workbook owns its lock file; never fail delivery for it.
                pass


def legacy_repeated_leaf_names(module_path: str, product_name: str | None = None) -> tuple[str, str] | None:
    parts = canonical_module_parts(module_path, product_name)
    if not parts or len(parts) >= 3:
        return None
    legacy_parts = parts + [parts[-1]] * (3 - len(parts))
    legacy_stem = safe_filename(">".join(legacy_parts))
    current_stem = safe_filename(">".join(parts))
    if legacy_stem == current_stem:
        return None
    return f"{legacy_stem}_测试设计.xlsx", f"{legacy_stem}_导入用例.xlsx"


def prepare_formal_workbook(template: Path, output: Path) -> None:
    """Copy the formal template and clear example data before model filling."""
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    wb = load_workbook(output)
    missing = [name for name in FORMAL_SHEETS if name not in wb.sheetnames]
    if missing:
        raise ValueError(f"Formal template is missing required sheets: {missing}")
    for sheet_name in FORMAL_SHEETS:
        clear_data_rows(wb[sheet_name])
    remove_workbook_tables_and_refresh_filters(wb)
    wb.save(output)


def _section_rows(data: object, section_name: str) -> list[dict[str, object]]:
    if not isinstance(data, dict):
        raise ValueError("JSON shard root must be an object")
    section = data.get(section_name)
    if section is None:
        return []
    if isinstance(section, dict):
        section = section.get("rows")
    if not isinstance(section, list):
        raise ValueError(f"{section_name} must be a rows array or an object containing a rows array")
    invalid_rows = [index for index, row in enumerate(section, start=1) if not isinstance(row, dict)]
    if invalid_rows:
        raise ValueError(f"{section_name} contains non-object rows at positions: {invalid_rows}")
    return section


def _normalized_row(row: dict[str, object]) -> dict[str, str]:
    return {
        str(key).strip(): "" if value is None else str(value).strip()
        for key, value in row.items()
    }


def _numbered_lines(value: str, label: str) -> None:
    lines = [line.strip() for line in value.splitlines() if line.strip()]
    if not lines:
        raise ValueError(f"{label} must not be empty")
    for expected_number, line in enumerate(lines, start=1):
        match = re.match(r"^(\d+)\.\s*\S+", line)
        if not match or int(match.group(1)) != expected_number:
            raise ValueError(f"{label} must use consecutive numbered lines; got: {line}")


def _validate_case_row(row: dict[str, str], source: Path, index: int) -> None:
    label = f"{source.name} 功能测试用例 row {index}"
    function_point = row["功能点"]
    if not row["用例标题"].startswith(f"{function_point}-"):
        raise ValueError(f"{label} title must start with 功能点-: {row['用例标题']}")
    _numbered_lines(row["操作步骤"], f"{label} 操作步骤")
    _numbered_lines(row["预期结果"], f"{label} 预期结果")
    first_steps = "\n".join(row["操作步骤"].splitlines()[:3])
    entry_markers = ("登录", "打开系统", "访问系统", "进入系统", "打开平台", "访问平台", "进入平台", "URL")
    if not any(marker in first_steps for marker in entry_markers):
        raise ValueError(f"{label} must start from the system/project entry")
    if not any(marker in first_steps for marker in ("菜单", "模块", "导航", "路径", ">", "-", "—", "－", "页面")):
        raise ValueError(f"{label} must include the business navigation path before control operations")
    expected = row["预期结果"]
    for pattern in AMBIGUOUS_EXPECTED_PATTERNS:
        if re.search(pattern, expected):
            raise ValueError(f"{label} contains an ambiguous expected result: {pattern}")
    executable_text = "\n".join([row.get("测试数据", ""), row["操作步骤"]])
    for pattern in DESTRUCTIVE_PAYLOAD_PATTERNS:
        if re.search(pattern, executable_text):
            raise ValueError(
                f"{label} contains a destructive executable payload. "
                "Use a non-destructive marker or read-only payload without masking internal test data."
            )


def _validate_performance_row(row: dict[str, str], source: Path, index: int) -> None:
    target = row.get("响应时间目标", "")
    if not any(marker in target for marker in PERFORMANCE_SOURCE_MARKERS):
        raise ValueError(
            f"{source.name} 性能测试设计 row {index} 响应时间目标 must identify its source as "
            f"one of {PERFORMANCE_SOURCE_MARKERS}"
        )
    if "实测基线" in target:
        evidence = "\n".join(
            [
                row.get("通过标准", ""),
                row.get("造数策略", ""),
                row.get("风险说明", ""),
            ]
        )
        if not any(marker in evidence for marker in ("采样", "测量", "记录", "基线")):
            raise ValueError(
                f"{source.name} 性能测试设计 row {index} claims an 实测基线 without measurement provenance"
            )


def _validate_shard_row(
    section_name: str,
    row: dict[str, str],
    allowed_headers: set[str],
    source: Path,
    index: int,
) -> None:
    unknown = sorted(set(row) - allowed_headers)
    if unknown:
        raise ValueError(f"{source.name} {section_name} row {index} has unknown fields: {unknown}")
    missing = [field for field in SHEET_REQUIRED_FIELDS[section_name] if not row.get(field, "").strip()]
    if missing:
        raise ValueError(f"{source.name} {section_name} row {index} is missing required fields: {missing}")
    if section_name == "功能测试用例":
        _validate_case_row(row, source, index)
    elif section_name == "性能测试设计":
        _validate_performance_row(row, source, index)


def load_design_shards(shards_dir: Path, formal_template: Path) -> dict[str, list[dict[str, str]]]:
    """Load, validate and deterministically merge existing per-function JSON shards."""
    if not shards_dir.exists():
        raise ValueError(f"JSON shard directory not found: {shards_dir}")
    generated_python = sorted(shards_dir.rglob("*.py"))
    if generated_python:
        names = ", ".join(path.name for path in generated_python[:5])
        raise ValueError(
            f"Run directories must not contain task-specific Python producers: {names}. "
            "Use compile-deliverables with JSON shards instead."
        )
    shard_files = sorted(shards_dir.rglob("*.json"))
    if not shard_files:
        raise ValueError(f"No JSON shards found under: {shards_dir}")

    template_wb = load_workbook(formal_template, read_only=True, data_only=False)
    allowed_by_sheet = {
        sheet_name: set(header_map(template_wb[sheet_name]))
        for sheet_name in FORMAL_SHEETS
    }
    template_wb.close()
    merged: dict[str, OrderedDict[str, dict[str, str]]] = {
        sheet_name: OrderedDict() for sheet_name in FORMAL_SHEETS
    }
    singleton_overview: dict[str, str] | None = None

    for shard_path in shard_files:
        try:
            with shard_path.open("r", encoding="utf-8-sig") as fp:
                data = json.load(fp)
        except json.JSONDecodeError as exc:
            raise ValueError(
                f"{shard_path} failed JSON syntax validation: line {exc.lineno}, "
                f"column {exc.colno}: {exc.msg}"
            ) from exc
        if not isinstance(data, dict):
            raise ValueError(f"{shard_path} must contain a JSON object")
        for section_name in FORMAL_SHEETS:
            for index, raw_row in enumerate(_section_rows(data, section_name), start=1):
                row = _normalized_row(raw_row)
                _validate_shard_row(
                    section_name,
                    row,
                    allowed_by_sheet[section_name],
                    shard_path,
                    index,
                )
                if section_name == "测试设计总览":
                    if singleton_overview is None:
                        singleton_overview = row
                    elif singleton_overview != row:
                        raise ValueError(
                            f"Conflicting 测试设计总览 rows across JSON shards: {shard_path}"
                        )
                    continue
                key_field = SHEET_ROW_KEYS[section_name]
                assert key_field
                key = row[key_field]
                existing = merged[section_name].get(key)
                if existing is None:
                    merged[section_name][key] = row
                elif existing != row:
                    raise ValueError(
                        f"Conflicting {section_name} key {key!r} across JSON shards: {shard_path}"
                    )

    if singleton_overview is None:
        raise ValueError("JSON shards must contain exactly one 测试设计总览 row")
    merged_rows: dict[str, list[dict[str, str]]] = {
        section_name: list(rows.values()) for section_name, rows in merged.items()
    }
    merged_rows["测试设计总览"] = [singleton_overview]
    for section_name in FORMAL_SHEETS[:-1]:
        if not merged_rows[section_name]:
            raise ValueError(f"JSON shards do not contain any rows for required section: {section_name}")

    case_rows = merged_rows["功能测试用例"]
    seen_bodies: dict[tuple[str, str], str] = {}
    for row in case_rows:
        body = (row["操作步骤"], row["预期结果"])
        if body in seen_bodies:
            raise ValueError(
                f"Cases {seen_bodies[body]} and {row['用例 ID']} have identical steps and expected results"
            )
        seen_bodies[body] = row["用例 ID"]
    automation_rows = merged_rows["自动化建议"]
    seen_automation_bodies: dict[tuple[tuple[str, str], ...], str] = {}
    for row in automation_rows:
        body = tuple(
            (field, value)
            for field, value in row.items()
            if field != SHEET_ROW_KEYS["自动化建议"]
        )
        if body in seen_automation_bodies:
            raise ValueError(
                f"Automation suggestions {seen_automation_bodies[body]} and "
                f"{row['用例 ID/场景 ID']} are identical; write a targeted suggestion "
                "or merge the referenced IDs into one row"
            )
        seen_automation_bodies[body] = row["用例 ID/场景 ID"]

    function_order: OrderedDict[str, None] = OrderedDict()
    for row in case_rows:
        function_order.setdefault(row["功能点"], None)
    function_rank = {name: index for index, name in enumerate(function_order)}
    original_rank = {row["用例 ID"]: index for index, row in enumerate(case_rows)}
    merged_rows["功能测试用例"] = sorted(
        case_rows,
        key=lambda row: (function_rank[row["功能点"]], original_rank[row["用例 ID"]]),
    )
    return merged_rows


def _parse_case_ids(value: str) -> list[str]:
    return [
        item.strip()
        for item in re.split(r"[,，;；\n]+", value or "")
        if item.strip()
    ]


def page_element_rows(
    page_discovery: Path,
    case_rows: list[dict[str, str]],
) -> list[dict[str, str]]:
    with page_discovery.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        headers = reader.fieldnames or []
        discovery_rows = list(reader)
    if not headers:
        raise ValueError(f"page-discovery.csv has no header row: {page_discovery}")
    case_by_id = {row["用例 ID"]: row for row in case_rows}
    elements: list[dict[str, str]] = []
    for index, source in enumerate(discovery_rows, start=1):
        element_name = (source.get("元素名称/文案") or "").strip()
        if not element_name:
            continue
        status = (source.get("事实状态") or "").strip()
        if status not in FACT_STATUSES:
            raise ValueError(
                f"page-discovery.csv row {index} has invalid 事实状态 {status!r}; expected {sorted(FACT_STATUSES)}"
            )
        linked_ids = _parse_case_ids(source.get("关联用例ID", ""))
        generated = (source.get("是否已生成用例") or "").strip()
        coverage = (source.get("覆盖状态") or "").strip()
        unknown_ids = [case_id for case_id in linked_ids if case_id not in case_by_id]
        if unknown_ids:
            raise ValueError(f"page-discovery.csv row {index} references unknown case IDs: {unknown_ids}")
        if linked_ids and generated != "是":
            raise ValueError(
                f"page-discovery.csv row {index} links cases but 是否已生成用例 is not 是"
            )
        if generated == "是" and not linked_ids:
            raise ValueError(
                f"page-discovery.csv row {index} is generated but has no 关联用例ID"
            )
        if coverage == "已覆盖" and not linked_ids:
            raise ValueError(
                f"page-discovery.csv row {index} is 已覆盖 but has no 关联用例ID"
            )
        if status == "已实测":
            element_type = source.get("元素类型", "")
            if any(marker in element_type for marker in ("输入", "下拉", "选择", "单选", "开关")):
                if not (source.get("选项取值/输入值") or "").strip():
                    raise ValueError(
                        f"page-discovery.csv row {index} is 已实测 but lacks 选项取值/输入值"
                    )
                if not (source.get("预期/观察行为") or "").strip():
                    raise ValueError(
                        f"page-discovery.csv row {index} is 已实测 but lacks 预期/观察行为"
                    )
        linked_dimensions = {
            case_by_id[case_id].get("DFX维度", "")
            for case_id in linked_ids
            if case_by_id[case_id].get("DFX维度", "")
        }
        declared_dimensions = {
            item.strip()
            for item in re.split(r"[,，;；]+", source.get("适用DFX维度", ""))
            if item.strip()
        }
        missing_dimensions = sorted(linked_dimensions - declared_dimensions)
        if missing_dimensions:
            raise ValueError(
                f"page-discovery.csv row {index} is missing linked case DFX dimensions: {missing_dimensions}"
            )
        story_ids = list(
            OrderedDict.fromkeys(
                case_by_id[case_id].get("Story ID/需求 ID", "")
                for case_id in linked_ids
                if case_by_id[case_id].get("Story ID/需求 ID", "")
            )
        )
        remarks = "\n".join(
            value
            for value in [
                (source.get("未覆盖/待确认原因") or "").strip(),
                (source.get("备注") or "").strip(),
                f"事实状态：{status}",
            ]
            if value
        )
        elements.append(
            {
                "元素 ID": f"EL-{len(elements) + 1:03d}",
                "Story ID/需求 ID": ",".join(story_ids),
                "页面/入口": (source.get("页面/入口") or "").strip(),
                "页面 URL/菜单路径": (source.get("菜单路径/URL") or "").strip(),
                "元素名称/文案": element_name,
                "元素类型": (source.get("元素类型") or "").strip(),
                "交互方式": (source.get("交互方式") or "").strip(),
                "适用DFX维度": (source.get("适用DFX维度") or "").strip(),
                "适用DFX场景": (source.get("适用DFX场景") or "").strip(),
                "前置状态/权限": ";".join(
                    value
                    for value in [
                        (source.get("角色/权限") or "").strip(),
                        (source.get("数据状态") or "").strip(),
                    ]
                    if value
                ),
                "预期行为": (source.get("预期/观察行为") or "").strip(),
                "业务依据/规则来源": (source.get("业务依据/规则来源") or "").strip(),
                "覆盖用例 ID": ",".join(linked_ids),
                "覆盖状态": coverage,
                "发现方式": (source.get("发现方式") or "").strip(),
                "素材来源": (source.get("测试数据来源") or "").strip(),
                "待确认问题/备注": remarks,
            }
        )
    if not elements:
        raise ValueError(f"page-discovery.csv contains no real page element rows: {page_discovery}")
    return elements


def compile_formal_workbook(
    formal_template: Path,
    shards_dir: Path,
    output: Path,
    page_discovery: Path | None = None,
) -> dict[str, list[dict[str, str]]]:
    """Compile all eight formal sheets from validated functional shards and page facts."""
    rows_by_sheet = load_design_shards(shards_dir, formal_template)
    if page_discovery:
        rows_by_sheet["页面元素覆盖清单"] = page_element_rows(
            page_discovery,
            rows_by_sheet["功能测试用例"],
        )
    if not rows_by_sheet["页面元素覆盖清单"]:
        raise ValueError(
            "页面元素覆盖清单 must be provided by page-discovery.csv or existing JSON shards"
        )
    prepare_formal_workbook(formal_template, output)
    workbook = load_workbook(output)
    for sheet_name in FORMAL_SHEETS:
        worksheet = workbook[sheet_name]
        headers = header_map(worksheet)
        for row_index, values in enumerate(rows_by_sheet[sheet_name], start=2):
            write_mapped_row(worksheet, headers, row_index, values)
    remove_workbook_tables_and_refresh_filters(workbook)
    workbook.save(output)
    return rows_by_sheet


def write_mapped_row(ws, headers: dict[str, int], row_index: int, values: dict[str, str]) -> None:
    copy_row_style(ws, 2 if ws.max_row >= 2 else 1, row_index)
    for field, value in values.items():
        column = headers.get(field)
        if column:
            ws.cell(row=row_index, column=column, value=value)


def append_mapped_row(ws, values: dict[str, str]) -> None:
    headers = header_map(ws)
    row_index = ws.max_row + 1 if ws.max_row >= 1 else 2
    write_mapped_row(ws, headers, row_index, values)


PRODUCT_MAP_TEMPLATE_MARKERS = ("示例", "FLOW-DEMO-", "CHG-DEMO-", "TC-DEMO-", "AI_TEST_DEMO")


def remove_rows_containing(ws, needles: list[str]) -> None:
    if not needles:
        return
    for row_index in range(ws.max_row, 1, -1):
        values = ["" if cell.value is None else str(cell.value) for cell in ws[row_index]]
        joined = "\n".join(values)
        if any(needle and needle in joined for needle in needles) or any(
            marker in joined for marker in PRODUCT_MAP_TEMPLATE_MARKERS
        ):
            ws.delete_rows(row_index, 1)


def update_batch_status_paths(
    batch_status: Path,
    batch_id: str | None,
    archive_rel: str,
    import_rel: str,
    function_case_count: int | None = None,
    performance_count: int | None = None,
    page_discovery_completed: bool = False,
    source_shards_validated: bool = False,
) -> list[dict[str, str]]:
    if not batch_status:
        return []
    with batch_status.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        headers = reader.fieldnames or []
        rows = list(reader)
    if not headers:
        raise ValueError(f"batch-status.csv has no header row: {batch_status}")
    required = {"批次ID", "归档路径", "导入文件路径"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"batch-status.csv is missing required finalize columns: {missing}")
    target_rows = [row for row in rows if not batch_id or row.get("批次ID") == batch_id]
    if not target_rows:
        raise ValueError(f"No matching batch row found for batch_id={batch_id!r}")
    changes: list[dict[str, str]] = []
    for row in target_rows:
        changes.append(
            {
                "批次ID": row.get("批次ID", ""),
                "旧归档路径": row.get("归档路径", ""),
                "旧导入文件路径": row.get("导入文件路径", ""),
                "归档路径": archive_rel,
                "导入文件路径": import_rel,
            }
        )
        row["归档路径"] = archive_rel
        row["导入文件路径"] = import_rel
        if "导入文件已生成" in row:
            row["导入文件已生成"] = "是"
        if "状态" in row:
            row["状态"] = "已完成"
        if page_discovery_completed and "页面实探状态" in row:
            row["页面实探状态"] = "已完成"
        if source_shards_validated and "JSON分片状态" in row:
            row["JSON分片状态"] = "已完成"
        if function_case_count is not None and "功能用例数" in row:
            row["功能用例数"] = str(function_case_count)
        if performance_count is not None and "性能场景数" in row:
            row["性能场景数"] = str(performance_count)
        if "最后更新时间" in row:
            row["最后更新时间"] = date.today().isoformat()
        if "下一步动作" in row:
            row["下一步动作"] = "执行一次最终语义Review"
    with batch_status.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
    return changes


def preflight_finalize_metadata(
    batch_status: Path | None,
    batch_id: str | None,
    product_map: Path | None,
    page_discovery: Path | None,
) -> None:
    if product_map and not page_discovery:
        raise ValueError("--product-map requires --page-discovery for product-map sync")
    if page_discovery and not batch_status:
        raise ValueError("--batch-status is required when --page-discovery is provided")
    if page_discovery and not page_discovery.exists():
        raise ValueError(f"page-discovery.csv not found: {page_discovery}")
    if product_map and not product_map.exists():
        raise ValueError(f"product-map.xlsx not found: {product_map}")
    if not batch_status:
        return
    if not batch_status.exists():
        raise ValueError(f"batch-status.csv not found: {batch_status}")
    with batch_status.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        headers = reader.fieldnames or []
        rows = list(reader)
    required = {"批次ID", "归档路径", "导入文件路径"}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"batch-status.csv is missing required finalize columns: {missing}")
    if not any(not batch_id or row.get("批次ID") == batch_id for row in rows):
        raise ValueError(f"No matching batch row found for batch_id={batch_id!r}")


def sync_batch_markdown_paths(batch_status: Path, changes: list[dict[str, str]]) -> None:
    for markdown_name in ["batch-plan.md", "final-review.md", "batch-review.md"]:
        markdown_path = batch_status.resolve().parent / markdown_name
        if not markdown_path.exists():
            continue
        text = markdown_path.read_text(encoding="utf-8-sig")
        for change in changes:
            for old_key, new_key in [("旧归档路径", "归档路径"), ("旧导入文件路径", "导入文件路径")]:
                old_value = change.get(old_key, "")
                new_value = change.get(new_key, "")
                if old_value and new_value:
                    text = text.replace(old_value, new_value)
            if change["归档路径"] not in text or change["导入文件路径"] not in text:
                text += (
                    "\n\n## 交付收口路径\n"
                    f"- {change['批次ID']} 归档路径：{change['归档路径']}\n"
                    f"- {change['批次ID']} 导入文件路径：{change['导入文件路径']}\n"
                )
        markdown_path.write_text(text, encoding="utf-8")


def cleanup_batch_artifacts(batch_status: Path | None) -> None:
    if not batch_status:
        return
    pycache = batch_status.resolve().parent / "artifacts" / "scripts" / "__pycache__"
    if pycache.exists():
        shutil.rmtree(pycache)


def copy_template_if_missing(source: Path, target: Path) -> bool:
    if target.exists():
        return False
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    return True


def write_single_csv_row(path: Path, values: dict[str, str]) -> None:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.reader(fp)
        headers = next(reader, [])
    if not headers:
        raise ValueError(f"CSV template has no header row: {path}")
    row = {header: "" for header in headers}
    for key, value in values.items():
        if key in row:
            row[key] = value
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=headers)
        writer.writeheader()
        writer.writerow(row)


def migrate_page_discovery_fact_status(path: Path) -> None:
    """Add provenance to historical discovery files without claiming old work was tested."""
    if not path.exists():
        return
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        headers = reader.fieldnames or []
        rows = list(reader)
    if not headers:
        raise ValueError(f"page-discovery.csv has no header row: {path}")
    if "事实状态" not in headers:
        insert_at = headers.index("是否已生成用例") if "是否已生成用例" in headers else len(headers)
        headers.insert(insert_at, "事实状态")
    for row in rows:
        if row.get("事实状态"):
            continue
        row["事实状态"] = "待确认" if row.get("覆盖状态") == "待确认" else "页面观察"
    with path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def init_batch_run(
    project_root: Path,
    run_id: str,
    module_path: str,
    batch_id: str,
    product_name: str | None = None,
    large_scope: bool = False,
) -> Path:
    run_dir = project_root / "docs" / "test-assets" / "batch-runs" / run_id
    templates_dir = project_root / "docs" / "test-assets" / "batch-runs" / "templates"
    required_templates = {
        "batch-status.csv": templates_dir / "batch-status-template.csv",
        "page-discovery.csv": templates_dir / "page-discovery-template.csv",
    }
    if large_scope:
        required_templates["batch-plan.md"] = templates_dir / "batch-plan-template.md"
    missing = [str(path) for path in required_templates.values() if not path.exists()]
    if missing:
        raise ValueError(f"Batch template files are missing: {missing}")

    run_dir.mkdir(parents=True, exist_ok=True)
    artifacts_dir = run_dir / "artifacts"
    shards_dir = artifacts_dir / "shards"
    shards_dir.mkdir(parents=True, exist_ok=True)

    created: dict[str, bool] = {}
    for target_name, template_path in required_templates.items():
        created[target_name] = copy_template_if_missing(template_path, run_dir / target_name)

    product, modules = split_module_parts(module_path, product_name)
    level1 = modules[0] if len(modules) > 0 else ""
    level2 = modules[1] if len(modules) > 1 else ""
    level3 = modules[2] if len(modules) > 2 else ""
    leaf_path = ">".join(modules) or module_path

    if created.get("batch-status.csv"):
        write_single_csv_row(
            run_dir / "batch-status.csv",
            {
                "批次ID": batch_id,
                "一级模块": level1,
                "二级菜单": level2,
                "三级菜单/页面域": level3,
                "最小标题路径": leaf_path,
                "状态": "待开始",
                "页面实探状态": "未开始",
                "JSON分片状态": "未开始",
                "功能用例数": "0",
                "性能场景数": "0",
                "最后更新时间": date.today().isoformat(),
                "下一步动作": "开始页面实探并补充 page-discovery.csv",
            },
        )
    if created.get("page-discovery.csv"):
        write_single_csv_row(
            run_dir / "page-discovery.csv",
            {
                "批次ID": batch_id,
                "一级模块": level1,
                "二级菜单": level2,
                "三级菜单/页面域": level3,
                "最小标题路径": leaf_path,
                "菜单路径/URL": leaf_path,
                "发现方式": "浏览器实探/页面资料",
                "事实状态": "待确认",
                "是否已生成用例": "否",
                "覆盖状态": "待确认",
                "备注": "按当前批次页面实探结果补充页面、元素、取值、联动和关联用例",
            },
        )
    else:
        migrate_page_discovery_fact_status(run_dir / "page-discovery.csv")

    init_note = (
        "\n\n## 运行范围\n"
        f"- 产品/系统：{product}\n"
        f"- 模块路径：{leaf_path}\n"
        f"- 批次ID：{batch_id}\n"
        "- 执行要求：先补全 page-discovery.csv，再生成测试设计、导入文件和 batch-status.csv 覆盖数据。\n"
    )
    for markdown_name in ["batch-plan.md"]:
        markdown_path = run_dir / markdown_name
        if not markdown_path.exists():
            continue
        text = markdown_path.read_text(encoding="utf-8-sig")
        if "## 运行范围" not in text:
            markdown_path.write_text(text.rstrip() + init_note, encoding="utf-8")

    print(f"Ready: {run_dir}")
    return run_dir


def split_module_parts(module_path: str, product_name: str | None = None) -> tuple[str, list[str]]:
    if product_name:
        return product_name, canonical_module_parts(module_path, product_name)
    parts = canonical_module_parts(module_path)
    if len(parts) >= 4:
        return parts[0], parts[1:]
    return (parts[0] if parts else "产品"), parts


def sync_product_map(
    product_map: Path,
    formal_workbook: Path,
    page_discovery: Path,
    module_path: str,
    archive_path: str,
    product_name: str | None = None,
) -> None:
    today = date.today().isoformat()
    product, modules = split_module_parts(module_path, product_name)
    module_label = ">".join(modules) or module_path
    level1 = modules[0] if len(modules) > 0 else ""
    level2 = modules[1] if len(modules) > 1 else ""
    level3 = modules[2] if len(modules) > 2 else ""

    formal_wb = load_workbook(formal_workbook, data_only=True)
    function_ws = formal_wb[FORMAL_FUNCTION_SHEET]
    function_rows = non_empty_rows(function_ws, header_map(function_ws))

    with page_discovery.open("r", encoding="utf-8-sig", newline="") as fp:
        discovery_rows = [row for row in csv.DictReader(fp) if any((value or "").strip() for value in row.values())]

    wb = load_workbook(product_map)
    for ws in wb.worksheets:
        remove_rows_containing(ws, [module_label, archive_path])

    pages = sorted({row.get("页面/入口", "") for row in discovery_rows if row.get("页面/入口")})
    for page in pages or [level3 or level2 or level1 or module_label]:
        append_mapped_row(
            wb["产品模块地图"],
            {
                "产品/系统": product,
                "一级模块": level1,
                "二级模块": level2,
                "三级模块": level3,
                "页面/入口": page,
                "菜单路径/URL": module_label,
                "模块功能摘要": f"{module_label} 页面功能、页面元素和测试资产已同步",
                "归档测试设计路径": archive_path,
                "覆盖状态": "已覆盖",
                "最后更新时间": today,
                "待确认问题": "无",
            },
        )

    linked_create_ids = ";".join(row.get("用例 ID", "") for row in function_rows[:3] if row.get("用例 ID"))
    append_mapped_row(
        wb["业务对象地图"],
        {
            "产品/系统": product,
            "业务对象": level3 or level2 or module_label,
            "来源模块": module_label,
            "消费模块": module_label,
            "关键字段": "名称、状态、配置项",
            "关键状态": "新增、编辑、删除、查询",
            "状态生产者": module_label,
            "状态消费者": module_label,
            "创建用例ID": linked_create_ids,
            "归档测试设计路径": archive_path,
            "待确认问题": "无",
        },
    )
    append_mapped_row(
        wb["业务链路地图"],
        {
            "链路ID": f"FLOW-{safe_filename(module_label)}",
            "链路名称": f"{module_label}主流程",
            "起始模块": module_label,
            "结束模块": module_label,
            "业务对象": level3 or level2 or module_label,
            "关键状态流转": "进入页面>新增/编辑/查询/删除>校验结果",
            "主流程用例ID": linked_create_ids,
            "依赖测试数据": "AI_TEST 测试数据",
            "风险点": "页面联动、数据一致性、权限状态",
            "归档测试设计路径": archive_path,
        },
    )

    for row in discovery_rows:
        append_mapped_row(
            wb["页面元素地图"],
            {
                "产品/系统": product,
                "模块": module_label,
                "页面/入口": row.get("页面/入口", ""),
                "菜单路径/URL": row.get("菜单路径/URL", "") or module_label,
                "元素名称/文案": row.get("元素名称/文案", ""),
                "元素类型": row.get("元素类型", ""),
                "交互方式": row.get("交互方式", ""),
                "前置状态/权限": row.get("角色/权限", ""),
                "关联用例ID": row.get("关联用例ID", ""),
                "覆盖状态": row.get("覆盖状态", ""),
                "发现来源": " / ".join(part for part in [row.get("事实状态", ""), row.get("发现方式", "")] if part),
                "最后更新时间": today,
                "备注": row.get("备注", ""),
            },
        )

    for row in function_rows:
        case_id = row.get("用例 ID", "")
        if not case_id:
            continue
        append_mapped_row(
            wb["用例资产索引"],
            {
                "产品/系统": product,
                "模块": module_label,
                "功能点": row.get("功能点", ""),
                "用例ID": case_id,
                "用例标题": row.get("用例标题", ""),
                "测试类型": row.get("测试类型", "功能测试") or "功能测试",
                "执行方式": "手动",
                "是否可复用为前置条件": "否",
                "是否跨模块": "否",
                "关联业务对象": level3 or level2 or module_label,
                "关联业务链路": f"{module_label}主流程",
                "归档测试设计路径": archive_path,
                "最后更新时间": today,
            },
        )

    for function_point in sorted({row.get("功能点", "") for row in function_rows if row.get("功能点")}):
        ids = ";".join(row.get("用例 ID", "") for row in function_rows if row.get("功能点") == function_point and row.get("用例 ID"))
        append_mapped_row(
            wb["模块能力索引"],
            {
                "产品/系统": product,
                "模块": module_label,
                "功能点": function_point,
                "能力/数据对象": level3 or level2 or module_label,
                "能力描述": f"{function_point} 已形成测试资产",
                "关键状态": "正常、异常、边界、权限/状态",
                "可复用前置条件": "按归档测试设计准备测试数据",
                "关联用例ID": ids,
                "归档测试设计路径": archive_path,
                "限制/待确认问题": "无",
                "最后更新时间": today,
            },
        )

    append_mapped_row(
        wb["跨模块依赖关系"],
        {
            "产品/系统": product,
            "当前模块": module_label,
            "依赖模块": "待识别",
            "依赖业务对象": level3 or level2 or module_label,
            "依赖功能点/能力": "页面入口、权限、数据状态",
            "依赖类型": "待确认",
            "引用用例ID": linked_create_ids,
            "当前模块用例ID": linked_create_ids,
            "使用方式": "作为页面实探和测试数据准备依据",
            "风险/待确认问题": "无明确跨模块依赖时保持待确认",
            "最后更新时间": today,
        },
    )
    append_mapped_row(
        wb["可复用测试数据"],
        {
            "产品/系统": product,
            "模块": module_label,
            "数据名称": "AI_TEST 测试数据",
            "数据类型": "页面实探数据",
            "数据内容/构造方式": "使用带 AI_TEST 或 CODEX_TEST 标识的数据",
            "适用用例ID": linked_create_ids,
            "是否可复用": "是",
            "限制/清理方式": "仅操作本次创建的数据，交付后按环境规则清理",
            "最后更新时间": today,
        },
    )
    append_mapped_row(
        wb["变更影响分析"],
        {
            "产品/系统": product,
            "变更模块": module_label,
            "变更点": "新增或更新模块测试设计",
            "影响模块": module_label,
            "影响业务对象": level3 or level2 or module_label,
            "影响用例ID": linked_create_ids,
            "回归建议": "回归页面入口、核心流程、异常边界、权限状态和数据一致性",
            "风险等级": "中",
            "最后更新时间": today,
        },
    )
    append_mapped_row(
        wb["变更记录"],
        {
            "变更日期": today,
            "变更类型": "测试资产同步",
            "变更内容": f"同步 {module_label} 测试设计、页面元素和用例资产",
            "影响范围": module_label,
            "关联归档路径": archive_path,
            "是否已同步产品版图": "是",
            "备注": "由 sync-product-map/finalize-deliverables 统一维护",
        },
    )
    remove_workbook_tables_and_refresh_filters(wb)
    wb.save(product_map)


def finalize_deliverables(
    project_root: Path,
    formal_workbook: Path,
    import_workbook: Path,
    module_path: str,
    batch_status: Path | None = None,
    batch_id: str | None = None,
    product_map: Path | None = None,
    page_discovery: Path | None = None,
    product_name: str | None = None,
    source_shards_validated: bool = False,
) -> dict[str, Path]:
    project_root = project_root.resolve()
    preflight_finalize_metadata(batch_status, batch_id, product_map, page_discovery)
    _, formal_name, import_name = deliverable_names(module_path, product_name)

    module_archive = project_root / "docs" / "test-assets" / "modules" / formal_name
    import_archive = project_root / "docs" / "test-assets" / "imports" / import_name
    current_copy = project_root / "docs" / "test-design" / "current" / formal_name
    deliverable_formal = project_root / "docs" / "test-design" / "deliverables" / formal_name
    deliverable_import = project_root / "docs" / "test-design" / "deliverables" / import_name

    with tempfile.TemporaryDirectory(prefix="test-design-publish-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        prepared_formal = temp_dir / formal_name
        prepared_import = temp_dir / import_name
        apply_formal_workbook_styles(formal_workbook, prepared_formal)
        shutil.copy2(import_workbook, prepared_import)
        import_wb = load_workbook(prepared_import)
        remove_workbook_tables_and_refresh_filters(import_wb)
        import_wb.save(prepared_import)
        atomic_publish_copies(
            [
                (prepared_formal, module_archive),
                (prepared_formal, current_copy),
                (prepared_formal, deliverable_formal),
                (prepared_import, import_archive),
                (prepared_import, deliverable_import),
            ]
        )

    target_dirs = {
        module_archive.parent,
        current_copy.parent,
        deliverable_formal.parent,
        import_archive.parent,
    }
    cleanup_excel_lock_files(target_dirs)
    legacy_names = legacy_repeated_leaf_names(module_path, product_name)
    if legacy_names:
        legacy_formal, legacy_import = legacy_names
        for directory, name in [
            (module_archive.parent, legacy_formal),
            (current_copy.parent, legacy_formal),
            (deliverable_formal.parent, legacy_formal),
            (import_archive.parent, legacy_import),
            (deliverable_import.parent, legacy_import),
        ]:
            stale = directory / name
            if stale.exists():
                stale.unlink()

    if batch_status:
        formal_wb = load_workbook(module_archive, read_only=True, data_only=True)
        function_case_count = len(non_empty_rows(
            formal_wb["功能测试用例"],
            header_map(formal_wb["功能测试用例"]),
        ))
        performance_count = len(non_empty_rows(
            formal_wb["性能测试设计"],
            header_map(formal_wb["性能测试设计"]),
        ))
        formal_wb.close()
        changes = update_batch_status_paths(
            batch_status,
            batch_id,
            relative_project_path(project_root, module_archive),
            relative_project_path(project_root, import_archive),
            function_case_count,
            performance_count,
            page_discovery is not None,
            source_shards_validated,
        )
        sync_batch_markdown_paths(batch_status, changes)
        cleanup_batch_artifacts(batch_status)
    if product_map and page_discovery:
        sync_product_map(
            product_map,
            module_archive,
            page_discovery,
            module_path,
            relative_project_path(project_root, module_archive),
            product_name,
        )
    return {
        "formal": deliverable_formal,
        "import": deliverable_import,
        "current_formal": current_copy,
        "formal_archive": module_archive,
        "import_archive": import_archive,
        "deliverable_formal": deliverable_formal,
    }


def run_python_script(script: Path, args: list[str]) -> None:
    completed = subprocess.run([sys.executable, str(script), *args], check=False)
    if completed.returncode:
        raise SystemExit(completed.returncode)


def complete_deliverables(
    project_root: Path,
    formal_workbook: Path,
    import_template: Path,
    module_path: str,
    import_workbook: Path | None = None,
    batch_status: Path | None = None,
    batch_id: str | None = None,
    product_map: Path | None = None,
    page_discovery: Path | None = None,
    product_name: str | None = None,
    scripts_path: Path | None = None,
    source_shards_validated: bool = False,
) -> dict[str, Path]:
    project_root = project_root.resolve()
    preflight_finalize_metadata(batch_status, batch_id, product_map, page_discovery)
    script_dir = Path(__file__).resolve().parent
    _, _, import_name = deliverable_names(module_path, product_name)
    if scripts_path and scripts_path.exists():
        run_python_script(script_dir / "validate-generated-python-scripts.py", ["--path", str(scripts_path)])

    with tempfile.TemporaryDirectory(prefix="test-design-complete-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        staged_formal = temp_dir / "formal.xlsx"
        staged_import = temp_dir / import_name
        apply_formal_workbook_styles(formal_workbook, staged_formal)
        generate_import_workbook(staged_formal, import_template, staged_import, module_path, product_name)
        staged_validator_args = ["--workbook", str(staged_formal), "--import-workbook", str(staged_import)]
        if batch_status:
            staged_validator_args.extend(["--batch-status", str(batch_status)])
        if page_discovery:
            staged_validator_args.extend(["--page-discovery", str(page_discovery)])
        run_python_script(script_dir / "validate-test-design-deliverable.py", staged_validator_args)
        paths = finalize_deliverables(
            project_root,
            staged_formal,
            staged_import,
            module_path,
            batch_status,
            batch_id,
            product_map,
            page_discovery,
            product_name,
            source_shards_validated,
        )

    if import_workbook and import_workbook.resolve() != paths["import_archive"].resolve():
        atomic_publish_copies([(paths["import_archive"], import_workbook)])

    if product_map and page_discovery:
        validator_args = [
            "--workbook",
            str(paths["formal_archive"]),
            "--import-workbook",
            str(paths["import_archive"]),
            "--product-map",
            str(product_map),
            "--page-discovery",
            str(page_discovery),
        ]
        if batch_status:
            validator_args.extend(["--batch-status", str(batch_status)])
        run_python_script(script_dir / "validate-test-design-deliverable.py", validator_args)
    return paths


def compile_deliverables(
    project_root: Path,
    shards_dir: Path,
    formal_template: Path,
    import_template: Path,
    module_path: str,
    batch_status: Path | None = None,
    batch_id: str | None = None,
    product_map: Path | None = None,
    page_discovery: Path | None = None,
    product_name: str | None = None,
) -> dict[str, Path]:
    """Compile validated shards and atomically publish the canonical workbook pair."""
    preflight_finalize_metadata(batch_status, batch_id, product_map, page_discovery)
    with tempfile.TemporaryDirectory(prefix="test-design-compile-") as temp_dir_name:
        formal_draft = Path(temp_dir_name) / "formal-from-shards.xlsx"
        compile_formal_workbook(
            formal_template=formal_template,
            shards_dir=shards_dir,
            output=formal_draft,
            page_discovery=page_discovery,
        )
        return complete_deliverables(
            project_root=project_root,
            formal_workbook=formal_draft,
            import_template=import_template,
            module_path=module_path,
            batch_status=batch_status,
            batch_id=batch_id,
            product_map=product_map,
            page_discovery=page_discovery,
            product_name=product_name,
            source_shards_validated=True,
        )


def generate_import_workbook(
    formal_workbook: Path,
    import_template: Path,
    output: Path,
    module_path: str,
    product_name: str | None = None,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(import_template, output)

    formal_wb = load_workbook(formal_workbook)
    if FORMAL_FUNCTION_SHEET not in formal_wb.sheetnames:
        raise ValueError(f"Formal workbook is missing sheet: {FORMAL_FUNCTION_SHEET}")
    function_ws = formal_wb[FORMAL_FUNCTION_SHEET]
    function_headers = header_map(function_ws)

    import_wb = load_workbook(output)
    import_ws = import_wb[import_wb.sheetnames[0]]
    import_headers = header_map(import_ws)
    clear_data_rows(import_ws)

    modules = import_module_names(module_path, product_name)
    write_row = 2
    for row_index in range(2, function_ws.max_row + 1):
        case = row_dict(function_ws, function_headers, row_index)
        if not case.get("用例 ID") and not case.get("用例标题"):
            continue
        copy_row_style(import_ws, 2 if import_ws.max_row >= 2 else 1, write_row)
        dfx_dimension = case.get("DFX维度", "")
        dfx_scenario = case.get("DFX场景", "")
        tags = ";".join(part for part in [case.get("模块", ""), case.get("功能点", ""), dfx_dimension, dfx_scenario] if part)
        dfx_note = f"DFX覆盖：{dfx_dimension}-{dfx_scenario}" if dfx_dimension and dfx_scenario else ""
        remarks = "\n".join(part for part in [dfx_note, case.get("备注", "")] if part)
        mapped = {
            "一级模块名称": modules[0],
            "二级模块名称": modules[1],
            "三级模块名称": modules[2],
            "四级模块名称": modules[3],
            "五级模块名称": modules[4],
            "测试用例序号": str(write_row - 1),
            "测试用例名称": case.get("用例标题", ""),
            "测试步骤描述": case.get("操作步骤", ""),
            "测试步骤预期结果": case.get("预期结果", ""),
            "测试类型": normalize_test_type(case.get("测试类型", "")),
            "测试用例级别": normalize_case_level(case.get("优先级", "")),
            "执行方式": execution_mode(case),
            "测试用例说明": case.get("功能点", ""),
            "前置条件": case.get("前置条件", ""),
            "标签": tags,
            "备注": remarks,
        }
        for field in IMPORT_AUTO_FIELDS:
            mapped[field] = ""
        for field, value in mapped.items():
            column = import_headers.get(field)
            if column:
                import_ws.cell(row=write_row, column=column, value=value)
        set_wrap(import_ws, import_headers, write_row, IMPORT_MULTILINE_FIELDS)
        import_ws.row_dimensions[write_row].height = max(import_ws.row_dimensions[write_row].height or 18, 60)
        write_row += 1

    template_wb = load_workbook(import_template)
    apply_template_workbook_format(import_wb, template_wb)
    for row_index in range(2, import_ws.max_row + 1):
        set_wrap(import_ws, import_headers, row_index, IMPORT_MULTILINE_FIELDS)
        import_ws.row_dimensions[row_index].height = max(import_ws.row_dimensions[row_index].height or 18, 60)
    remove_workbook_tables_and_refresh_filters(import_wb)
    import_wb.save(output)


def apply_formal_workbook_styles(workbook: Path, output: Path | None = None, template: Path | None = None) -> None:
    target = output or workbook
    if output:
        output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(workbook, output)
    wb = load_workbook(target)
    template_path = template or (Path(__file__).resolve().parents[1] / "docs" / "test-design" / "codebuddy-test-design-template.xlsx")
    if template_path.exists():
        template_wb = load_workbook(template_path)
        apply_template_workbook_format(wb, template_wb)
    for sheet_name, fields in FORMAL_MULTILINE_FIELDS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = header_map(ws)
        for row_index in range(2, ws.max_row + 1):
            set_wrap(ws, headers, row_index, fields)
            ws.row_dimensions[row_index].height = max(ws.row_dimensions[row_index].height or 18, 60)
    remove_workbook_tables_and_refresh_filters(wb)
    wb.save(target)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate or normalize test design Excel deliverables.")
    sub = parser.add_subparsers(dest="command", required=True)

    gen = sub.add_parser("generate-import", help="Generate test-system import workbook from formal test design workbook.")
    gen.add_argument("--formal-workbook", required=True, type=Path)
    gen.add_argument("--import-template", required=True, type=Path)
    gen.add_argument("--output", required=True, type=Path)
    gen.add_argument("--module-path", required=True)
    gen.add_argument("--product-name")

    style = sub.add_parser("fix-formal-styles", help="Apply required multiline wrapping styles to a formal workbook.")
    style.add_argument("--workbook", required=True, type=Path)
    style.add_argument("--output", type=Path)
    style.add_argument("--template", type=Path)

    prepare = sub.add_parser("prepare-formal", help="Copy the formal template and clear all example data before filling.")
    prepare.add_argument("--template", required=True, type=Path)
    prepare.add_argument("--output", required=True, type=Path)

    init = sub.add_parser("init-batch-run", help="Create a standard batch-run ledger from templates before page discovery.")
    init.add_argument("--project-root", required=True, type=Path)
    init.add_argument("--run-id", required=True)
    init.add_argument("--module-path", required=True)
    init.add_argument("--batch-id", default="BATCH-001")
    init.add_argument("--product-name")
    init.add_argument("--large-scope", action="store_true", help="Also create a compact batch-plan.md for multi-page/module work.")

    finalize = sub.add_parser("finalize-deliverables", help="Copy validated workbooks to current/deliverables/internal archives and update batch-status paths.")
    finalize.add_argument("--project-root", required=True, type=Path)
    finalize.add_argument("--formal-workbook", required=True, type=Path)
    finalize.add_argument("--import-workbook", required=True, type=Path)
    finalize.add_argument("--module-path", required=True)
    finalize.add_argument("--batch-status", type=Path)
    finalize.add_argument("--batch-id")
    finalize.add_argument("--product-map", type=Path)
    finalize.add_argument("--page-discovery", type=Path)
    finalize.add_argument("--product-name")

    complete = sub.add_parser("complete-deliverables", help="One-shot precheck, style, import generation, finalize, and delivery validation.")
    complete.add_argument("--project-root", required=True, type=Path)
    complete.add_argument("--formal-workbook", required=True, type=Path)
    complete.add_argument("--import-template", required=True, type=Path)
    complete.add_argument("--module-path", required=True)
    complete.add_argument("--import-workbook", type=Path)
    complete.add_argument("--batch-status", type=Path)
    complete.add_argument("--batch-id")
    complete.add_argument("--product-map", type=Path)
    complete.add_argument("--page-discovery", type=Path)
    complete.add_argument("--product-name")
    complete.add_argument("--scripts-path", type=Path)

    compile_cmd = sub.add_parser(
        "compile-deliverables",
        help="Compile existing per-function JSON shards into all eight sheets and atomically publish both Excel files.",
    )
    compile_cmd.add_argument("--project-root", required=True, type=Path)
    compile_cmd.add_argument("--shards-dir", required=True, type=Path)
    compile_cmd.add_argument("--formal-template", required=True, type=Path)
    compile_cmd.add_argument("--import-template", required=True, type=Path)
    compile_cmd.add_argument("--module-path", required=True)
    compile_cmd.add_argument("--batch-status", type=Path)
    compile_cmd.add_argument("--batch-id")
    compile_cmd.add_argument("--product-map", type=Path)
    compile_cmd.add_argument("--page-discovery", type=Path)
    compile_cmd.add_argument("--product-name")

    sync = sub.add_parser("sync-product-map", help="Sync product-map.xlsx from a formal workbook and page-discovery.csv.")
    sync.add_argument("--product-map", required=True, type=Path)
    sync.add_argument("--formal-workbook", required=True, type=Path)
    sync.add_argument("--page-discovery", required=True, type=Path)
    sync.add_argument("--module-path", required=True)
    sync.add_argument("--archive-path", required=True)
    sync.add_argument("--product-name")

    args = parser.parse_args()
    if args.command == "generate-import":
        generate_import_workbook(args.formal_workbook, args.import_template, args.output, args.module_path, args.product_name)
    elif args.command == "fix-formal-styles":
        apply_formal_workbook_styles(args.workbook, args.output, args.template)
    elif args.command == "prepare-formal":
        prepare_formal_workbook(args.template, args.output)
    elif args.command == "init-batch-run":
        run_dir = init_batch_run(
            args.project_root,
            args.run_id,
            args.module_path,
            args.batch_id,
            args.product_name,
            args.large_scope,
        )
        print(f"RUN_DIR={run_dir.resolve()}")
    elif args.command == "finalize-deliverables":
        if args.page_discovery and not args.batch_status:
            raise SystemExit(
                "ERROR: --batch-status is required when --page-discovery is provided. "
                "Keep batch-status.csv and page-discovery.csv in the same run directory."
            )
        paths = finalize_deliverables(
            args.project_root,
            args.formal_workbook,
            args.import_workbook,
            args.module_path,
            args.batch_status,
            args.batch_id,
            args.product_map,
            args.page_discovery,
            args.product_name,
        )
        print(f"FORMAL_WORKBOOK={paths['formal'].resolve()}")
        print(f"IMPORT_WORKBOOK={paths['import'].resolve()}")
    elif args.command == "complete-deliverables":
        if args.page_discovery and not args.batch_status:
            raise SystemExit(
                "ERROR: --batch-status is required when --page-discovery is provided. "
                "Keep batch-status.csv and page-discovery.csv in the same run directory."
            )
        paths = complete_deliverables(
            args.project_root,
            args.formal_workbook,
            args.import_template,
            args.module_path,
            args.import_workbook,
            args.batch_status,
            args.batch_id,
            args.product_map,
            args.page_discovery,
            args.product_name,
            args.scripts_path,
        )
        print(f"FORMAL_WORKBOOK={paths['formal'].resolve()}")
        print(f"IMPORT_WORKBOOK={paths['import'].resolve()}")
    elif args.command == "compile-deliverables":
        if args.page_discovery and not args.batch_status:
            raise SystemExit(
                "ERROR: --batch-status is required when --page-discovery is provided. "
                "Keep batch-status.csv and page-discovery.csv in the same run directory."
            )
        paths = compile_deliverables(
            project_root=args.project_root,
            shards_dir=args.shards_dir,
            formal_template=args.formal_template,
            import_template=args.import_template,
            module_path=args.module_path,
            batch_status=args.batch_status,
            batch_id=args.batch_id,
            product_map=args.product_map,
            page_discovery=args.page_discovery,
            product_name=args.product_name,
        )
        print(f"FORMAL_WORKBOOK={paths['formal'].resolve()}")
        print(f"IMPORT_WORKBOOK={paths['import'].resolve()}")
    elif args.command == "sync-product-map":
        sync_product_map(
            args.product_map,
            args.formal_workbook,
            args.page_discovery,
            args.module_path,
            args.archive_path,
            args.product_name,
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
