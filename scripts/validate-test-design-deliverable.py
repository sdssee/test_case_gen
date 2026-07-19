# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import math
import re
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from test_design.formal_assembler import FUNCTION_SHEET, SHEETS
from test_design.excel_utils import header_map


def _nonempty_rows(ws) -> list[int]:
    return [row for row in range(2, ws.max_row + 1) if any(ws.cell(row, col).value not in (None, "") for col in range(1, ws.max_column + 1))]


def _assert_readable_rows(ws, headers: dict[str, int], rows: list[int], multiline_headers: set[str]) -> None:
    for row in rows:
        lines = 1
        for header in multiline_headers & set(headers):
            column = headers[header]
            dimension = ws.column_dimensions.get(ws.cell(1, column).column_letter)
            width = (dimension.width if dimension else ws.sheet_format.defaultColWidth) or 10
            text = str(ws.cell(row, column).value or "")
            visual = sum(max(1, math.ceil(len(line) / max(4, int(width)))) for line in text.splitlines()) if text else 1
            lines = max(lines, visual)
        required = float(max(36, min(360, 8 + lines * 16)))
        actual = float(ws.row_dimensions[row].height or ws.sheet_format.defaultRowHeight or 15)
        if actual + 1 < required:
            raise ValueError(f"{ws.title} row {row} is visually clipped: height {actual} < {required}")


def validate_workbook(path: Path) -> int:
    if not path.is_file():
        raise ValueError(f"workbook not found: {path}")
    with zipfile.ZipFile(path) as archive:
        if any(name.startswith("xl/tables/table") for name in archive.namelist()):
            raise ValueError("workbook must not retain Excel Table parts")
    workbook = load_workbook(path, data_only=False)
    if workbook.sheetnames != SHEETS:
        raise ValueError(f"formal workbook must contain exactly the standard 8 sheets: {SHEETS}")
    ws = workbook[FUNCTION_SHEET]
    headers = header_map(ws)
    required = {"用例 ID", "功能点", "用例标题", "操作步骤", "预期结果"}
    if not required <= set(headers):
        raise ValueError(f"function-case sheet lacks headers: {sorted(required - set(headers))}")
    rows = _nonempty_rows(ws)
    if not rows:
        raise ValueError("function-case sheet has no cases")
    if rows != list(range(2, 2 + len(rows))):
        raise ValueError("function-case sheet contains blank rows between cases")
    ids: set[str] = set()
    signatures: set[tuple[str, str]] = set()
    for row in rows:
        values = {name: str(ws.cell(row, col).value or "").strip() for name, col in headers.items()}
        case_id = values["用例 ID"]
        if not case_id or case_id in ids:
            raise ValueError(f"empty or duplicate case id at row {row}")
        ids.add(case_id)
        if not values["功能点"] or not values["用例标题"]:
            raise ValueError(f"case test point or title is empty at row {row}")
        steps = values["操作步骤"]
        expected = values["预期结果"]
        if not steps or not expected or not re.search(r"(?m)^1\.\s+", steps) or not re.search(r"(?m)^1\.\s+", expected):
            raise ValueError(f"steps/expected are empty or unnumbered at row {row}")
        signature = (steps, expected)
        if signature in signatures:
            raise ValueError(f"duplicate steps and expected results at row {row}")
        signatures.add(signature)
        text = "\n".join(values.values()).lower()
        if any(marker in text for marker in ("截图", "screenshot", "uid", "fact_id", "todo", "tbd")):
            raise ValueError(f"non-executable/internal wording found at row {row}")
    _assert_readable_rows(ws, headers, rows, {"前置条件", "测试数据", "操作步骤", "预期结果", "备注"})
    matrix = workbook["测试场景矩阵"]
    matrix_headers = header_map(matrix)
    matrix_rows = _nonempty_rows(matrix)
    by_case = {
        str(matrix.cell(row, matrix_headers["场景 ID"]).value or "").strip(): row
        for row in matrix_rows
    }
    for row in rows:
        case_id = str(ws.cell(row, headers["用例 ID"]).value or "").strip()
        matrix_row = by_case.get(case_id)
        if not matrix_row:
            raise ValueError(f"scenario matrix is missing case {case_id}")
        for case_header, matrix_header in (("功能点", "功能点"), ("优先级", "优先级"), ("DFX维度", "DFX维度"), ("DFX场景", "DFX场景")):
            case_value = str(ws.cell(row, headers[case_header]).value or "").strip()
            matrix_value = str(matrix.cell(matrix_row, matrix_headers[matrix_header]).value or "").strip()
            if case_value != matrix_value:
                raise ValueError(f"scenario matrix {case_id} {matrix_header} differs from function cases")
    coverage = workbook["页面元素覆盖清单"]
    coverage_headers = header_map(coverage)
    for row in _nonempty_rows(coverage):
        status = str(coverage.cell(row, coverage_headers["覆盖状态"]).value or "").strip()
        note = str(coverage.cell(row, coverage_headers["待确认问题/备注"]).value or "").strip()
        if status == "未覆盖" and not note:
            name = str(coverage.cell(row, coverage_headers["元素名称/文案"]).value or "").strip()
            raise ValueError(f"interactive element {name!r} is uncovered without a real blocker")
    return len(rows)


def validate_import(formal_path: Path, path: Path, expected_count: int) -> None:
    formal = load_workbook(formal_path, data_only=False)
    formal_ws = formal[FUNCTION_SHEET]
    formal_headers = header_map(formal_ws)
    formal_rows = _nonempty_rows(formal_ws)
    workbook = load_workbook(path, data_only=False)
    ws = workbook[workbook.sheetnames[0]]
    headers = header_map(ws)
    required = {"测试用例名称", "测试步骤描述", "测试步骤预期结果"}
    if not required <= set(headers):
        raise ValueError(f"import workbook lacks headers: {sorted(required - set(headers))}")
    rows = _nonempty_rows(ws)
    if len(rows) != expected_count:
        raise ValueError(f"import workbook case count {len(rows)} != formal workbook {expected_count}")
    if rows != list(range(2, 2 + len(rows))):
        raise ValueError("import workbook contains blank rows between cases")
    mappings = {
        "用例标题": "测试用例名称",
        "操作步骤": "测试步骤描述",
        "预期结果": "测试步骤预期结果",
        "前置条件": "前置条件",
    }
    for formal_row, import_row in zip(formal_rows, rows):
        for formal_header, import_header in mappings.items():
            formal_value = str(formal_ws.cell(formal_row, formal_headers[formal_header]).value or "").strip()
            import_value = str(ws.cell(import_row, headers[import_header]).value or "").strip()
            if formal_value != import_value:
                raise ValueError(
                    f"import row {import_row} {import_header} does not match formal case row {formal_row}"
                )
    _assert_readable_rows(ws, headers, rows, {"测试步骤描述", "测试步骤预期结果", "测试用例说明", "前置条件", "备注"})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--import-workbook", type=Path)
    args = parser.parse_args()
    count = validate_workbook(args.workbook)
    if args.import_workbook:
        validate_import(args.workbook, args.import_workbook, count)
    print(f"OK: deliverables are valid ({count} function cases).")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
