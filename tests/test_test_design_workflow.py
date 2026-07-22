# -*- coding: utf-8 -*-
from __future__ import annotations

import importlib.util
import csv
import os
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from scripts import test_design_excel_tools as excel_tools


REPO_ROOT = Path(__file__).resolve().parents[1]
FORMAL_TEMPLATE = REPO_ROOT / "docs" / "test-design" / "codebuddy-test-design-template.xlsx"
IMPORT_TEMPLATE = REPO_ROOT / "docs" / "test-design" / "测试用例模板.xlsx"
PRODUCT_MAP_TEMPLATE = REPO_ROOT / "docs" / "test-assets" / "product-map.xlsx"


def load_deliverable_validator():
    path = REPO_ROOT / "scripts" / "validate-test-design-deliverable.py"
    spec = importlib.util.spec_from_file_location("deliverable_validator", path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def load_shard_validator():
    path = REPO_ROOT / "scripts" / "validate-generated-python-scripts.py"
    spec = importlib.util.spec_from_file_location("shard_validator", path)
    assert spec and spec.loader
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class TestDesignWorkflowTests(unittest.TestCase):
    def test_product_map_template_has_no_builtin_product_facts(self) -> None:
        workbook = load_workbook(PRODUCT_MAP_TEMPLATE, data_only=False)
        self.assertEqual(len(workbook.sheetnames), 10)
        for worksheet in workbook.worksheets:
            headers = [cell.value for cell in worksheet[1]]
            self.assertTrue(any(value not in (None, "") for value in headers), worksheet.title)
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                self.assertFalse(
                    any(value not in (None, "") for value in row),
                    f"{worksheet.title} contains built-in product facts: {row}",
                )

    def test_blank_product_map_template_accepts_real_sync(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            formal = temp / "formal.xlsx"
            product_map = temp / "product-map.xlsx"
            page_discovery = temp / "page-discovery.csv"
            self.build_minimal_valid_formal(formal)
            shutil.copy2(PRODUCT_MAP_TEMPLATE, product_map)

            template = REPO_ROOT / "docs" / "test-assets" / "batch-runs" / "templates" / "page-discovery-template.csv"
            with template.open("r", encoding="utf-8-sig", newline="") as fp:
                headers = next(csv.reader(fp))
            row = {header: "" for header in headers}
            row.update(
                {
                    "批次ID": "BATCH-001",
                    "一级模块": "一级功能",
                    "二级菜单": "目标页面",
                    "最小标题路径": "一级功能>目标页面",
                    "页面/入口": "目标页面",
                    "菜单路径/URL": "一级功能>目标页面",
                    "发现方式": "页面实探",
                    "元素名称/文案": "执行按钮",
                    "元素类型": "按钮",
                    "交互方式": "点击",
                    "事实状态": "已实测",
                    "是否已生成用例": "是",
                    "关联用例ID": "TC-001",
                    "覆盖状态": "已覆盖",
                }
            )
            with page_discovery.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=headers)
                writer.writeheader()
                writer.writerow(row)

            excel_tools.sync_product_map(
                product_map=product_map,
                formal_workbook=formal,
                page_discovery=page_discovery,
                module_path="一级功能>目标页面",
                archive_path="docs/test-assets/modules/一级功能_目标页面_测试设计.xlsx",
            )

            workbook = load_workbook(product_map, data_only=True)
            for worksheet in workbook.worksheets:
                values = [
                    "" if value is None else str(value)
                    for row_values in worksheet.iter_rows(min_row=2, values_only=True)
                    for value in row_values
                ]
                self.assertTrue(any(value for value in values), worksheet.title)
                self.assertFalse(any("示例" in value for value in values), worksheet.title)

    def build_minimal_valid_formal(self, output: Path) -> None:
        excel_tools.prepare_formal_workbook(FORMAL_TEMPLATE, output)
        workbook = load_workbook(output)
        rows = {
            "测试场景矩阵": {
                "场景 ID": "SC-001",
                "功能点": "连接测试",
                "测试维度": "功能测试",
                "DFX维度": "DFT功能",
                "DFX场景": "正向流程",
                "测试对象/页面元素": "连接按钮",
                "输入数据/状态条件": "有效内网地址",
                "观察点": "连接结果",
                "风险等级": "中",
                "优先级": "P1",
                "是否生成用例": "是",
            },
            "功能测试用例": {
                "用例 ID": "TC-001",
                "模块": "一级模块-二级页面",
                "功能点": "连接测试",
                "用例标题": "连接测试-使用有效内网地址执行",
                "优先级": "P1",
                "测试类型": "功能测试",
                "DFX维度": "DFT功能",
                "DFX场景": "正向流程",
                "前置条件": "1. 使用内网测试账号登录。",
                "测试数据": "http://192.168.1.20；token=abc123",
                "操作步骤": "1. 登录系统，进入一级模块-二级页面。\n2. 输入内网地址并点击连接按钮。",
                "预期结果": "1. 页面保留输入的内网地址。\n2. 连接结果显示成功。",
                "执行状态": "待执行",
                "是否适合自动化": "否",
            },
            "性能测试设计": {
                "性能场景 ID": "PF-001",
                "业务链路": "打开页面并建立连接",
                "性能测试类型": "响应时间",
                "DFX维度": "DFP性能",
                "DFX场景": "响应时间",
                "目标用户量/并发数": "建议目标：1个用户单次操作",
                "响应时间目标": "建议目标：以实测基线为准",
                "数据量级": "建议目标：1条有效测试数据",
                "测试时长": "建议目标：完成3次重复采样",
                "监控指标": "页面响应时间",
                "通过标准": "不显著劣于实测基线",
                "造数策略": "使用本轮创建的有效测试数据",
                "风险说明": "无正式阈值，交付前待确认",
                "是否纳入本轮测试": "是",
            },
            "页面元素覆盖清单": {
                "元素 ID": "EL-001",
                "页面/入口": "二级页面",
                "页面 URL/菜单路径": "一级模块>二级页面",
                "元素名称/文案": "连接按钮",
                "元素类型": "按钮",
                "交互方式": "点击",
                "适用DFX维度": "DFT功能",
                "适用DFX场景": "正向流程",
                "预期行为": "执行连接并展示结果",
                "覆盖用例 ID": "TC-001",
                "覆盖状态": "已覆盖",
                "发现方式": "已实测",
                "素材来源": "页面实探",
            },
        }
        for sheet_name, values in rows.items():
            worksheet = workbook[sheet_name]
            excel_tools.write_mapped_row(worksheet, excel_tools.header_map(worksheet), 2, values)
        workbook.save(output)

    def test_prepare_formal_clears_all_template_examples(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "formal.xlsx"
            excel_tools.prepare_formal_workbook(FORMAL_TEMPLATE, output)
            workbook = load_workbook(output)
            self.assertEqual(excel_tools.FORMAL_SHEETS, workbook.sheetnames)
            for sheet_name in excel_tools.FORMAL_SHEETS:
                worksheet = workbook[sheet_name]
                for row in worksheet.iter_rows(min_row=2):
                    self.assertTrue(all(cell.value is None for cell in row), sheet_name)
            for sheet_name, fields in excel_tools.FORMAL_MULTILINE_FIELDS.items():
                headers = excel_tools.header_map(workbook[sheet_name])
                self.assertTrue(set(fields) <= set(headers), f"{sheet_name}: {set(fields) - set(headers)}")

    def test_two_level_path_keeps_canonical_name_and_fills_import_third_level(self) -> None:
        self.assertEqual(
            excel_tools.deliverable_names("一级模块>二级页面")[1:],
            ("一级模块_二级页面_测试设计.xlsx", "一级模块_二级页面_导入用例.xlsx"),
        )
        self.assertEqual(excel_tools.import_module_names("一级模块>二级页面")[:3], ["一级模块", "二级页面", "二级页面"])

    def test_run_directory_is_lightweight_unless_scope_is_large(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project_root = Path(temp_dir)
            template_source = REPO_ROOT / "docs" / "test-assets" / "batch-runs" / "templates"
            template_target = project_root / "docs" / "test-assets" / "batch-runs" / "templates"
            template_target.mkdir(parents=True)
            for name in ["batch-status-template.csv", "page-discovery-template.csv", "batch-plan-template.md"]:
                (template_target / name).write_bytes((template_source / name).read_bytes())

            simple = excel_tools.init_batch_run(project_root, "simple", "一级模块>二级页面", "BATCH-001")
            self.assertTrue((simple / "batch-status.csv").exists())
            self.assertTrue((simple / "page-discovery.csv").exists())
            self.assertFalse((simple / "batch-plan.md").exists())
            self.assertFalse((simple / "batch-review.md").exists())
            with (simple / "page-discovery.csv").open("r", encoding="utf-8-sig", newline="") as fp:
                row = next(csv.DictReader(fp))
            self.assertEqual(row["事实状态"], "待确认")
            original_discovery = (simple / "page-discovery.csv").read_text(encoding="utf-8-sig")
            excel_tools.init_batch_run(project_root, "simple", "一级模块>二级页面", "BATCH-001")
            self.assertEqual((simple / "page-discovery.csv").read_text(encoding="utf-8-sig"), original_discovery)

            large = excel_tools.init_batch_run(
                project_root,
                "large",
                "一级模块>二级页面",
                "BATCH-001",
                large_scope=True,
            )
            self.assertTrue((large / "batch-plan.md").exists())

    def test_batch_status_is_checkpoint_not_quality_gate(self) -> None:
        validator = load_deliverable_validator()
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "batch-status.csv"
            path.write_text(
                "批次ID,一级模块,二级菜单,三级菜单/页面域,最小标题路径,状态,页面实探状态,JSON分片状态,功能用例数,性能场景数,归档路径,导入文件路径,最后更新时间,下一步动作\n"
                "BATCH-001,一级模块,二级页面,,一级模块>二级页面,已完成,已完成,已完成,3,0,,,,完成\n",
                encoding="utf-8-sig",
            )
            rows = validator.validate_batch_status(path)
            self.assertEqual(rows[0]["性能场景数"], "0")

    def test_case_shard_accepts_combined_business_navigation_step(self) -> None:
        validator = load_shard_validator()
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "cases.json"
            path.write_text(
                '{"功能测试用例":{"rows":[{'
                '"用例 ID":"TC-001","功能点":"任务执行",'
                '"用例标题":"任务执行-使用有效输入",'
                '"操作步骤":"1. 登录系统，进入一级功能-目标页面。\\n2. 输入有效数据并执行任务。",'
                '"预期结果":"1. 成功进入目标页面。\\n2. 页面显示本次任务的具体处理结果。"'
                '}]}}',
                encoding="utf-8",
            )
            rows = validator.validate_json(path)
            self.assertEqual(len(rows), 1)

    def test_current_producer_rejects_temporary_repair_script(self) -> None:
        validator = load_shard_validator()
        with self.assertRaises(AssertionError):
            validator.validate_compile(Path("fix_retry.py"))

    def test_historical_discovery_migration_does_not_claim_unverified_work(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "page-discovery.csv"
            path.write_text(
                "页面/入口,元素名称/文案,是否已生成用例,覆盖状态\n"
                "目标页面,目标按钮,是,已覆盖\n",
                encoding="utf-8-sig",
            )
            excel_tools.migrate_page_discovery_fact_status(path)
            with path.open("r", encoding="utf-8-sig", newline="") as fp:
                row = next(csv.DictReader(fp))
            self.assertEqual(row["事实状态"], "页面观察")

    def test_import_preserves_internal_test_values_without_masking(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            formal = temp / "formal.xlsx"
            imported = temp / "import.xlsx"
            excel_tools.prepare_formal_workbook(FORMAL_TEMPLATE, formal)
            workbook = load_workbook(formal)
            worksheet = workbook[excel_tools.FORMAL_FUNCTION_SHEET]
            headers = excel_tools.header_map(worksheet)
            values = {
                "用例 ID": "TC-001",
                "模块": "一级模块-二级页面",
                "功能点": "连接测试",
                "用例标题": "连接测试-使用内网地址执行",
                "优先级": "P1",
                "测试类型": "功能测试",
                "DFX维度": "DFT功能",
                "DFX场景": "正向流程",
                "前置条件": "1. 使用http://192.168.1.20，password=internal-test，token=abc123。",
                "测试数据": "URL：http://192.168.1.20；password=internal-test；token=abc123",
                "操作步骤": "1. 登录系统，进入一级模块-二级页面。\n2. 输入内网地址并执行连接测试。",
                "预期结果": "1. 页面显示当前内网地址。\n2. 连接结果显示成功。",
            }
            excel_tools.write_mapped_row(worksheet, headers, 2, values)
            workbook.save(formal)
            excel_tools.generate_import_workbook(formal, IMPORT_TEMPLATE, imported, "一级模块>二级页面")
            import_workbook = load_workbook(imported, data_only=True)
            import_sheet = import_workbook[import_workbook.sheetnames[0]]
            import_headers = excel_tools.header_map(import_sheet)
            self.assertEqual(import_sheet.cell(2, import_headers["三级模块名称"]).value, "二级页面")
            self.assertIn("192.168.1.20", import_sheet.cell(2, import_headers["前置条件"]).value)

    def test_business_closure_does_not_require_generic_recovery(self) -> None:
        validator = load_deliverable_validator()
        validator.assert_complete_operation_steps(
            "1. 登录系统，进入一级功能-目标页面。\n2. 点击刷新并观察列表更新时间。",
            "combined business navigation",
        )
        validator.assert_business_transaction_closed(
            "1. 登录系统，进入模块-页面。\n2. 选择选项B并观察联动字段。",
            "1. 下拉框显示选项B。\n2. 联动字段按选项B展示。",
            "selection",
        )
        validator.assert_business_transaction_closed(
            "1. 登录系统，进入模块-页面。\n2. 点击创建并提交。",
            "1. 创建成功并进入详情页面。",
            "create",
        )
        with self.assertRaises(AssertionError):
            validator.assert_business_transaction_closed(
                "1. 登录系统，进入模块-页面。\n2. 点击创建并提交。",
                "1. 页面正常。",
                "create",
            )

    def test_atomic_publish_rolls_back_all_targets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp = Path(temp_dir)
            source_one = temp / "source-one.xlsx"
            source_two = temp / "source-two.xlsx"
            target_one = temp / "target-one.xlsx"
            target_two = temp / "target-two.xlsx"
            source_one.write_bytes(b"new-one")
            source_two.write_bytes(b"new-two")
            target_one.write_bytes(b"old-one")
            target_two.write_bytes(b"old-two")

            original_replace = os.replace
            call_count = 0

            def fail_during_second_publish(source, target):
                nonlocal call_count
                call_count += 1
                if call_count == 4:
                    raise OSError("simulated publish failure")
                return original_replace(source, target)

            with patch.object(excel_tools.os, "replace", side_effect=fail_during_second_publish):
                with self.assertRaises(OSError):
                    excel_tools.atomic_publish_copies([(source_one, target_one), (source_two, target_two)])

            self.assertEqual(target_one.read_bytes(), b"old-one")
            self.assertEqual(target_two.read_bytes(), b"old-two")

    def test_complete_delivery_publishes_one_canonical_pair(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            project_root = Path(temp_dir)
            formal = project_root / "draft.xlsx"
            self.build_minimal_valid_formal(formal)
            legacy_dir = project_root / "docs" / "test-design" / "deliverables"
            legacy_dir.mkdir(parents=True, exist_ok=True)
            legacy_formal = legacy_dir / "一级模块_二级页面_二级页面_测试设计.xlsx"
            legacy_import = legacy_dir / "一级模块_二级页面_二级页面_导入用例.xlsx"
            legacy_formal.write_bytes(b"stale")
            legacy_import.write_bytes(b"stale")

            paths = excel_tools.complete_deliverables(
                project_root=project_root,
                formal_workbook=formal,
                import_template=IMPORT_TEMPLATE,
                module_path="一级模块>二级页面",
            )

            self.assertEqual(paths["formal"].name, "一级模块_二级页面_测试设计.xlsx")
            self.assertEqual(paths["import"].name, "一级模块_二级页面_导入用例.xlsx")
            self.assertEqual(paths["formal"].parent, paths["import"].parent)
            self.assertEqual(paths["formal"].parent.name, "deliverables")
            self.assertTrue(paths["formal"].exists())
            self.assertTrue(paths["import"].exists())
            self.assertFalse(legacy_formal.exists())
            self.assertFalse(legacy_import.exists())


if __name__ == "__main__":
    unittest.main()
